# -*- coding: utf-8 -*-
"""
Web Admin – Cấp hồ sơ nhân viên (Flask).
API: GET/POST /api/employees, DELETE /api/employees/:id
Đăng nhập: POST /api/login (username, password) -> JWT. Các API (trừ login, config) cần header Authorization: Bearer <token>.
Hỗ trợ bulk-import nhân viên: POST /api/employees/bulk-import (file Excel).
"""
import os
import csv
import sqlite3
import smtplib
import ssl
import random
import secrets
import base64
import urllib.error
import urllib.parse
import urllib.request
from email.message import EmailMessage
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, g, send_file, make_response
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from database import DB_PATH, get_db, init_db
from config import (
    SECRET_KEY, MAX_CONTENT_LENGTH, JWT_ALGORITHM, JWT_EXPIRY_HOURS,
    get_upload_folder, get_chat_upload_folder, get_arcface_model_path,
    FACE_RESET_MIN_SIMILARITY, PASSWORD_RESET_OTP_MINUTES,
    COMPANY_NAME, ATTENDANCE_RADIUS, LATE_THRESHOLD, LANGUAGE,
    DEFAULT_PORT, GOOGLE_PLACES_API_KEY,
)

try:
    import jwt
    if not hasattr(jwt, 'encode'):
        raise ImportError("Cần gói PyJWT. Chạy: pip uninstall jwt python-jwt ; pip install PyJWT")
except ImportError as e:
    raise ImportError("Cài PyJWT: pip install PyJWT (không phải 'jwt' hay 'python-jwt')") from e

import json
import io
import numpy as np
from PIL import Image


def _get_request_data():
    """Parse request body — hỗ trợ cả JSON và FormData."""
    if request.is_json:
        return request.get_json(silent=True) or {}
    # FormData: values from request.form are lists → unwrap single values
    return {k: v[0] if isinstance(v, (list, tuple)) and len(v) == 1 else v
            for k, v in dict(request.form).items()}


def load_local_env():
    """Đọc file web_admin/.env (dòng KEY=VALUE). Không ghi đè biến môi trường đã có (ưu tiên hệ thống/Docker)."""
    base = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(base, '.env')
    if not os.path.isfile(path):
        return
    try:
        with open(path, encoding='utf-8-sig') as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith('#'):
                    continue
                if line.startswith('export '):
                    line = line[7:].strip()
                if '=' not in line:
                    continue
                key, _, val = line.partition('=')
                key = key.strip()
                if not key:
                    continue
                val = val.strip()
                if len(val) >= 2 and val[0] == val[-1] and val[0] in '"\'':
                    val = val[1:-1]
                if key not in os.environ:
                    os.environ[key] = val
        print(f"[INFO] Đã nạp cấu hình từ {path}")
    except OSError as e:
        print(f"[WARN] Không đọc được .env: {e}")


load_local_env()

# Model ArcFace để tính embedding (512-d, input 112x112)
ARCFACE_MODEL_PATH = get_arcface_model_path()
_arcface_model = None

def get_arcface_model():
    global _arcface_model
    if _arcface_model is None:
        try:
            from tensorflow.keras.models import load_model
            _arcface_model = load_model(ARCFACE_MODEL_PATH, compile=False, safe_mode=False)
        except Exception as e:
            print(f"[WARN] Không load được ArcFace model: {e}")
            _arcface_model = False
    return _arcface_model if _arcface_model else None

def compute_embedding(image_path):
    """Tính embedding 512-d từ ảnh (112x112). Trả về list hoặc None nếu lỗi."""
    model = get_arcface_model()
    if model is None:
        return None
    try:
        img = Image.open(image_path).convert('RGB').resize((112, 112))
        arr = np.array(img, dtype=np.float32) / 255.0
        arr = np.expand_dims(arr, axis=0)  # (1, 112, 112, 3)
        emb = model.predict(arr, verbose=0)[0]
        # L2 normalize
        norm = np.linalg.norm(emb)
        if norm > 0:
            emb = emb / norm
        return emb.tolist()
    except Exception as e:
        print(f"[WARN] compute_embedding lỗi: {e}")
        return None


def compute_embedding_from_images(image_paths, max_images=5):
    """Tính embedding trung bình từ nhiều ảnh (tối đa 5). Mỗi ảnh L2-normalize, trung bình rồi L2-normalize lại. Trả về list 512-d hoặc None."""
    if not image_paths:
        return None
    paths = image_paths[:max_images]
    embeddings = []
    for p in paths:
        emb = compute_embedding(p)
        if emb is not None and len(emb) == 512:
            embeddings.append(np.array(emb, dtype=np.float32))
    if not embeddings:
        return None
    # Trung bình các vector đã L2-normalize
    mean_emb = np.mean(embeddings, axis=0)
    norm = np.linalg.norm(mean_emb)
    if norm > 0:
        mean_emb = mean_emb / norm
    return mean_emb.tolist()


# ─── Flask App ─────────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder='static', static_url_path='')
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH
app.config['SECRET_KEY'] = SECRET_KEY

UPLOAD_FOLDER = get_upload_folder()
CHAT_UPLOAD_FOLDER = get_chat_upload_folder()
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CHAT_UPLOAD_FOLDER, exist_ok=True)


# ─── Rate Limiting (in-memory, single-instance) ───────────────────────────────
from threading import Lock
from collections import defaultdict
import time as _time

_RATE_LOCK = Lock()
_login_attempts = defaultdict(list)     # ip -> [(timestamp, username), ...]
_otp_attempts = defaultdict(list)       # ip -> [(timestamp), ...]
_api_attempts = defaultdict(list)      # ip -> [(timestamp, path), ...]

_RATE_LIMITS = {
    'login': {'max': 10, 'window_sec': 300},      # 10 lần / 5 phút
    'otp':   {'max': 5,  'window_sec': 900},       # 5 lần / 15 phút
    'api':   {'max': 200,'window_sec': 60},        # 200 request / phút
}

def _cleanup_rate(bucket: dict, window_sec: int, max_age: float):
    """Xóa entries cũ hơn window_sec."""
    now = _time.time()
    expired = [k for k, v in bucket.items() if not v or (now - max(v)[0]) > window_sec]
    for k in expired:
        bucket.pop(k, None)

def check_rate_limit(bucket_name: str, key: str, username: str = None, path: str = None) -> tuple:
    """Kiểm tra rate limit. Trả về (allowed, remaining, retry_after_sec)."""
    cfg = _RATE_LIMITS.get(bucket_name, _RATE_LIMITS['api'])
    max_req, window = cfg['max'], cfg['window_sec']
    bucket = {'login': _login_attempts, 'otp': _otp_attempts, 'api': _api_attempts}.get(bucket_name, _api_attempts)
    now = _time.time()

    with _RATE_LOCK:
        _cleanup_rate(bucket, window, now - window)
        entries = bucket[key]
        # Lọc entries trong window
        if username:
            valid = [(t, u) for t, u in entries if now - t <= window and u == username]
        elif path:
            valid = [(t, p) for t, p in entries if now - t <= window and p == path]
        else:
            valid = [(t, u) for t, u in entries if now - t <= window]

        if len(valid) >= max_req:
            oldest = valid[0][0]
            retry_after = int(oldest + window - now) + 1
            return False, 0, retry_after

        if username:
            valid.append((now, username))
        elif path:
            valid.append((now, path))
        else:
            valid.append((now, key))
        bucket[key] = valid
        return True, max(0, max_req - len(valid)), 0

def clear_rate_limit(bucket_name: str, key: str):
    """Xóa rate limit entry (dùng khi đăng nhập thành công)."""
    bucket = {'login': _login_attempts, 'otp': _otp_attempts, 'api': _api_attempts}.get(bucket_name)
    if bucket:
        with _RATE_LOCK:
            bucket.pop(key, None)


def send_email(to_email, subject, body):
    """Gửi email qua SMTP. Trả (True, None) nếu thành công; (False, thông báo lý do) nếu lỗi."""
    host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    port = int(os.environ.get('SMTP_PORT', '587'))
    user = os.environ.get('SMTP_USER') or ''
    password = os.environ.get('SMTP_PASSWORD') or ''
    sender = os.environ.get('SMTP_FROM') or user
    use_ssl = (os.environ.get('SMTP_USE_SSL') or '').strip().lower() in ('1', 'true', 'yes', 'on')
    if port == 465:
        use_ssl = True

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = sender or 'no-reply@example.com'
    msg['To'] = to_email
    msg.set_content(body)

    if not (user and password and sender and to_email):
        print(f"[INFO] send_email skip SMTP (thiếu cấu hình). To: {to_email}, Subject: {subject}")
        return False, (
            'Chưa cấu hình SMTP đầy đủ. Trong web_admin/.env cần: SMTP_USER, SMTP_PASSWORD, SMTP_FROM '
            '(và tùy chọn SMTP_HOST, SMTP_PORT).'
        )
    try:
        context = ssl.create_default_context()
        if use_ssl:
            with smtplib.SMTP_SSL(host, port, context=context) as server:
                server.login(user, password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(host, port) as server:
                server.starttls(context=context)
                server.login(user, password)
                server.send_message(msg)
        return True, None
    except Exception as e:
        print(f"[WARN] send_email SMTP failed: {e}\nTo: {to_email}")
        return False, f'Lỗi SMTP: {e}'


def validate_password(pw: str):
    """Kiểm tra mật khẩu đủ mạnh: >= 6 ký tự, có số và chữ cái."""
    if len(pw) < 6:
        return False, 'Mật khẩu phải có ít nhất 6 ký tự.'
    if not any(c.isdigit() for c in pw):
        return False, 'Mật khẩu phải chứa ít nhất 1 chữ số.'
    if not any(c.isalpha() for c in pw):
        return False, 'Mật khẩu phải chứa ít nhất 1 chữ cái.'
    return True, None


def generate_random_password():
    """Mật khẩu ngẫu nhiên đạt validate_password (8 ký tự: chữ + số)."""
    letters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
    digits = '0123456789'
    rng = secrets.SystemRandom()
    while True:
        parts = [secrets.choice(letters) for _ in range(4)] + [secrets.choice(digits) for _ in range(4)]
        rng.shuffle(parts)
        pw = ''.join(parts)
        ok, _ = validate_password(pw)
        if ok:
            return pw


def send_sms_twilio(to_phone: str, body: str):
    """Gửi SMS qua Twilio. Trả (True, None) nếu thành công; (False, thông báo lý do) nếu lỗi."""
    sid = (os.environ.get('TWILIO_ACCOUNT_SID') or '').strip()
    tok = (os.environ.get('TWILIO_AUTH_TOKEN') or '').strip()
    from_n = (os.environ.get('TWILIO_PHONE_NUMBER') or '').strip()
    if not (sid and tok and from_n and to_phone):
        print(f"[INFO] SMS skip Twilio (thiếu cấu hình hoặc SĐT). Đến: {to_phone}")
        return False, (
            'Chưa cấu hình Twilio. Trong web_admin/.env cần: TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, '
            'TWILIO_PHONE_NUMBER (số gửi SMS, dạng +84...).'
        )
    try:
        url = f'https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json'
        data = urllib.parse.urlencode({'To': to_phone, 'From': from_n, 'Body': body}).encode()
        req = urllib.request.Request(url, data=data, method='POST')
        req.add_header('Authorization', 'Basic ' + base64.b64encode(f'{sid}:{tok}'.encode()).decode())
        with urllib.request.urlopen(req, timeout=30):
            pass
        return True, None
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode('utf-8', errors='replace')[:500]
        except Exception:
            body = str(e)
        print(f"[WARN] Twilio SMS HTTP {e.code}: {body}")
        return False, f'Twilio trả lỗi HTTP {e.code}: {body}'
    except Exception as e:
        print(f"[WARN] Twilio SMS: {e}")
        return False, f'Lỗi gửi SMS: {e}'


def normalize_phone_digits(p: str) -> str:
    if not p:
        return ''
    return ''.join(c for c in p if c.isdigit())


def employee_face_similarity(stored_embedding_json, probe_embedding_list):
    """Trả về độ tương đồng [0..1] giữa embedding lưu DB và embedding ảnh mới."""
    if not stored_embedding_json or not probe_embedding_list:
        return 0.0
    try:
        a = np.array(json.loads(stored_embedding_json), dtype=np.float32)
        b = np.array(probe_embedding_list, dtype=np.float32)
        na, nb = np.linalg.norm(a), np.linalg.norm(b)
        if na <= 0 or nb <= 0:
            return 0.0
        a, b = a / na, b / nb
        return float(np.dot(a, b))
    except Exception:
        return 0.0


def _password_reset_row(conn, username: str, otp: str):
    return conn.execute(
        '''SELECT id, otp, expires_at, used FROM password_resets
           WHERE username = ? AND otp = ? AND used = 0
           ORDER BY id DESC LIMIT 1''',
        (username, otp),
    ).fetchone()


def _age_from_birth_date(bd):
    if not bd or not str(bd).strip():
        return None
    try:
        raw = str(bd).strip()[:10]
        born = datetime.strptime(raw, '%Y-%m-%d').date()
        today = datetime.utcnow().date()
        age = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
        return max(0, age)
    except Exception:
        return None


def _sender_display_name(conn, sender_username, sender_type):
    """Tên hiển thị khi chat: ưu tiên họ tên nhân viên từ users→employees, admin không gắn NV → Quản trị viên."""
    if not sender_username:
        return '—'
    row = conn.execute(
        '''SELECT e.name FROM users u
           LEFT JOIN employees e ON u.employee_id = e.id
           WHERE u.username = ?''',
        (sender_username,),
    ).fetchone()
    if row and row['name']:
        return row['name']
    if (sender_type or '').lower() == 'admin':
        return 'Quản trị viên'
    return sender_username


# CORS whitelist - chỉ những origin này mới được phép (thêm qua env ALLOWED_ORIGINS)
_ALLOWED_ORIGINS = set()
for o in (os.environ.get('ALLOWED_ORIGINS') or '').split(','):
    o = o.strip().rstrip('/')
    if o:
        _ALLOWED_ORIGINS.add(o)


@app.before_request
def handle_cors_preflight():
    """Xử lý preflight OPTIONS request trước khi route chạy."""
    import sys
    print(f'[PREFLIGHT] {request.method} {request.path} | Content-Type: {request.headers.get("Content-Type")} | Origin: {request.headers.get("Origin")}', file=sys.stderr)
    if request.method == 'OPTIONS':
        origin = request.headers.get('Origin', '')
        allowed = True
        if _ALLOWED_ORIGINS and origin not in _ALLOWED_ORIGINS:
            allowed = False
        if allowed:
            resp = make_response('', 204)
            resp.headers['Access-Control-Allow-Origin'] = origin or '*'
            resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, PATCH, DELETE, OPTIONS'
            resp.headers['Access-Control-Allow-Headers'] = 'Authorization, Content-Type'
            resp.headers['Access-Control-Max-Age'] = '3600'
            print(f'[PREFLIGHT] -> 204 OK', file=sys.stderr)
            return resp
        print(f'[PREFLIGHT] -> 403 Forbidden', file=sys.stderr)
        return '', 403


@app.after_request
def add_cors_headers(resp):
    if request.method == 'OPTIONS':
        return resp  # before_request đã xử lý
    origin = request.headers.get('Origin', '')
    if _ALLOWED_ORIGINS:
        if origin in _ALLOWED_ORIGINS:
            resp.headers['Access-Control-Allow-Origin'] = origin
            resp.headers['Vary'] = 'Origin'
        else:
            pass
    else:
        if origin:
            resp.headers['Access-Control-Allow-Origin'] = origin
            resp.headers['Vary'] = 'Origin'
        else:
            resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Access-Control-Allow-Headers'] = 'Authorization, Content-Type'
    resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, PATCH, DELETE, OPTIONS'
    # Đảm bảo UTF-8 charset cho static files (JS, CSS) tránh lỗi hiển thị tiếng Việt
    path = request.path.lower()
    if path.endswith('.js'):
        resp.headers['Content-Type'] = 'application/javascript; charset=utf-8'
    elif path.endswith('.css'):
        resp.headers['Content-Type'] = 'text/css; charset=utf-8'
    return resp


@app.errorhandler(415)
def handle_unsupported_media(e):
    """Debug lỗi 415 - cho biết body không parse được."""
    import sys, traceback
    traceback.print_exc()
    print(f'[DEBUG 415] Content-Type: {request.headers.get("Content-Type")}', file=sys.stderr)
    print(f'[DEBUG 415] Data: {request.get_data()}', file=sys.stderr)
    return jsonify({
        'error': 'Unsupported Media Type',
        'detail': str(e),
        'content_type': request.headers.get('Content-Type'),
        'body_preview': request.get_data()[:200],
    }), 415


def get_token_payload():
    """Lấy payload từ header Authorization: Bearer <token>. Trả về None nếu không hợp lệ."""
    auth = request.headers.get('Authorization')
    if not auth or not auth.startswith('Bearer '):
        return None
    token = auth[7:].strip()
    if not token:
        return None
    try:
        return jwt.decode(token, app.config['SECRET_KEY'], algorithms=[JWT_ALGORITHM])
    except jwt.InvalidTokenError:
        return None


def require_auth(f):
    """Decorator: yêu cầu JWT hợp lệ. Trả về 401 nếu chưa đăng nhập."""
    @wraps(f)
    def wrapped(*args, **kwargs):
        payload = get_token_payload()
        if not payload or not payload.get('sub'):
            return jsonify({'error': 'Chưa đăng nhập hoặc phiên hết hạn'}), 401
        g.current_user = payload.get('sub')
        return f(*args, **kwargs)
    return wrapped


def get_user_employee_id(conn, username):
    row = conn.execute(
        'SELECT employee_id FROM users WHERE username = ?',
        (username,)
    ).fetchone()
    if not row or row['employee_id'] is None:
        return None
    return int(row['employee_id'])


@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')


@app.route('/dashboard')
@app.route('/employees')
@app.route('/positions')
@app.route('/shifts')
@app.route('/departments')
@app.route('/attendance')
@app.route('/attendance-log')
@app.route('/leave-requests')
@app.route('/payroll')
@app.route('/chat')
@app.route('/account')
@app.route('/settings')
def index_views():
    """Các route SPA riêng cho từng chức năng."""
    return send_from_directory(app.static_folder, 'index.html')


@app.route('/login')
def login_page():
    return send_from_directory(app.static_folder, 'login.html')


@app.route('/api/login', methods=['POST'])
def login():
    """Dăng nhập: body JSON { username, password }. Trả về { token, username }."""
    conn = None
    try:
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip() or 'unknown'

        allowed, remaining, retry_after = check_rate_limit('login', client_ip, username='')
        if not allowed:
            return jsonify({
                'error': 'Quá nhiều lần đăng nhập. Vui lòng thử lại sau.',
                'retry_after': retry_after,
            }), 429

        data = _get_request_data()
        username = (data.get('username') or '').strip()
        password = data.get('password') or ''
        if not username or not password:
            return jsonify({'error': 'Thiếu tên đăng nhập hoặc mật khẩu'}), 400

        conn = get_db()

        # Kiểm tra bảng admin_users trước
        cur = conn.execute(
            'SELECT id, password_hash, display_name, role FROM admin_users WHERE username = ? AND is_active = 1',
            (username,)
        )
        row = cur.fetchone()

        # Nếu không có trong admin_users, kiểm tra bảng users cũ (backward compat)
        if not row:
            cur = conn.execute('SELECT id, password_hash FROM users WHERE username = ?', (username,))
            row = cur.fetchone()

        conn.close()

        if not row or not check_password_hash(row['password_hash'], password):
            return jsonify({'error': 'Sai tên đăng nhập hoặc mật khẩu'}), 401

        # Xóa rate limit khi đăng nhập thành công
        clear_rate_limit('login', client_ip)

        payload = {'sub': username, 'exp': datetime.utcnow() + timedelta(hours=JWT_EXPIRY_HOURS)}
        token = jwt.encode(payload, app.config['SECRET_KEY'], algorithm=JWT_ALGORITHM)
        if hasattr(token, 'decode'):
            token = token.decode('utf-8')
        return jsonify({'token': token, 'username': username})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    """Gửi OTP quên mật khẩu qua email hoặc SMS (theo hồ sơ nhân viên / users).

    Body JSON:
      - username (bắt buộc)
      - channel: \"email\" | \"sms\" (mặc định email)
      - email (tùy chọn): nếu gửi thì phải khớp email đã lưu (không gửi = dùng email trên hồ sơ users/employees)
      - phone (tùy chọn): nếu gửi thì phải khớp SĐT đã lưu (không gửi = dùng SĐT employee hoặc users.phone)
    """
    conn = None
    try:
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip() or 'unknown'

        allowed, remaining, retry_after = check_rate_limit('otp', client_ip)
        if not allowed:
            return jsonify({
                'error': 'Quá nhiều yêu cầu OTP. Vui lòng thử lại sau.',
                'retry_after': retry_after,
            }), 429

        data = _get_request_data()
        username = (data.get('username') or '').strip()
        channel = (data.get('channel') or 'email').strip().lower()
        email_hint = (data.get('email') or '').strip()
        phone_hint = (data.get('phone') or '').strip()

        if not username:
            return jsonify({'error': 'Thiếu tên đăng nhập'}), 400

        # Form cũ chỉ gửi username + email → coi là kênh email và kiểm tra khớp
        if email_hint and 'channel' not in data:
            channel = 'email'

        if channel not in ('email', 'sms'):
            return jsonify({'error': 'channel phải là email hoặc sms'}), 400

        conn = None
        conn = get_db()
        row = conn.execute(
            '''SELECT u.username, u.email AS u_email, u.phone AS u_phone, u.employee_id,
                      e.email AS e_email, e.phone AS e_phone
               FROM users u
               LEFT JOIN employees e ON u.employee_id = e.id
               WHERE u.username = ?''',
            (username,),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy tài khoản với tên đăng nhập này.'}), 404

        email_to = (row['e_email'] or row['u_email'] or '').strip()
        phone_to = (row['e_phone'] or row['u_phone'] or '').strip()

        if channel == 'email':
            if not email_to:
                conn.close()
                return jsonify({
                    'error': 'Tài khoản chưa có email trên hồ sơ. Vui lòng cập nhật tại quản trị, hoặc dùng SMS / xác thực khuôn mặt.',
                }), 400
            if email_hint and email_hint.lower() != email_to.lower():
                conn.close()
                return jsonify({'error': 'Email không khớp với hồ sơ nhân viên.'}), 400
            dest_label = 'email'
            send_dest = email_to
        else:
            if not phone_to:
                conn.close()
                return jsonify({
                    'error': 'Tài khoản chưa có số điện thoại trên hồ sơ. Cập nhật tại quản trị, hoặc dùng email / khuôn mặt.',
                }), 400
            if phone_hint and normalize_phone_digits(phone_hint) != normalize_phone_digits(phone_to):
                conn.close()
                return jsonify({'error': 'Số điện thoại không khớp với hồ sơ.'}), 400
            dest_label = 'sms'
            send_dest = phone_to

        otp = f"{random.randint(0, 999999):06d}"
        now = datetime.utcnow()
        expires_at = now + timedelta(minutes=PASSWORD_RESET_OTP_MINUTES)
        _m = PASSWORD_RESET_OTP_MINUTES

        if dest_label == 'email':
            subject = 'Mã OTP đặt lại mật khẩu'
            body = (
                f'Xin chào {username},\n\n'
                f'Mã OTP để đặt lại mật khẩu của bạn là: {otp}\n'
                f'Mã có hiệu lực trong {_m} phút.\n\n'
                f'Nếu bạn không yêu cầu đặt lại mật khẩu, vui lòng bỏ qua email này.'
            )
            ok_send, send_err = send_email(send_dest, subject, body)
            if not ok_send:
                conn.close()
                return jsonify({
                    'error': 'Không gửi được email.',
                    'detail': send_err or 'Kiểm tra SMTP trong web_admin/.env và log máy chủ.',
                }), 502
        else:
            sms_body = f'Ma OTP dat lai mat khau ({username}): {otp}. Hieu luc {_m} phut.'
            ok_send, send_err = send_sms_twilio(send_dest, sms_body)
            if not ok_send:
                conn.close()
                return jsonify({
                    'error': 'Không gửi được SMS.',
                    'detail': send_err or 'Kiểm tra Twilio trong web_admin/.env và log máy chủ.',
                }), 502

        # Chỉ INSERT OTP SAU KHI GỬI THÀNH CÔNG - ngăn brute force khi gửi thất bại
        conn.execute('UPDATE password_resets SET used = 1 WHERE username = ? AND used = 0', (username,))
        conn.execute(
            'INSERT INTO password_resets (username, otp, expires_at, used, created_at, channel) VALUES (?, ?, ?, 0, ?, ?)',
            (username, otp, expires_at.isoformat() + 'Z', now.isoformat() + 'Z', channel),
        )
        conn.commit()
        conn.close()
        print(f"[INFO] OTP {channel} đã gửi cho user {username}")
        return jsonify({
            'ok': True,
            'delivered': True,
            'message': (
                'Đã gửi mã OTP tới email đã lưu. Kiểm tra hộp thư (và thư mục spam).'
                if dest_label == 'email'
                else 'Đã gửi mã OTP tới số điện thoại đã lưu.'
            ),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forgot-password-face', methods=['POST'])
def forgot_password_face():
    """Xác thực khuôn mặt (ảnh) khớp embedding nhân viên → trả reset_token dùng thay OTP khi đặt lại mật khẩu.

    multipart/form-data: username, photo (file)
    """
    conn = None
    try:
        username = (request.form.get('username') or '').strip()
        if not username:
            return jsonify({'error': 'Thiếu tên đăng nhập'}), 400
        if 'photo' not in request.files:
            return jsonify({'error': 'Thiếu ảnh khuôn mặt'}), 400
        f = request.files['photo']
        if not f or not f.filename:
            return jsonify({'error': 'Chưa chọn ảnh'}), 400

        conn = get_db()
        row = conn.execute(
            '''SELECT u.username, u.employee_id, e.embedding
               FROM users u
               JOIN employees e ON u.employee_id = e.id
               WHERE u.username = ?''',
            (username,),
        ).fetchone()
        if not row or not row['embedding']:
            conn.close()
            return jsonify({'error': 'Không thể xác thực khuôn mặt cho tài khoản này (thiếu dữ liệu khuôn mặt).'}), 400

        ext = os.path.splitext(f.filename)[1] or '.jpg'
        tmp_path = os.path.join(UPLOAD_FOLDER, f"_reset_face_{int(datetime.now().timestamp())}_{secrets.token_hex(4)}{ext}")
        f.save(tmp_path)
        try:
            probe = compute_embedding(tmp_path)
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass

        if not probe:
            conn.close()
            return jsonify({'error': 'Không trích xuất được khuôn mặt từ ảnh. Chụp mặt rõ, đủ sáng.'}), 400

        sim = employee_face_similarity(row['embedding'], probe)
        if sim < FACE_RESET_MIN_SIMILARITY:
            conn.close()
            return jsonify({'error': 'Khuôn mặt không khớp với hồ sơ.'}), 400

        token = secrets.token_hex(24)
        now = datetime.utcnow()
        expires_at = now + timedelta(minutes=PASSWORD_RESET_OTP_MINUTES)
        conn.execute('UPDATE password_resets SET used = 1 WHERE username = ? AND used = 0', (username,))
        conn.execute(
            'INSERT INTO password_resets (username, otp, expires_at, used, created_at, channel) VALUES (?, ?, ?, 0, ?, ?)',
            (username, token, expires_at.isoformat() + 'Z', now.isoformat() + 'Z', 'face'),
        )
        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'reset_token': token,
            'message': 'Đã xác thực khuôn mặt. Dán mã này vào ô OTP / mã xác thực khi đặt lại mật khẩu.',
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    """Đặt lại mật khẩu bằng OTP 6 số (email/SMS) hoặc mã reset_token từ xác thực khuôn mặt.

    Body JSON: { username, otp | reset_token, new_password }
    """
    conn = None
    try:
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip() or 'unknown'
        allowed, remaining, retry_after = check_rate_limit('otp', client_ip)
        if not allowed:
            return jsonify({
                'error': 'Quá nhiều lần thử. Vui lòng thử lại sau.',
                'retry_after': retry_after,
            }), 429

        data = _get_request_data()
        username = (data.get('username') or '').strip()
        otp = (data.get('otp') or data.get('reset_token') or '').strip()
        new_password = data.get('new_password') or ''
        if not username or not otp or not new_password:
            return jsonify({'error': 'Thiếu tên đăng nhập, mã OTP (hoặc mã khuôn mặt) hoặc mật khẩu mới'}), 400

        ok, msg = validate_password(new_password)
        if not ok:
            return jsonify({'error': msg}), 400

        conn = get_db()
        row = _password_reset_row(conn, username, otp)
        if not row:
            conn.close()
            return jsonify({'error': 'Mã OTP không hợp lệ hoặc đã sử dụng'}), 400

        raw_exp = row['expires_at'].replace('Z', '').replace('+00:00', '')
        expires_at = datetime.fromisoformat(raw_exp)
        if datetime.utcnow() > expires_at:
            conn.execute('UPDATE password_resets SET used = 1 WHERE id = ?', (row['id'],))
            conn.commit()
            conn.close()
            return jsonify({'error': 'Mã OTP đã hết hạn'}), 400

        # Cập nhật mật khẩu - kiểm tra cả users và admin_users
        pw_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
        cur = conn.execute('UPDATE users SET password_hash = ? WHERE username = ?', (pw_hash, username))
        if cur.rowcount == 0:
            cur = conn.execute('UPDATE admin_users SET password_hash = ? WHERE username = ?', (pw_hash, username))
            if cur.rowcount == 0:
                conn.close()
                return jsonify({'error': 'Không tìm thấy tài khoản'}), 404

        conn.execute('UPDATE password_resets SET used = 1 WHERE id = ?', (row['id'],))
        conn.commit()
        conn.close()

        # Xóa rate limit khi đặt lại thành công
        clear_rate_limit('otp', client_ip)

        return jsonify({'ok': True, 'message': 'Đã đặt lại mật khẩu. Vui lòng đăng nhập lại.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/profile', methods=['GET'])
@require_auth
def admin_profile():
    """Trả về thông tin tài khoản đang đăng nhập + hồ sơ nhân viên (nếu có)."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute(
            '''SELECT u.id, u.username, u.email AS u_email, u.phone AS u_phone, u.employee_id, u.created_at,
                      u.display_name AS u_display_name, u.profile_birth_date, u.profile_department,
                      u.profile_position_id, u.profile_shift_id, u.profile_employee_code,
                      e.name AS employee_name, e.code AS employee_code, e.department, e.department_id,
                      COALESCE(ed.name, e.department) AS employee_department_display, e.phone AS e_phone,
                      e.email AS e_email, e.birth_date, e.photo_filename, e.position_id AS e_position_id, e.shift_id AS e_shift_id,
                      p.name AS position_name, s.name AS shift_name,
                      pa.name AS profile_position_name, sa.name AS profile_shift_name
               FROM users u
               LEFT JOIN employees e ON e.id = u.employee_id
               LEFT JOIN departments ed ON e.department_id = ed.id
               LEFT JOIN positions p ON e.position_id = p.id
               LEFT JOIN shifts s ON e.shift_id = s.id
               LEFT JOIN positions pa ON pa.id = u.profile_position_id
               LEFT JOIN shifts sa ON sa.id = u.profile_shift_id
               WHERE u.username = ?''',
            (g.current_user,)
        )
        row = cur.fetchone()
        conn.close()
        if not row:
            return jsonify({'error': 'Không tìm thấy tài khoản'}), 404
        base = request.url_root.rstrip('/')
        photo_fn = row['photo_filename']
        photo_url = f'{base}/uploads/{photo_fn}' if photo_fn else None
        emp_id = row['employee_id']
        is_pure_admin = emp_id is None
        emp_name = (row['employee_name'] or '').strip()
        u_disp = (row['u_display_name'] or '').strip()
        if is_pure_admin:
            display_name = u_disp if u_disp else (row['username'] or '')
            email_eff = (row['u_email'] or '').strip()
            phone_eff = (row['u_phone'] or '').strip()
            bd = row['profile_birth_date']
            dept = (row['profile_department'] or '').strip() or None
            pos_name = (row['profile_position_name'] or '').strip() or None
            shift_nm = (row['profile_shift_name'] or '').strip() or None
            emp_code = (row['profile_employee_code'] or '').strip() or None
            pos_id = row['profile_position_id']
            shift_id = row['profile_shift_id']
        else:
            display_name = emp_name if emp_name else (row['username'] or '')
            email_eff = (row['e_email'] or row['u_email'] or '').strip()
            phone_eff = (row['e_phone'] or row['u_phone'] or '').strip()
            bd = row['birth_date']
            dept = (row['employee_department_display'] or '').strip() or None
            pos_name = (row['position_name'] or '').strip() or None
            shift_nm = (row['shift_name'] or '').strip() or None
            emp_code = (row['employee_code'] or '').strip() or None
            pos_id = row['e_position_id']
            shift_id = row['e_shift_id']
        age = _age_from_birth_date(bd)
        return jsonify({
            'id': row['id'],
            'username': row['username'],
            'email': email_eff,
            'phone': phone_eff,
            'employee_id': emp_id,
            'is_admin': is_pure_admin,
            'created_at': row['created_at'],
            'photo_url': photo_url,
            'display_name': display_name,
            'employee_name': (emp_name or None) if not is_pure_admin else (u_disp or None),
            'employee_code': emp_code,
            'department': dept,
            'department_id': row['department_id'] if not is_pure_admin else None,
            'position_name': pos_name,
            'shift_name': shift_nm,
            'position_id': pos_id,
            'shift_id': shift_id,
            'birth_date': (str(bd).strip() if bd else None),
            'age': age,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _parse_optional_fk(val):
    if val is None or val == '':
        return None
    try:
        v = int(val)
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None


@app.route('/api/admin/profile', methods=['PUT'])
@require_auth
def admin_update_profile():
    conn = None
    try:
        data = _get_request_data()
        email = (data.get('email') or '').strip()
        phone = (data.get('phone') or '').strip()
        if email and '@' not in email:
            return jsonify({'error': 'Email không hợp lệ'}), 400

        conn = get_db()

        # Kiểm tra bảng admin_users trước
        urow = conn.execute(
            'SELECT id, role FROM admin_users WHERE username = ?',
            (g.current_user,),
        ).fetchone()

        if urow:
            # Cập nhật bảng admin_users
            display_name = (data.get('employee_name') or data.get('display_name') or '').strip()
            email = (data.get('email') or '').strip()
            phone = (data.get('phone') or '').strip()
            if email and '@' not in email:
                conn.close()
                return jsonify({'error': 'Email không hợp lệ'}), 400
            conn.execute(
                'UPDATE admin_users SET display_name = ?, email = ?, phone = ? WHERE username = ?',
                (display_name or None, email or None, phone or None, g.current_user),
            )
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Cập nhật hồ sơ thành công'})

        # Kiểm tra bảng users cũ
        urow = conn.execute(
            'SELECT id, employee_id FROM users WHERE username = ?',
            (g.current_user,),
        ).fetchone()
        if not urow:
            conn.close()
            return jsonify({'error': 'Không tìm thấy tài khoản'}), 404
        emp_id = urow['employee_id']

        if emp_id is not None:
            if email:
                er = conn.execute(
                    'SELECT id FROM employees WHERE email = ? AND id != ? AND email IS NOT NULL AND email != ""',
                    (email, emp_id),
                ).fetchone()
                if er:
                    conn.close()
                    return jsonify({'error': 'Email này đã được sử dụng bởi nhân viên khác'}), 400
            if phone:
                pr = conn.execute(
                    'SELECT id FROM employees WHERE phone = ? AND id != ? AND phone IS NOT NULL AND phone != ""',
                    (phone, emp_id),
                ).fetchone()
                if pr:
                    conn.close()
                    return jsonify({'error': 'Số điện thoại này đã được sử dụng bởi nhân viên khác'}), 400

        conn.execute(
            'UPDATE users SET email = ?, phone = ? WHERE username = ?',
            (email or None, phone or None, g.current_user),
        )

        if emp_id is not None:
            eu, ea = [], []
            if 'employee_name' in data:
                name = (data.get('employee_name') or '').strip()
                if not name:
                    conn.close()
                    return jsonify({'error': 'Họ và tên không được để trống'}), 400
                eu.append('name = ?')
                ea.append(name)
            if 'department_id' in data:
                did = _parse_optional_fk(data.get('department_id'))
                if did is not None:
                    dr = conn.execute('SELECT name FROM departments WHERE id = ?', (did,)).fetchone()
                    if not dr:
                        conn.close()
                        return jsonify({'error': 'Phòng ban không hợp lệ'}), 400
                    eu.append('department_id = ?')
                    ea.append(did)
                    eu.append('department = ?')
                    ea.append(dr['name'])
                else:
                    eu.append('department_id = ?')
                    ea.append(None)
                    eu.append('department = ?')
                    ea.append(None)
            elif 'department' in data:
                eu.append('department = ?')
                ea.append((data.get('department') or '').strip() or None)
                eu.append('department_id = ?')
                ea.append(None)
            if 'birth_date' in data:
                birth_raw = (data.get('birth_date') or '').strip()
                bd = birth_raw[:10] if birth_raw else None
                if bd:
                    try:
                        datetime.strptime(bd, '%Y-%m-%d')
                    except ValueError:
                        conn.close()
                        return jsonify({'error': 'Ngày sinh không hợp lệ (YYYY-MM-DD)'}), 400
                eu.append('birth_date = ?')
                ea.append(bd)
            if 'position_id' in data:
                eu.append('position_id = ?')
                ea.append(_parse_optional_fk(data.get('position_id')))
            if 'shift_id' in data:
                eu.append('shift_id = ?')
                ea.append(_parse_optional_fk(data.get('shift_id')))
            if 'employee_code' in data:
                code = (data.get('employee_code') or '').strip()
                if not code:
                    conn.close()
                    return jsonify({'error': 'Mã nhân viên không được để trống'}), 400
                dup = conn.execute(
                    'SELECT id FROM employees WHERE code = ? AND id != ?',
                    (code, emp_id),
                ).fetchone()
                if dup:
                    conn.close()
                    return jsonify({'error': 'Mã nhân viên đã tồn tại'}), 400
                eu.append('code = ?')
                ea.append(code)
            eu.extend(['email = ?', 'phone = ?'])
            ea.extend([email or None, phone or None])
            ea.append(emp_id)
            conn.execute('UPDATE employees SET ' + ', '.join(eu) + ' WHERE id = ?', tuple(ea))
        else:
            disp = (data.get('employee_name') or data.get('display_name') or '').strip() or None
            birth_raw = (data.get('birth_date') or '').strip()
            bd = birth_raw[:10] if birth_raw else None
            if bd:
                try:
                    datetime.strptime(bd, '%Y-%m-%d')
                except ValueError:
                    conn.close()
                    return jsonify({'error': 'Ngày sinh không hợp lệ (YYYY-MM-DD)'}), 400
            dept = (data.get('department') or '').strip() or None
            pid = _parse_optional_fk(data.get('position_id'))
            sid = _parse_optional_fk(data.get('shift_id'))
            pcode = (data.get('employee_code') or '').strip() or None
            if pcode:
                dup = conn.execute('SELECT id FROM employees WHERE code = ?', (pcode,)).fetchone()
                if dup:
                    conn.close()
                    return jsonify({'error': 'Mã này đã được dùng cho một nhân viên trong hệ thống'}), 400
            conn.execute(
                '''UPDATE users SET display_name = ?, profile_birth_date = ?, profile_department = ?,
                   profile_position_id = ?, profile_shift_id = ?, profile_employee_code = ?
                   WHERE username = ?''',
                (disp, bd, dept, pid, sid, pcode, g.current_user),
            )

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Cập nhật hồ sơ thành công'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/me', methods=['GET'])
@require_auth
def me():
    """Xác thực token. Trả về username, employee_id (nếu tài khoản gắn nhân viên) và thông tin nhân viên.
    Kiểm tra cả admin_users và users (admin users có thể không có employee_id)."""
    conn = None
    try:
        username = g.current_user
        conn = get_db()

        # Ưu tiên: lấy từ users (user gắn nhân viên)
        cur = conn.execute(
            'SELECT u.employee_id, e.id AS emp_id, e.code, e.name, e.department_id, '
            'COALESCE(d.name, e.department) AS department, e.position_id, e.shift_id, e.paid_leave_days_per_year, '
            'e.photo_filename, '
            'p.name AS position_name, s.name AS shift_name, s.start_time AS shift_start, s.end_time AS shift_end '
            'FROM users u LEFT JOIN employees e ON u.employee_id = e.id '
            'LEFT JOIN departments d ON e.department_id = d.id '
            'LEFT JOIN positions p ON e.position_id = p.id LEFT JOIN shifts s ON e.shift_id = s.id '
            'WHERE u.username = ?',
            (username,)
        )
        row = cur.fetchone()

        # Nếu không có trong users → thử admin_users (admin không gắn employee)
        if not row:
            cur = conn.execute(
                'SELECT username, display_name, role FROM admin_users WHERE username = ?',
                (username,)
            )
            admin_row = cur.fetchone()
            conn.close()
            if admin_row:
                return jsonify({
                    'username': username,
                    'display_name': admin_row['display_name'],
                    'role': admin_row['role'],
                    'is_admin': True,
                })
            return jsonify({'username': username})

        conn.close()
        base = request.url_root.rstrip('/')
        photo_fn = row['photo_filename']
        return jsonify({
            'username': username,
            'employee_id': row['emp_id'],
            'employee_code': row['code'],
            'employee_name': row['name'],
            'department_id': row['department_id'],
            'department': row['department'],
            'position_id': row['position_id'],
            'position_name': row['position_name'],
            'shift_id': row['shift_id'],
            'shift_name': row['shift_name'],
            'shift_start': row['shift_start'],
            'shift_end': row['shift_end'],
            'paid_leave_days_per_year': row['paid_leave_days_per_year'] or 12,
            'photo_url': f'{base}/uploads/{photo_fn}' if photo_fn else None,
        })
    except Exception as e:
        return jsonify({'username': g.current_user, 'error': str(e)}), 500


@app.route('/api/ping')
def ping():
    """Keep-alive endpoint cho Render auto-sleep prevention."""
    return jsonify({'ok': True, 'time': datetime.now().isoformat()}), 200


@app.route('/api/config', methods=['GET'])
def get_config():
    """Cấu hình frontend. Đặt GOOGLE_PLACES_API_KEY (env) để bật tìm địa điểm như Google Maps."""
    return jsonify({
        'google_places_api_key': GOOGLE_PLACES_API_KEY,
    })


@app.route('/api/employees', methods=['GET'])
@require_auth
def list_employees():
    conn = None
    try:
        conn = get_db()
        
        # Parse query params
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        search = (request.args.get('search') or '').strip()
        status = (request.args.get('status') or '').strip()
        department_id = request.args.get('department_id', type=int)
        department_name = (request.args.get('department') or '').strip()
        
        print(f"[DEBUG] department_id={department_id}, status={status}")
        
        # Build WHERE clause
        conditions = []
        args = []
        if search:
            conditions.append('(e.name LIKE ? OR e.code LIKE ?)')
            args.extend([f'%{search}%', f'%{search}%'])
        if status and status != 'all':
            conditions.append('e.status = ?')
            args.append(status)
        if department_id:
            conditions.append('e.department_id = ?')
            args.append(department_id)
        elif department_name:
            conditions.append('COALESCE(d.name, e.department) LIKE ?')
            args.append(f'%{department_name}%')
        
        where_clause = ' AND '.join(conditions) if conditions else '1=1'

        # Build base JOINs — dùng chung cho cả data query lẫn count query
        base_joins = (
            'FROM employees e '
            'LEFT JOIN departments d ON e.department_id = d.id '
            'LEFT JOIN positions p ON e.position_id = p.id '
            'LEFT JOIN shifts s ON e.shift_id = s.id '
            'LEFT JOIN offices o ON e.office_id = o.id '
            'LEFT JOIN employee_types et ON e.employee_type_id = et.id '
            'LEFT JOIN salary_policies sp ON e.salary_policy_id = sp.id '
            'LEFT JOIN users u ON u.employee_id = e.id'
        )

        # Get paginated data
        offset = (page - 1) * limit
        sql = f'''
            SELECT e.id, e.code, e.name, e.department_id,
                   COALESCE(d.name, e.department) AS department, e.photo_filename, e.allowed_checkin, e.created_at,
                   e.position_id, e.shift_id, e.paid_leave_days_per_year,
                   e.daily_wage, e.status, e.email, e.phone, e.birth_date,
                   e.office_id, e.employee_type_id, e.salary_policy_id,
                   p.name AS position_name, s.name AS shift_name, s.start_time AS shift_start, s.end_time AS shift_end,
                   o.name AS office_name, et.name AS employee_type_name, sp.name AS salary_policy_name,
                   u.username AS account_username
            {base_joins}
            WHERE {where_clause}
            ORDER BY e.name
        '''

        # Count total — dùng cùng base_joins để tất cả JOIN và WHERE đều hợp lệ
        count_sql = f'SELECT COUNT(*) {base_joins} WHERE {where_clause}'
        total = conn.execute(count_sql, args).fetchone()[0]
        
        # Get paginated rows
        cur = conn.execute(sql + f' LIMIT {limit} OFFSET {offset}', args)
        rows = cur.fetchall()
        base = request.url_root.rstrip('/')
        list_ = []
        for r in rows:
            zones_cur = conn.execute(
                'SELECT id, name, latitude, longitude, radius_meters, created_at FROM employee_zones WHERE employee_id = ? ORDER BY id',
                (r['id'],)
            )
            zones = [{
                'id': z['id'],
                'name': z['name'],
                'latitude': z['latitude'],
                'longitude': z['longitude'],
                'radius_meters': z['radius_meters'],
                'created_at': z['created_at'],
            } for z in zones_cur.fetchall()]
            offices_cur = conn.execute(
                'SELECT eo.id, eo.office_id, o.name AS office_name '
                'FROM employee_offices eo '
                'JOIN offices o ON eo.office_id = o.id '
                'WHERE eo.employee_id = ? ORDER BY o.name',
                (r['id'],)
            )
            offices_list = offices_cur.fetchall()
            office_ids = [o['office_id'] for o in offices_list]
            office_name = offices_list[0]['office_name'] if offices_list else (r['office_name'] or None)
            list_.append({
                'id': r['id'],
                'code': r['code'],
                'name': r['name'],
                'department_id': r['department_id'],
                'department': r['department'] or None,
                'department_name': r['department'] or None,
                'photo_url': f"{base}/uploads/{r['photo_filename']}" if r['photo_filename'] else None,
                'allowed_checkin': bool(r['allowed_checkin'] if r['allowed_checkin'] is not None else 1),
                'zones': zones,
                'position_id': r['position_id'],
                'position_name': r['position_name'],
                'shift_id': r['shift_id'],
                'shift_name': r['shift_name'],
                'shift_start': r['shift_start'],
                'shift_end': r['shift_end'],
                'paid_leave_days_per_year': r['paid_leave_days_per_year'] if r['paid_leave_days_per_year'] is not None else 12,
                'daily_wage': float(r['daily_wage'] if r['daily_wage'] is not None else 0),
                'created_at': r['created_at'],
                'status': r['status'] or 'active',
                'email': r['email'],
                'phone': r['phone'],
                'birth_date': r['birth_date'],
                'office_ids': office_ids,
                'office_name': office_name,
                'employee_type_id': r['employee_type_id'],
                'employee_type_name': r['employee_type_name'],
                'salary_policy_id': r['salary_policy_id'],
                'salary_policy_name': r['salary_policy_name'],
                'account_username': r['account_username'],
            })
        conn.close()
        return jsonify({'data': list_, 'total': total, 'page': page, 'limit': limit})
    except Exception as e:
        import traceback
        print(f"[ERROR] list_employees: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/<int:eid>', methods=['GET'])
@require_auth
def get_employee(eid):
    conn = None
    try:
        conn = get_db()
        cur = conn.execute(
            '''SELECT e.*, d.name AS department_name, p.name AS position_name,
                      s.name AS shift_name, s.start_time AS shift_start, s.end_time AS shift_end,
                      u.username AS account_username, o.name AS office_name,
                      et.name AS employee_type_name,
                      sp.name AS salary_policy_name
               FROM employees e
               LEFT JOIN departments d ON e.department_id = d.id
               LEFT JOIN positions p ON e.position_id = p.id
               LEFT JOIN shifts s ON e.shift_id = s.id
               LEFT JOIN offices o ON e.office_id = o.id
               LEFT JOIN users u ON u.employee_id = e.id
               LEFT JOIN employee_types et ON e.employee_type_id = et.id
               LEFT JOIN salary_policies sp ON e.salary_policy_id = sp.id
               WHERE e.id = ?''', (eid,)
        )
        r = cur.fetchone()
        if not r:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        base = request.url_root.rstrip('/')
        zones_cur = conn.execute(
            'SELECT id, name, latitude, longitude, radius_meters, created_at FROM employee_zones WHERE employee_id = ?', (eid,)
        )
        zones = [{'id': z['id'], 'name': z['name'], 'latitude': z['latitude'],
                  'longitude': z['longitude'], 'radius_meters': z['radius_meters'], 'created_at': z['created_at']}
                 for z in zones_cur.fetchall()]
        # Lấy danh sách office_ids từ bảng employee_offices
        offices_cur = conn.execute(
            'SELECT office_id FROM employee_offices WHERE employee_id = ?', (eid,)
        )
        office_ids = [o['office_id'] for o in offices_cur.fetchall()]
        r_oc = r['office_id']
        if not office_ids and r_oc:
            office_ids = [r_oc]
        conn.close()
        return jsonify({
            'id': r['id'], 'code': r['code'], 'name': r['name'],
            'department_id': r['department_id'], 'department_name': r['department_name'],
            'photo_url': f"{base}/{r['photo_filename']}" if r['photo_filename'] else None,
            'allowed_checkin': bool(r['allowed_checkin'] if r['allowed_checkin'] is not None else 1),
            'zones': zones, 'position_id': r['position_id'], 'position_name': r['position_name'],
            'shift_id': r['shift_id'], 'shift_name': r['shift_name'],
            'shift_start': r['shift_start'], 'shift_end': r['shift_end'],
            'paid_leave_days_per_year': r['paid_leave_days_per_year'] if r['paid_leave_days_per_year'] is not None else 12,
            'daily_wage': float(r['daily_wage'] if r['daily_wage'] is not None else 0),
            'created_at': r['created_at'], 'status': r['status'] or 'active',
            'email': r['email'], 'phone': r['phone'], 'birth_date': r['birth_date'],
            'office_id': r_oc,
            'office_name': r['office_name'],
            'office_ids': office_ids,
            'employee_type_id': r['employee_type_id'],
            'employee_type_name': r['employee_type_name'],
            'salary_policy_id': r['salary_policy_id'],
            'salary_policy_name': r['salary_policy_name'],
            'account_username': r['account_username'],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/<int:eid>', methods=['POST'])
@require_auth
def update_employee_profile(eid):
    conn = None
    try:
        # ... parse các field như cũ ...

        # ✅ Lấy danh sách office (multi-select)
        office_ids_raw = request.form.getlist('office_ids') or request.form.getlist('office_id')
        office_ids = [int(x) for x in office_ids_raw if x]

        conn = get_db()
        cur = conn.execute('SELECT id FROM employees WHERE id = ?', (eid,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404

        # Lấy office_id chính (cái đầu tiên)
        main_office_id = office_ids[0] if office_ids else None

        conn.execute(
            '''UPDATE employees SET name=?, department_id=?, email=?, phone=?,
               paid_leave_days_per_year=?, status=?, position_id=?, shift_id=?,
               office_id=?, employee_type_id=?, salary_policy_id=?
               WHERE id=?''',
            (name, department_id, email, phone,
             int(paid_leave) if paid_leave else None, status,
             position_id, shift_id, main_office_id, employee_type_id, salary_policy_id, eid)
        )

        conn.execute('DELETE FROM employee_offices WHERE employee_id = ?', (eid,))
        for ofid in office_ids:
            conn.execute(
                'INSERT OR IGNORE INTO employee_offices (employee_id, office_id) VALUES (?, ?)',
                (eid, ofid)
            )

        conn.commit()
        conn.close()
        return jsonify({'message': 'Cập nhật thành công'}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        if conn:
            conn.close()  
        return jsonify({'error': str(e)}), 500


# NOTE: duplicate DELETE endpoint removed. Gộp đầy đủ vào delete_employee ở dưới
# (xóa face_embeddings vì bảng này không tồn tại trong schema)

@app.route('/api/employees/stats', methods=['GET'])
@require_auth
def employee_stats():
    conn = None
    try:
        conn = get_db()
        active = conn.execute("SELECT COUNT(*) FROM employees WHERE status='active'").fetchone()[0]
        inactive = conn.execute("SELECT COUNT(*) FROM employees WHERE status='inactive'").fetchone()[0]
        conn.close()
        return jsonify({'active': active, 'inactive': inactive})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/stats', methods=['GET'])
@require_auth
def dashboard_stats():
    conn = None
    try:
        conn = get_db()
        total_emp = conn.execute('SELECT COUNT(*) FROM employees').fetchone()[0]
        today = datetime.utcnow().strftime('%Y-%m-%d')
        today_att = conn.execute(
            "SELECT COUNT(DISTINCT employee_id) FROM attendance WHERE date(check_at) = ?",
            (today,)
        ).fetchone()[0]
        pending_leave = conn.execute(
            "SELECT COUNT(*) FROM leave_requests WHERE status = 'pending'"
        ).fetchone()[0]

        # Weekly attendance data (last 7 days)
        weekly = []
        for i in range(6, -1, -1):
            d = datetime.strptime(today, '%Y-%m-%d') - timedelta(days=i)
            d_str = d.strftime('%Y-%m-%d')
            count = conn.execute(
                "SELECT COUNT(*) FROM attendance WHERE date(check_at) = ?",
                (d_str,)
            ).fetchone()[0]
            weekly.append({'date': d_str, 'label': d.strftime('%d/%m') + ' (' + ['CN','T2','T3','T4','T5','T6','T7'][d.weekday()] + ')', 'count': count})

        # Late count today - use LATE_THRESHOLD from config
        late_rows = conn.execute('''
            SELECT COUNT(DISTINCT a.employee_id) FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            JOIN shifts s ON e.shift_id = s.id
            WHERE date(a.check_at) = ?
              AND a.check_type = 'in'
              AND datetime(a.check_at) > datetime(date(a.check_at) || ' ' || s.start_time, '+' || ? || ' minutes')
        ''', (today, LATE_THRESHOLD)).fetchone()[0]

        # OT count
        ot_rows = conn.execute('''
            SELECT COUNT(DISTINCT employee_id) FROM attendance
            WHERE date(check_at) = ? AND check_type = 'outside'
        ''', (today,)).fetchone()[0]

        conn.close()
        return jsonify({
            'total_employees': total_emp,
            'today_attendance': today_att,
            'pending_leave': pending_leave,
            'late_count': late_rows,
            'ot_count': ot_rows,
            'weekly_data': weekly,
            'employee_stats': {'active': total_emp, 'inactive': 0},
            'leave_balance': 0,
            'leave_total': 0,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/positions/<int:pid>', methods=['POST'])
@require_auth
def update_position_api(pid):
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        description = (data.get('description') or '').strip() or None
        conn = get_db()
        cur = conn.execute('SELECT id FROM positions WHERE id = ?', (pid,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'error': 'Không tìm thấy chức vụ'}), 404
        conn.execute('UPDATE positions SET name=?, description=? WHERE id=?', (name, description, pid))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Cập nhật thành công'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/positions/<int:pid>', methods=['DELETE'])
@require_auth
def delete_position_api(pid):
    conn = None
    try:
        conn = get_db()
        conn.execute('UPDATE employees SET position_id = NULL WHERE position_id = ?', (pid,))
        conn.execute('DELETE FROM positions WHERE id = ?', (pid,))
        conn.commit()
        # Reset AUTOINCREMENT if table is empty
        cur = conn.execute('SELECT COUNT(*) FROM positions')
        if cur.fetchone()[0] == 0:
            conn.execute('DELETE FROM sqlite_sequence WHERE name = ?', ('positions',))
            conn.commit()
        conn.close()
        return jsonify({'message': 'Đã xóa chức vụ'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/shifts/<int:sid>', methods=['DELETE'])
@require_auth
def delete_shift(sid):
    conn = None
    try:
        conn = get_db()
        conn.execute('UPDATE employees SET shift_id = NULL WHERE shift_id = ?', (sid,))
        conn.execute('DELETE FROM shifts WHERE id = ?', (sid,))
        conn.commit()
        cur = conn.execute('SELECT COUNT(*) FROM shifts')
        if cur.fetchone()[0] == 0:
            conn.execute('DELETE FROM sqlite_sequence WHERE name = ?', ('shifts',))
            conn.commit()
        conn.close()
        return jsonify({'message': 'Đã xóa ca'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/departments/<int:did>', methods=['DELETE'])
@require_auth
def delete_department(did):
    conn = None
    try:
        conn = get_db()
        conn.execute('UPDATE employees SET department_id = NULL WHERE department_id = ?', (did,))
        conn.execute('DELETE FROM departments WHERE id = ?', (did,))
        conn.commit()
        cur = conn.execute('SELECT COUNT(*) FROM departments')
        if cur.fetchone()[0] == 0:
            conn.execute('DELETE FROM sqlite_sequence WHERE name = ?', ('departments',))
            conn.commit()
        conn.close()
        return jsonify({'message': 'Đã xóa phòng ban'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/leave-requests/<int:req_id>/approve', methods=['POST'])
@require_auth
def approve_leave_request(req_id):
    conn = None
    try:
        conn = get_db()
        payload = get_token_payload()
        admin_user = payload.get('sub') if payload else 'admin'
        reviewed_at = datetime.utcnow().isoformat() + 'Z'
        conn.execute(
            "UPDATE leave_requests SET status='approved', reviewed_at=?, reviewed_by=? WHERE id=?",
            (reviewed_at, admin_user, req_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'message': 'Đã duyệt đơn'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/leave-requests/<int:req_id>/reject', methods=['POST'])
@require_auth
def reject_leave_request(req_id):
    conn = None
    try:
        conn = get_db()
        payload = get_token_payload()
        admin_user = payload.get('sub') if payload else 'admin'
        reviewed_at = datetime.utcnow().isoformat() + 'Z'
        conn.execute(
            "UPDATE leave_requests SET status='rejected', reviewed_at=?, reviewed_by=? WHERE id=?",
            (reviewed_at, admin_user, req_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'message': 'Đã từ chối đơn'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat', methods=['GET'])
@require_auth
def admin_get_chat():
    """Admin lấy tin nhắn gần nhất (mọi cuộc hội thoại)."""
    conn = None
    try:
        conn = get_db()
        limit = request.args.get('limit', type=int) or 200
        rows = conn.execute('''
            SELECT m.id, m.employee_id, m.message, m.created_at,
                   e.name AS sender_name, e.code AS sender_code,
                   CASE WHEN e.id IS NULL THEN 1 ELSE 0 END AS is_admin
            FROM chat_messages m
            LEFT JOIN employees e ON m.employee_id = e.id
            ORDER BY m.created_at DESC
            LIMIT ?
        ''', (limit,)).fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'], 'employee_id': r['employee_id'], 'message': r['message'],
            'created_at': r['created_at'], 'sender_name': r['sender_name'] or 'Admin',
            'sender_code': r['sender_code'], 'is_admin': r['is_admin']
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat', methods=['POST'])
@require_auth
def admin_send_chat():
    """Admin gửi tin nhắn."""
    conn = None
    try:
        data = _get_request_data()
        message = (data.get('message') or '').strip()
        if not message:
            return jsonify({'error': 'Tin nhắn trống'}), 400
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        conn.execute(
            'INSERT INTO chat_messages (employee_id, message, created_at) VALUES (?, ?, ?)',
            (None, message, created_at)
        )
        conn.commit()
        msg_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
        conn.close()
        return jsonify({'id': msg_id, 'message': message, 'created_at': created_at, 'is_admin': 1, 'sender_name': 'Admin'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/me', methods=['GET'])
@require_auth
def admin_me():
    conn = None
    try:
        payload = get_token_payload()
        username = payload.get('sub') if payload else None
        conn = get_db()
        cur = conn.execute(
            '''SELECT u.username, u.display_name, e.name AS employee_name
               FROM users u LEFT JOIN employees e ON u.employee_id = e.id
               WHERE u.username = ?''',
            (username,)
        )
        row = cur.fetchone()
        conn.close()
        if not row:
            return jsonify({'username': username, 'full_name': 'Administrator', 'email': None, 'role': 'admin'})
        display_name = (row['display_name'] or row['employee_name'] or 'Administrator').strip()
        return jsonify({'username': username, 'full_name': display_name, 'role': 'admin'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/logout', methods=['POST'])
@require_auth
def admin_logout():
    return jsonify({'message': 'OK'}), 200


@app.route('/api/admin/change-password', methods=['POST'])
@require_auth
def admin_change_password():
    """Đổi mật khẩu admin hiện tại. Body JSON: { current, new_password }."""
    conn = None
    try:
        data = _get_request_data()
        current_password = data.get('current') or ''
        new_password = data.get('new_password') or ''
        if not current_password or not new_password:
            return jsonify({'error': 'Thiếu mật khẩu hiện tại hoặc mật khẩu mới'}), 400

        ok, msg = validate_password(new_password)
        if not ok:
            return jsonify({'error': msg}), 400

        conn = get_db()

        # Kiểm tra bảng admin_users trước
        row = conn.execute(
            'SELECT id, password_hash FROM admin_users WHERE username = ?',
            (g.current_user,)
        ).fetchone()

        # Nếu không có trong admin_users, kiểm tra bảng users cũ
        if not row:
            row = conn.execute(
                'SELECT id, password_hash FROM users WHERE username = ?',
                (g.current_user,)
            ).fetchone()
            table = 'users'
        else:
            table = 'admin_users'

        if not row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy tài khoản'}), 404

        if not check_password_hash(row['password_hash'], current_password):
            conn.close()
            return jsonify({'error': 'Mật khẩu hiện tại không đúng'}), 401

        pw_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
        conn.execute(f'UPDATE {table} SET password_hash = ? WHERE username = ?', (pw_hash, g.current_user))
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'message': 'Đã đổi mật khẩu'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Admin Users CRUD ──────────────────────────────────────────────────────────
@app.route('/api/admin/users', methods=['GET'])
@require_auth
def get_admin_users():
    """Lấy danh sách tài khoản admin."""
    conn = None
    try:
        conn = get_db()
        rows = conn.execute(
            'SELECT id, username, display_name, email, phone, role, is_active, created_at FROM admin_users ORDER BY id'
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users', methods=['POST'])
@require_auth
def create_admin_user():
    """Tạo tài khoản admin mới. Body: { username, password, display_name, email, phone, role }."""
    conn = None
    try:
        data = _get_request_data()
        username = (data.get('username') or '').strip()
        password = data.get('password') or ''
        display_name = (data.get('display_name') or '').strip()
        email = (data.get('email') or '').strip()
        phone = (data.get('phone') or '').strip()
        role = (data.get('role') or 'admin').strip()

        if not username or not password:
            return jsonify({'error': 'Thiếu tên đăng nhập hoặc mật khẩu'}), 400

        ok, msg = validate_password(password)
        if not ok:
            return jsonify({'error': msg}), 400

        pw_hash = generate_password_hash(password, method='pbkdf2:sha256')
        conn = get_db()
        try:
            conn.execute(
                '''INSERT INTO admin_users (username, password_hash, display_name, email, phone, role, is_active, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, 1, ?)''',
                (username, pw_hash, display_name, email, phone, role, datetime.utcnow().isoformat() + 'Z')
            )
            conn.commit()
            user_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            conn.close()
            return jsonify({'ok': True, 'id': user_id, 'message': 'Đã tạo tài khoản'})
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': 'Tên đăng nhập đã tồn tại'}), 409
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@require_auth
def update_admin_user(user_id):
    """Sửa tài khoản admin. Body: { display_name, email, phone, role, is_active, password }."""
    conn = None
    try:
        data = _get_request_data()
        display_name = (data.get('display_name') or '').strip()
        email = (data.get('email') or '').strip()
        phone = (data.get('phone') or '').strip()
        role = (data.get('role') or 'admin').strip()
        is_active = 1 if data.get('is_active') else 0
        password = data.get('password') or ''

        conn = get_db()
        row = conn.execute('SELECT id FROM admin_users WHERE id = ?', (user_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy tài khoản'}), 404

        if password:
            ok, msg = validate_password(password)
            if not ok:
                conn.close()
                return jsonify({'error': msg}), 400
            pw_hash = generate_password_hash(password, method='pbkdf2:sha256')
            conn.execute(
                '''UPDATE admin_users SET display_name=?, email=?, phone=?, role=?, is_active=?, password_hash=? WHERE id=?''',
                (display_name, email, phone, role, is_active, pw_hash, user_id)
            )
        else:
            conn.execute(
                '''UPDATE admin_users SET display_name=?, email=?, phone=?, role=?, is_active=? WHERE id=?''',
                (display_name, email, phone, role, is_active, user_id)
            )
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'message': 'Đã cập nhật'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@require_auth
def delete_admin_user(user_id):
    """Xóa tài khoản admin."""
    conn = None
    try:
        conn = get_db()
        # Không cho xóa chính mình
        if g.current_user:
            cur = conn.execute('SELECT username FROM admin_users WHERE id = ?', (user_id,))
            row = cur.fetchone()
            if row and row['username'] == g.current_user:
                conn.close()
                return jsonify({'error': 'Không thể xóa tài khoản của chính bạn'}), 400

        conn.execute('DELETE FROM admin_users WHERE id = ?', (user_id,))
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'message': 'Đã xóa'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/settings', methods=['GET'])
@require_auth
def get_admin_settings():
    """Lấy cài đặt hệ thống."""
    try:
        settings = {
            'company_name': COMPANY_NAME,
            'attendance_radius': ATTENDANCE_RADIUS,
            'late_threshold': LATE_THRESHOLD,
            'language': LANGUAGE,
        }
        return jsonify(settings)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/settings', methods=['PUT'])
@require_auth
def update_admin_settings():
    """Cập nhật cài đặt hệ thống (lưu vào session/.env tùy cấu hình)."""
    try:
        data = _get_request_data()
        if 'company_name' in data:
            os.environ['COMPANY_NAME'] = str(data['company_name'])
        if 'attendance_radius' in data:
            os.environ['ATTENDANCE_RADIUS'] = str(int(data['attendance_radius']))
        if 'late_threshold' in data:
            os.environ['LATE_THRESHOLD'] = str(int(data['late_threshold']))
        if 'language' in data:
            os.environ['LANGUAGE'] = str(data['language'])
        return jsonify({'ok': True, 'message': 'Đã lưu cài đặt'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees', methods=['POST'])
@require_auth
def add_employee():
    try:
        code = (request.form.get('code') or '').strip()
        name = (request.form.get('name') or '').strip()
        if not code or not name:
            return jsonify({'error': 'Thiếu mã hoặc tên nhân viên'}), 400

        photo_filename = None
        embedding_json = None
        # Nhận tối đa 5 ảnh: photo_1 .. photo_5 (hoặc 1 ảnh tên photo để tương thích cũ)
        saved_paths = []
        for key in ['photo_1', 'photo_2', 'photo_3', 'photo_4', 'photo_5']:
            if key in request.files and request.files[key].filename:
                f = request.files[key]
                ext = os.path.splitext(f.filename)[1] or '.jpg'
                base_name = f"{int(datetime.now().timestamp())}_{key}"
                fn = f"{base_name}{ext.replace(' ', '_')}"
                path = os.path.join(UPLOAD_FOLDER, fn)
                f.save(path)
                saved_paths.append(path)
                if photo_filename is None:
                    photo_filename = fn  # Ảnh đại diện dùng ảnh đầu
        if not saved_paths and 'photo' in request.files and request.files['photo'].filename:
            f = request.files['photo']
            ext = os.path.splitext(f.filename)[1] or '.jpg'
            photo_filename = f"{int(datetime.now().timestamp())}{ext.replace(' ', '_')}"
            photo_path = os.path.join(UPLOAD_FOLDER, photo_filename)
            f.save(photo_path)
            saved_paths.append(photo_path)
        if saved_paths:
            emb = compute_embedding_from_images(saved_paths, max_images=5)
            if emb:
                embedding_json = json.dumps(emb)

        position_id = request.form.get('position_id', type=int) or None
        shift_id = request.form.get('shift_id', type=int) or None
        office_id = request.form.get('office_id', type=int) or None
        employee_type_id = request.form.get('employee_type_id', type=int) or None
        salary_policy_id = request.form.get('salary_policy_id', type=int) or None
        paid_leave = request.form.get('paid_leave_days_per_year', type=int)
        paid_leave = paid_leave if paid_leave is not None else 12
        daily_wage = request.form.get('daily_wage', type=float)
        daily_wage = daily_wage if daily_wage is not None and daily_wage >= 0 else 0
        email = (request.form.get('email') or '').strip() or None
        phone = (request.form.get('phone') or '').strip() or None
        account_password_raw = (request.form.get('account_password') or '').strip()
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = None
        conn = get_db()
        department_id = request.form.get('department_id', type=int) or None
        department = None
        if department_id:
            dr = conn.execute('SELECT name FROM departments WHERE id = ?', (department_id,)).fetchone()
            if not dr:
                conn.close()
                return jsonify({'error': 'Phòng ban không hợp lệ'}), 400
            department = dr['name']
        # Kiểm tra trùng email
        if email:
            existing = conn.execute('SELECT id FROM employees WHERE email = ? AND email IS NOT NULL AND email != ""', (email,)).fetchone()
            if existing:
                conn.close()
                return jsonify({'error': 'Email này đã được sử dụng bởi nhân viên khác'}), 400
        # Kiểm tra trùng số điện thoại
        if phone:
            existing = conn.execute('SELECT id FROM employees WHERE phone = ? AND phone IS NOT NULL AND phone != ""', (phone,)).fetchone()
            if existing:
                conn.close()
                return jsonify({'error': 'Số điện thoại này đã được sử dụng bởi nhân viên khác'}), 400
        dup_login = conn.execute('SELECT id FROM users WHERE username = ?', (code,)).fetchone()
        if dup_login:
            conn.close()
            return jsonify({'error': 'Mã nhân viên trùng với tên đăng nhập đã tồn tại. Đổi mã NV hoặc xóa tài khoản cũ.'}), 400

        if account_password_raw:
            pw_ok, pw_msg = validate_password(account_password_raw)
            if not pw_ok:
                conn.close()
                return jsonify({'error': pw_msg}), 400

        conn.execute('BEGIN')
        conn.execute(
            '''INSERT INTO employees (code, name, department, department_id, photo_filename, allowed_checkin, position_id, shift_id, office_id, employee_type_id, salary_policy_id, paid_leave_days_per_year, embedding, daily_wage, email, phone, status, created_at)
               VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?)''',
            (code, name, department, department_id, photo_filename, position_id, shift_id, office_id, employee_type_id, salary_policy_id, paid_leave, embedding_json, daily_wage, email, phone, created_at)
        )
        row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]

        plain_password = account_password_raw if account_password_raw else generate_random_password()
        pw_hash = generate_password_hash(plain_password, method='pbkdf2:sha256')
        try:
            conn.execute(
                'INSERT INTO users (username, password_hash, employee_id, email, phone, created_at) VALUES (?, ?, ?, ?, ?, ?)',
                (code, pw_hash, row_id, email, phone, created_at),
            )
        except sqlite3.IntegrityError:
            conn.rollback()
            conn.close()
            return jsonify({'error': 'Không tạo được tài khoản đăng nhập (lỗi trùng dữ liệu).'}), 400
        conn.commit()
        # Lưu office_ids sau khi có row_id
        office_ids_raw = request.form.getlist('office_ids')
        for ofid in office_ids_raw:
            if ofid:
                conn.execute(
                    'INSERT OR IGNORE INTO employee_offices (employee_id, office_id) VALUES (?, ?)',
                    (row_id, int(ofid))
                )
        conn.close()

        base = request.url_root.rstrip('/')
        payload = {
            'id': row_id,
            'code': code,
            'name': name,
            'department_id': department_id,
            'department': department,
            'photo_url': f"{base}/uploads/{photo_filename}" if photo_filename else None,
            'allowed_checkin': True,
            'position_id': position_id,
            'shift_id': shift_id,
            'office_id': office_id,
            'office_ids': [int(ofid) for ofid in office_ids_raw if ofid],
            'employee_type_id': employee_type_id,
            'salary_policy_id': salary_policy_id,
            'paid_leave_days_per_year': paid_leave,
            'daily_wage': daily_wage,
            'created_at': created_at,
            'account_created': True,
            'account_username': code,
        }
        if not account_password_raw:
            payload['initial_password'] = plain_password
        else:
            payload['initial_password'] = None
        return jsonify(payload), 201
    except sqlite3.IntegrityError:
        if 'conn' in locals() and conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({'error': 'Mã nhân viên đã tồn tại'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/<int:emp_id>', methods=['PATCH'])
@require_auth
def update_employee(emp_id):
    """Cập nhật quyền chấm công hoặc thông tin nhân viên."""
    conn = get_db()
    try:
        data = _get_request_data()
        updates = []
        args = []
        if 'allowed_checkin' in data:
            updates.append('allowed_checkin = ?')
            args.append(1 if data['allowed_checkin'] else 0)
        if 'name' in data and (data['name'] or '').strip():
            updates.append('name = ?')
            args.append((data['name'] or '').strip())
        if 'department_id' in data:
            did = _parse_optional_fk(data.get('department_id'))
            if did is not None:
                conn_v = get_db()
                dr = conn_v.execute('SELECT name FROM departments WHERE id = ?', (did,)).fetchone()
                conn_v.close()
                if not dr:
                    return jsonify({'error': 'Phòng ban không hợp lệ'}), 400
                updates.append('department_id = ?')
                args.append(did)
                updates.append('department = ?')
                args.append(dr['name'])
            else:
                updates.append('department_id = ?')
                args.append(None)
                updates.append('department = ?')
                args.append(None)
        elif 'department' in data:
            updates.append('department = ?')
            args.append((data['department'] or '').strip() or None)
            updates.append('department_id = ?')
            args.append(None)
        if 'position_id' in data:
            updates.append('position_id = ?')
            args.append(data['position_id'] if data['position_id'] else None)
        if 'shift_id' in data:
            updates.append('shift_id = ?')
            args.append(data['shift_id'] if data['shift_id'] else None)
        if 'paid_leave_days_per_year' in data:
            v = data['paid_leave_days_per_year']
            updates.append('paid_leave_days_per_year = ?')
            args.append(int(v) if v is not None else 12)
        if 'daily_wage' in data:
            v = data['daily_wage']
            dv = float(v) if v is not None else 0.0
            if dv < 0:
                return jsonify({'error': 'daily_wage không được âm'}), 400
            updates.append('daily_wage = ?')
            args.append(dv)
        if 'status' in data:
            s = (data['status'] or 'active').strip()
            if s not in ('active', 'inactive'):
                return jsonify({'error': 'status phải là active hoặc inactive'}), 400
            updates.append('status = ?')
            args.append(s)
        if 'email' in data:
            updates.append('email = ?')
            args.append((data['email'] or '').strip() or None)
        if 'phone' in data:
            updates.append('phone = ?')
            args.append((data['phone'] or '').strip() or None)
        if 'birth_date' in data:
            bd = (data.get('birth_date') or '').strip()
            if bd and len(bd) >= 10:
                try:
                    datetime.strptime(bd[:10], '%Y-%m-%d')
                except ValueError:
                    return jsonify({'error': 'birth_date phải là YYYY-MM-DD'}), 400
            updates.append('birth_date = ?')
            args.append(bd[:10] if bd else None)
        if 'office_id' in data:
            oid = _parse_optional_fk(data.get('office_id'))
            updates.append('office_id = ?')
            args.append(oid)
        if 'office_ids' in data:
            oid_list = data.get('office_ids')
            if isinstance(oid_list, list):
                conn.execute('DELETE FROM employee_offices WHERE employee_id = ?', (emp_id,))
                for ofid in oid_list:
                    if ofid:
                        conn.execute(
                            'INSERT OR IGNORE INTO employee_offices (employee_id, office_id) VALUES (?, ?)',
                            (emp_id, int(ofid))
                        )
                # Cập nhật office_id chính (trường office_id của bảng employees)
                main_office = oid_list[0] if oid_list else None
                if main_office:
                    updates.append('office_id = ?')
                    args.append(int(main_office))
                else:
                    updates.append('office_id = ?')
                    args.append(None)
        if 'employee_type_id' in data:
            etid = _parse_optional_fk(data.get('employee_type_id'))
            updates.append('employee_type_id = ?')
            args.append(etid)
        if 'salary_policy_id' in data:
            spid = _parse_optional_fk(data.get('salary_policy_id'))
            updates.append('salary_policy_id = ?')
            args.append(spid)
        if not updates:
            return jsonify({'error': 'Không có trường nào để cập nhật'}), 400
        args.append(emp_id)
        # Kiểm tra trùng email trước khi update
        new_email = (data.get('email', '') or '').strip() or None
        if new_email:
            existing = conn.execute('SELECT id FROM employees WHERE email = ? AND id != ? AND email IS NOT NULL AND email != ""', (new_email, emp_id)).fetchone()
            if existing:
                conn.close()
                return jsonify({'error': 'Email này đã được sử dụng bởi nhân viên khác'}), 400
        new_phone = (data.get('phone', '') or '').strip() or None
        if new_phone:
            existing = conn.execute('SELECT id FROM employees WHERE phone = ? AND id != ? AND phone IS NOT NULL AND phone != ""', (new_phone, emp_id)).fetchone()
            if existing:
                conn.close()
                return jsonify({'error': 'Số điện thoại này đã được sử dụng bởi nhân viên khác'}), 400
        cur = conn.execute(
            'UPDATE employees SET ' + ', '.join(updates) + ' WHERE id = ?',
            args
        )
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        conn.commit()
        if 'email' in data or 'phone' in data:
            erow = conn.execute('SELECT email, phone FROM employees WHERE id = ?', (emp_id,)).fetchone()
            if erow:
                conn.execute(
                    'UPDATE users SET email = ?, phone = ? WHERE employee_id = ?',
                    (
                        ((erow['email'] or '').strip() or None),
                        ((erow['phone'] or '').strip() or None),
                        emp_id,
                    ),
                )
                conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/<int:emp_id>/photo', methods=['POST'])
@require_auth
def update_employee_photo(emp_id):
    """Cập nhật ảnh khuôn mặt nhân viên (upload lại 1-5 ảnh mới)."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT photo_filename FROM employees WHERE id = ?', (emp_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        old_photo = row['photo_filename']
        saved_paths = []
        photo_filename = None
        embedding_json = None
        for i in range(1, 6):
            f = request.files.get(f'photo_{i}')
            if f and f.filename:
                ext = os.path.splitext(f.filename)[1] or '.jpg'
                fn = f"{int(datetime.now().timestamp())}_{i}{ext.replace(' ', '_')}"
                path = os.path.join(UPLOAD_FOLDER, fn)
                f.save(path)
                saved_paths.append(path)
                if photo_filename is None:
                    photo_filename = fn
        if not saved_paths:
            f = request.files.get('photo')
            if f and f.filename:
                ext = os.path.splitext(f.filename)[1] or '.jpg'
                photo_filename = f"{int(datetime.now().timestamp())}{ext.replace(' ', '_')}"
                path = os.path.join(UPLOAD_FOLDER, photo_filename)
                f.save(path)
                saved_paths.append(path)
        if not saved_paths:
            conn.close()
            return jsonify({'error': 'Chưa chọn ảnh nào'}), 400
        emb = compute_embedding_from_images(saved_paths, max_images=5)
        if emb:
            embedding_json = json.dumps(emb)
        conn.execute(
            'UPDATE employees SET photo_filename = ?, embedding = ? WHERE id = ?',
            (photo_filename, embedding_json, emp_id)
        )
        conn.commit()
        conn.close()
        if old_photo:
            old_path = os.path.join(UPLOAD_FOLDER, old_photo)
            if os.path.exists(old_path):
                try:
                    os.unlink(old_path)
                except Exception:
                    pass
        base = request.url_root.rstrip('/')
        return jsonify({
            'ok': True,
            'photo_url': f"{base}/uploads/{photo_filename}" if photo_filename else None,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/<int:emp_id>/zones', methods=['POST'])
@require_auth
def add_employee_zone(emp_id):
    """Thêm vùng chấm công cho nhân viên (cho phép check-in tại đâu)."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        try:
            latitude = float(data.get('latitude', 0))
            longitude = float(data.get('longitude', 0))
            radius_meters = float(data.get('radius_meters', 100))
        except (TypeError, ValueError):
            return jsonify({'error': 'Vĩ độ, kinh độ, bán kính phải là số'}), 400
        if not name:
            return jsonify({'error': 'Thiếu tên vùng'}), 400
        if radius_meters < 50:
            return jsonify({'error': 'Bán kính tối thiểu 50m'}), 400
        conn = get_db()
        cur = conn.execute('SELECT id FROM employees WHERE id = ?', (emp_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn.execute(
            'INSERT INTO employee_zones (employee_id, name, latitude, longitude, radius_meters, created_at) VALUES (?, ?, ?, ?, ?, ?)',
            (emp_id, name, latitude, longitude, radius_meters, created_at)
        )
        conn.commit()
        row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
        conn.close()
        return jsonify({
            'id': row_id,
            'employee_id': emp_id,
            'name': name,
            'latitude': latitude,
            'longitude': longitude,
            'radius_meters': radius_meters,
            'created_at': created_at,
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/location-zones', methods=['GET'])
@require_auth
def list_location_zones():
    """API cho mobile app: trả về danh sách tất cả vùng chấm công (không gắn theo nhân viên).

    Response: List[ { name, latitude, longitude, radius_meters } ]
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.execute(
            'SELECT name, latitude, longitude, radius_meters FROM employee_zones'
        )
        rows = cur.fetchall()
        conn.close()
        return jsonify([
            {
                'name': r['name'],
                'latitude': r['latitude'],
                'longitude': r['longitude'],
                'radius_meters': r['radius_meters'],
            }
            for r in rows
        ])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/location-zones/admin', methods=['GET'])
@require_auth
def list_location_zones_admin():
    """API cho admin: trả về danh sách tất cả vùng chấm công kèm thông tin nhân viên."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('''
            SELECT ez.id, ez.employee_id, ez.name, ez.latitude, ez.longitude, ez.radius_meters, ez.created_at,
                   e.name AS employee_name, e.code AS employee_code
            FROM employee_zones ez
            LEFT JOIN employees e ON ez.employee_id = e.id
            ORDER BY ez.created_at DESC
        ''')
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'employee_id': r['employee_id'],
            'employee_name': r['employee_name'],
            'employee_code': r['employee_code'],
            'name': r['name'],
            'latitude': r['latitude'],
            'longitude': r['longitude'],
            'radius_meters': r['radius_meters'],
            'created_at': r['created_at'],
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _geocode_address(address):
    """Geocode địa chỉ sang tọa độ bằng Nominatim (OpenStreetMap). Trả về (lat, lon) hoặc (None, None)."""
    try:
        import urllib.request
        import urllib.parse
        import json as _json
        url = 'https://nominatim.openstreetmap.org/search?' + urllib.parse.urlencode({
            'q': address,
            'format': 'json',
            'limit': '1',
        })
        req = urllib.request.Request(url, headers={'User-Agent': 'FaceRecognitionApp/1.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read().decode())
            if data:
                return float(data[0]['lat']), float(data[0]['lon'])
    except Exception as e:
        print(f'[WARN] Geocode failed for "{address}": {e}')
    return None, None


@app.route('/api/location-zones/import', methods=['POST'])
@require_auth
def import_location_zones():
    """Import vùng chấm công từ file Excel.

    File Excel format (dùng địa chỉ, tự geocode):
      employee_code,name,address,radius_meters
      NV001,Văn phòng chính,123 Nguyễn Trãi, Quận 5, HCM,100
      NV002,Văn phòng Q1,1 Lê Duẩn, Quận 1, HCM,150

    Hoặc với tọa độ trực tiếp:
      employee_code,name,latitude,longitude,radius_meters
      NV001,Văn phòng chính,10.8231,106.6297,100
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel (.xlsx, .xls)'}), 400

        content = f.read()
        import io as _io
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(content))
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return jsonify({'error': 'File trống'}), 400

        header = None
        rows_data = []
        for i, row in enumerate(rows):
            parts = [str(c).strip() if c is not None else '' for c in row]

            if i == 0:
                header = [h.lower().strip() for h in parts]
                if 'employee_code' not in header and 'name' not in header:
                    return jsonify({'error': 'File Excel thiếu cột "employee_code" và "name"'}), 400
                continue
            if not parts or not any(parts):
                continue
            row_dict = dict(zip(header, parts))
            rows_data.append(row_dict)

        if not rows_data:
            return jsonify({'error': 'Không có dữ liệu hợp lệ trong file'}), 400

        conn = None
        conn = get_db()
        created = 0
        skipped = 0
        errors = []
        geocode_cache = {}
        created_at = datetime.utcnow().isoformat() + 'Z'

        for i, row in enumerate(rows_data, start=2):
            emp_code = (row.get('employee_code') or '').strip()
            zone_name = (row.get('name') or '').strip()
            address = (row.get('address') or '').strip()
            radius = int(float(row.get('radius_meters') or 100) or 100)

            if not emp_code:
                errors.append(f'Dòng {i}: thiếu mã nhân viên')
                continue
            if not zone_name:
                errors.append(f'Dòng {i}: thiếu tên vùng')
                continue
            if radius < 50:
                radius = 50

            emp_row = conn.execute('SELECT id FROM employees WHERE code = ?', (emp_code,)).fetchone()
            if not emp_row:
                errors.append(f'Dòng {i}: không tìm thấy nhân viên mã "{emp_code}"')
                continue
            emp_id = emp_row['id']

            lat = None
            lon = None

            if 'latitude' in row and 'longitude' in row and row.get('latitude') and row.get('longitude'):
                try:
                    lat = float(row.get('latitude'))
                    lon = float(row.get('longitude'))
                except (ValueError, TypeError):
                    pass

            if lat is None and address:
                if address in geocode_cache:
                    lat, lon = geocode_cache[address]
                else:
                    lat, lon = _geocode_address(address)
                    geocode_cache[address] = (lat, lon)

            if lat is None or lon is None:
                errors.append(f'Dòng {i}: không tìm thấy tọa độ cho "{address or zone_name}"')
                skipped += 1
                continue

            existing = conn.execute('SELECT id FROM employee_zones WHERE employee_id = ? AND name = ?',
                                   (emp_id, zone_name)).fetchone()
            if existing:
                conn.execute('UPDATE employee_zones SET latitude=?, longitude=?, radius_meters=? WHERE id=?',
                           (lat, lon, radius, existing['id']))
            else:
                conn.execute('INSERT INTO employee_zones (employee_id, name, latitude, longitude, radius_meters, created_at) VALUES (?, ?, ?, ?, ?, ?)',
                           (emp_id, zone_name, lat, lon, radius, created_at))
                created += 1

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'created': created,
            'skipped': skipped,
            'errors': errors,
            'total': len(rows_data),
            'geocode_cache_size': len(geocode_cache),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-zones', methods=['POST'])
@require_auth
def add_employee_zone_admin():
    """Thêm vùng chấm công cho nhân viên."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        employee_id = data.get('employee_id')
        try:
            latitude = float(data.get('latitude', 0))
            longitude = float(data.get('longitude', 0))
            radius_meters = float(data.get('radius_meters', 100))
        except (TypeError, ValueError):
            return jsonify({'error': 'Vĩ độ, kinh độ, bán kính phải là số'}), 400
        if not name:
            return jsonify({'error': 'Thiếu tên vùng'}), 400
        if radius_meters < 50:
            return jsonify({'error': 'Bán kính tối thiểu 50m'}), 400
        if not employee_id:
            return jsonify({'error': 'Thiếu employee_id'}), 400
        conn = get_db()
        cur = conn.execute('SELECT id FROM employees WHERE id = ?', (employee_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn.execute(
            'INSERT INTO employee_zones (employee_id, name, latitude, longitude, radius_meters, created_at) VALUES (?, ?, ?, ?, ?, ?)',
            (employee_id, name, latitude, longitude, radius_meters, created_at)
        )
        conn.commit()
        row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
        conn.close()
        return jsonify({
            'id': row_id,
            'employee_id': employee_id,
            'name': name,
            'latitude': latitude,
            'longitude': longitude,
            'radius_meters': radius_meters,
            'created_at': created_at,
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-zones/<int:zone_id>', methods=['PUT'])
@require_auth
def update_employee_zone(zone_id):
    """Cập nhật vùng chấm công."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        employee_id = data.get('employee_id')
        try:
            latitude = float(data.get('latitude', 0))
            longitude = float(data.get('longitude', 0))
            radius_meters = float(data.get('radius_meters', 100))
        except (TypeError, ValueError):
            return jsonify({'error': 'Vĩ độ, kinh độ, bán kính phải là số'}), 400
        if not name:
            return jsonify({'error': 'Thiếu tên vùng'}), 400
        if radius_meters < 50:
            return jsonify({'error': 'Bán kính tối thiểu 50m'}), 400
        conn = get_db()
        cur = conn.execute('SELECT id FROM employee_zones WHERE id = ?', (zone_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'error': 'Không tìm thấy vùng'}), 404
        if employee_id:
            cur = conn.execute('SELECT id FROM employees WHERE id = ?', (employee_id,))
            if not cur.fetchone():
                conn.close()
                return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        conn.execute(
            'UPDATE employee_zones SET name=?, employee_id=?, latitude=?, longitude=?, radius_meters=? WHERE id=?',
            (name, employee_id, latitude, longitude, radius_meters, zone_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-zones/<int:zone_id>', methods=['DELETE'])
@require_auth
def delete_employee_zone(zone_id):
    conn = None
    try:
        conn = get_db()
        conn.execute('DELETE FROM employee_zones WHERE id = ?', (zone_id,))
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-offices', methods=['POST'])
@require_auth
def add_employee_office():
    conn = None
    try:
        data = request.get_json() or {}
        emp_id = data.get('employee_id')
        office_id = data.get('office_id')
        if not emp_id or not office_id:
            return jsonify({'error': 'Thiếu employee_id hoặc office_id'}), 400
        conn = get_db()
        try:
            cur = conn.execute(
                'INSERT INTO employee_offices (employee_id, office_id) VALUES (?, ?)',
                (emp_id, office_id)
            )
            conn.commit()
            rid = cur.lastrowid
            conn.close()
            return jsonify({'ok': True, 'id': rid}), 201
        except Exception as ie:
            conn.close()
            if 'UNIQUE' in str(ie) or 'UNIQUE constraint' in str(ie):
                return jsonify({'error': 'Văn phòng này đã được gán cho nhân viên'}), 400
            return jsonify({'error': str(ie)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-offices/<int:rel_id>', methods=['DELETE'])
@require_auth
def delete_employee_office(rel_id):
    conn = None
    try:
        conn = get_db()
        conn.execute('DELETE FROM employee_offices WHERE id = ?', (rel_id,))
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/<int:emp_id>/offices', methods=['GET'])
@require_auth
def list_employee_offices(emp_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.execute(
            'SELECT eo.id, eo.office_id, o.name AS office_name '
            'FROM employee_offices eo '
            'JOIN offices o ON eo.office_id = o.id '
            'WHERE eo.employee_id = ? '
            'ORDER BY o.name',
            (emp_id,)
        )
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'office_id': r['office_id'],
            'office_name': r['office_name']
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/positions', methods=['GET'])
@require_auth
def list_positions():
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT id, name, description, created_at, base_salary, allowance, standard_hours, hourly_rate, overtime_multiplier FROM positions ORDER BY name')
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'name': r['name'],
            'description': r['description'],
            'created_at': r['created_at'],
            'base_salary': float(r['base_salary'] or 0),
            'allowance': float(r['allowance'] or 0),
            'standard_hours': float(r['standard_hours'] or 8),
            'hourly_rate': float(r['hourly_rate'] or 0),
            'overtime_multiplier': float(r['overtime_multiplier'] or 1.5)
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/positions', methods=['POST'])
@require_auth
def add_position():
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        description = (data.get('description') or '').strip() or None
        base_salary = float(data.get('base_salary') or 0)
        allowance = float(data.get('allowance') or 0)
        standard_hours = float(data.get('standard_hours') or 8)
        hourly_rate = float(data.get('hourly_rate') or 0)
        overtime_multiplier = float(data.get('overtime_multiplier') or 1.5)
        if not name:
            return jsonify({'error': 'Thiếu tên chức vụ'}), 400
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        conn.execute('INSERT INTO positions (name, description, created_at, base_salary, allowance, standard_hours, hourly_rate, overtime_multiplier) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                     (name, description, created_at, base_salary, allowance, standard_hours, hourly_rate, overtime_multiplier))
        conn.commit()
        row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
        conn.close()
        return jsonify({
            'id': row_id, 'name': name, 'description': description, 'created_at': created_at,
            'base_salary': base_salary, 'allowance': allowance, 'standard_hours': standard_hours,
            'hourly_rate': hourly_rate, 'overtime_multiplier': overtime_multiplier
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/positions/<int:pid>', methods=['PATCH'])
@require_auth
def update_position(pid):
    conn = None
    try:
        data = _get_request_data()
        updates, args = [], []
        if 'name' in data and (data.get('name') or '').strip():
            updates.append('name = ?')
            args.append((data['name'] or '').strip())
        if 'description' in data:
            updates.append('description = ?')
            args.append((data.get('description') or '').strip() or None)
        if 'base_salary' in data:
            updates.append('base_salary = ?')
            args.append(float(data.get('base_salary') or 0))
        if 'allowance' in data:
            updates.append('allowance = ?')
            args.append(float(data.get('allowance') or 0))
        if 'standard_hours' in data:
            updates.append('standard_hours = ?')
            args.append(float(data.get('standard_hours') or 8))
        if 'hourly_rate' in data:
            updates.append('hourly_rate = ?')
            args.append(float(data.get('hourly_rate') or 0))
        if 'overtime_multiplier' in data:
            updates.append('overtime_multiplier = ?')
            args.append(float(data.get('overtime_multiplier') or 1.5))
        if not updates:
            return jsonify({'error': 'Không có trường cập nhật'}), 400
        args.append(pid)
        conn = get_db()
        cur = conn.execute('UPDATE positions SET ' + ', '.join(updates) + ' WHERE id = ?', args)
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy chức vụ'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/positions/import', methods=['POST'])
@require_auth
def import_positions():
    """Import chức vụ từ file Excel.

    File Excel format:
      name,description,base_salary,allowance,standard_hours,hourly_rate,overtime_multiplier
      Nhân viên,Nhân viên thường,5000000,0,8,0,1.5
      Trưởng phòng,Trưởng phòng,10000000,2000000,8,0,2.0
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel (.xlsx, .xls)'}), 400

        content = f.read()
        import io as _io
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(content))
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return jsonify({'error': 'File trống'}), 400

        header = None
        rows_data = []
        for i, row in enumerate(rows):
            parts = [str(c).strip() if c is not None else '' for c in row]

            if i == 0:
                header = [h.lower().strip() for h in parts]
                if 'name' not in header:
                    return jsonify({'error': 'File Excel thiếu cột "name"'}), 400
                continue
            if not parts or not any(parts):
                continue
            row_dict = dict(zip(header, parts))
            rows_data.append(row_dict)

        if not rows_data:
            return jsonify({'error': 'Không có dữ liệu hợp lệ trong file'}), 400

        conn = None
        conn = get_db()
        created = 0
        updated = 0
        errors = []
        created_at = datetime.utcnow().isoformat() + 'Z'

        for i, row in enumerate(rows_data, start=2):
            name = (row.get('name') or '').strip()
            if not name:
                errors.append(f'Dòng {i}: thiếu tên chức vụ')
                continue
            description = (row.get('description') or '').strip() or None
            base_salary = float(row.get('base_salary') or 0) or 0
            allowance = float(row.get('allowance') or 0) or 0
            standard_hours = float(row.get('standard_hours') or 8) or 8
            hourly_rate = float(row.get('hourly_rate') or 0) or 0
            overtime_multiplier = float(row.get('overtime_multiplier') or 1.5) or 1.5

            existing = conn.execute('SELECT id FROM positions WHERE name = ?', (name,)).fetchone()
            if existing:
                conn.execute('''UPDATE positions SET description=?, base_salary=?, allowance=?,
                               standard_hours=?, hourly_rate=?, overtime_multiplier=? WHERE name=?''',
                           (description, base_salary, allowance, standard_hours, hourly_rate, overtime_multiplier, name))
                updated += 1
            else:
                conn.execute('''INSERT INTO positions (name, description, created_at, base_salary, allowance,
                               standard_hours, hourly_rate, overtime_multiplier) VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                           (name, description, created_at, base_salary, allowance, standard_hours, hourly_rate, overtime_multiplier))
                created += 1

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'created': created,
            'updated': updated,
            'errors': errors,
            'total': len(rows_data),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/shifts', methods=['GET'])
@require_auth
def list_shifts():
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT id, name, start_time, end_time, description, created_at, base_salary, hourly_rate, overtime_multiplier FROM shifts ORDER BY name')
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'], 'name': r['name'], 'start_time': r['start_time'], 'end_time': r['end_time'],
            'description': r['description'], 'created_at': r['created_at'],
            'base_salary': float(r['base_salary'] or 0),
            'hourly_rate': float(r['hourly_rate'] or 0),
            'overtime_multiplier': float(r['overtime_multiplier'] or 1.5),
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/shifts', methods=['POST'])
@require_auth
def add_shift():
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        start_time = (data.get('start_time') or '08:00').strip()
        end_time = (data.get('end_time') or '17:00').strip()
        description = (data.get('description') or '').strip() or None
        if not name:
            return jsonify({'error': 'Thiếu tên ca'}), 400
        base_salary = float(data.get('base_salary') or 0)
        hourly_rate = float(data.get('hourly_rate') or 0)
        overtime_multiplier = float(data.get('overtime_multiplier') or 1.5)
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        conn.execute('INSERT INTO shifts (name, start_time, end_time, description, created_at, base_salary, hourly_rate, overtime_multiplier) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                     (name, start_time, end_time, description, created_at, base_salary, hourly_rate, overtime_multiplier))
        conn.commit()
        row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
        conn.close()
        return jsonify({'id': row_id, 'name': name, 'start_time': start_time, 'end_time': end_time, 'description': description, 'created_at': created_at,
                        'base_salary': base_salary, 'hourly_rate': hourly_rate, 'overtime_multiplier': overtime_multiplier}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/shifts/<int:sid>', methods=['PATCH'])
@require_auth
def update_shift(sid):
    try:
        data = _get_request_data()
        updates, args = [], []
        if 'name' in data and (data.get('name') or '').strip():
            updates.append('name = ?')
            args.append((data.get('name') or '').strip())
        if 'start_time' in data:
            updates.append('start_time = ?')
            args.append((data.get('start_time') or '08:00').strip())
        if 'end_time' in data:
            updates.append('end_time = ?')
            args.append((data.get('end_time') or '17:00').strip())
        if 'description' in data:
            updates.append('description = ?')
            args.append((data.get('description') or '').strip() or None)
        if 'base_salary' in data:
            updates.append('base_salary = ?')
            args.append(float(data.get('base_salary') or 0))
        if 'hourly_rate' in data:
            updates.append('hourly_rate = ?')
            args.append(float(data.get('hourly_rate') or 0))
        if 'overtime_multiplier' in data:
            v = float(data.get('overtime_multiplier') or 1.5)
            if v < 1.0:
                return jsonify({'error': 'Hệ số tăng ca phải >= 1.0'}), 400
            updates.append('overtime_multiplier = ?')
            args.append(v)
        if not updates:
            return jsonify({'error': 'Không có trường cập nhật'}), 400
        args.append(sid)
        conn = None
        conn = get_db()
        cur = conn.execute('UPDATE shifts SET ' + ', '.join(updates) + ' WHERE id = ?', args)
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy ca'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/departments', methods=['GET'])
@require_auth
def list_departments():
    conn = None
    try:
        conn = get_db()
        cur = conn.execute(
            '''SELECT d.id, d.name, d.description, d.created_at,
                      (SELECT COUNT(*) FROM employees e WHERE e.department_id = d.id) AS employee_count
               FROM departments d ORDER BY d.name'''
        )
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'name': r['name'],
            'description': r['description'],
            'created_at': r['created_at'],
            'employee_count': int(r['employee_count'] or 0),
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/departments', methods=['POST'])
@require_auth
def add_department():
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        description = (data.get('description') or '').strip() or None
        if not name:
            return jsonify({'error': 'Thiếu tên phòng ban'}), 400
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        conn.execute(
            'INSERT INTO departments (name, description, created_at) VALUES (?, ?, ?)',
            (name, description, created_at),
        )
        conn.commit()
        row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
        conn.close()
        return jsonify({'id': row_id, 'name': name, 'description': description, 'created_at': created_at}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Tên phòng ban đã tồn tại'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/departments/<int:did>', methods=['PATCH'])
@require_auth
def update_department(did):
    conn = None
    try:
        data = _get_request_data()
        updates, args = [], []
        if 'name' in data and (data.get('name') or '').strip():
            updates.append('name = ?')
            args.append((data.get('name') or '').strip())
        if 'description' in data:
            updates.append('description = ?')
            args.append((data.get('description') or '').strip() or None)
        if not updates:
            return jsonify({'error': 'Không có trường cập nhật'}), 400
        args.append(did)
        conn = get_db()
        cur = conn.execute('UPDATE departments SET ' + ', '.join(updates) + ' WHERE id = ?', args)
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Không tìm thấy phòng ban'}), 404
        if 'name' in data and (data.get('name') or '').strip():
            new_name = (data.get('name') or '').strip()
            conn.execute('UPDATE employees SET department = ? WHERE department_id = ?', (new_name, did))
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Tên phòng ban đã tồn tại'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/departments/import', methods=['POST'])
@require_auth
def import_departments():
    """Import phòng ban từ file Excel.

    File Excel format:
      name,description
      Kỹ thuật,Khối kỹ thuật
      Kinh doanh,Khối kinh doanh

    Header dòng đầu: name,description (bắt buộc)
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel (.xlsx, .xls)'}), 400

        content = f.read()
        import io as _io
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(content))
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return jsonify({'error': 'File trống'}), 400

        header = None
        rows_data = []
        for i, row in enumerate(rows):
            parts = [str(c).strip() if c is not None else '' for c in row]

            if i == 0:
                header = [h.lower().strip() for h in parts]
                if 'name' not in header:
                    return jsonify({'error': 'File Excel thiếu cột "name"'}), 400
                continue
            if not parts or not any(parts):
                continue
            row_dict = dict(zip(header, parts))
            rows_data.append(row_dict)

        if not rows_data:
            return jsonify({'error': 'Không có dữ liệu hợp lệ trong file'}), 400

        conn = None
        conn = get_db()
        created = 0
        updated = 0
        errors = []
        created_at = datetime.utcnow().isoformat() + 'Z'

        for i, row in enumerate(rows_data, start=2):
            name = (row.get('name') or '').strip()
            if not name:
                errors.append(f'Dòng {i}: thiếu tên phòng ban')
                continue
            description = (row.get('description') or '').strip() or None

            existing = conn.execute('SELECT id FROM departments WHERE name = ?', (name,)).fetchone()
            if existing:
                if description is not None or 'description' in header:
                    conn.execute('UPDATE departments SET description = ? WHERE name = ?', (description, name))
                updated += 1
            else:
                conn.execute('INSERT INTO departments (name, description, created_at) VALUES (?, ?, ?)',
                           (name, description, created_at))
                created += 1

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'created': created,
            'updated': updated,
            'errors': errors,
            'total': len(rows_data),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════════
# OFFICES (Văn phòng / Chi nhánh)
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/offices', methods=['GET'])
@require_auth
def list_offices():
    """Lấy danh sách văn phòng."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT id, name, code, address, description, latitude, longitude, radius_meters, is_active, created_at FROM offices ORDER BY name')
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'name': r['name'],
            'code': r['code'],
            'address': r['address'],
            'description': r['description'],
            'latitude': r['latitude'],
            'longitude': r['longitude'],
            'radius_meters': r['radius_meters'],
            'is_active': bool(r['is_active']),
            'created_at': r['created_at'],
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offices', methods=['POST'])
@require_auth
def add_office():
    """Tạo văn phòng mới."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        code = (data.get('code') or '').strip() or None
        address = (data.get('address') or '').strip() or None
        description = (data.get('description') or '').strip() or None
        latitude = float(data['latitude']) if data.get('latitude') not in (None, '') else None
        longitude = float(data['longitude']) if data.get('longitude') not in (None, '') else None
        radius_meters = int(data['radius_meters']) if data.get('radius_meters') not in (None, '') else None
        if radius_meters is not None and radius_meters < 50:
            return jsonify({'error': 'Bán kính tối thiểu 50m'}), 400
        if not name:
            return jsonify({'error': 'Thiếu tên văn phòng'}), 400
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        try:
            conn.execute('INSERT INTO offices (name, code, address, description, latitude, longitude, radius_meters, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                        (name, code, address, description, latitude, longitude, radius_meters, created_at))
            conn.commit()
            row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
            conn.close()
            return jsonify({
                'id': row_id, 'name': name, 'code': code, 'address': address,
                'description': description, 'latitude': latitude, 'longitude': longitude,
                'radius_meters': radius_meters, 'is_active': True, 'created_at': created_at,
            }), 201
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': 'Tên văn phòng đã tồn tại'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offices/<int:oid>', methods=['PATCH'])
@require_auth
def update_office(oid):
    """Cập nhật văn phòng."""
    try:
        data = _get_request_data()
        updates, args = [], []
        for field, key in [('name', 'name'), ('code', 'code'), ('address', 'address'), ('description', 'description')]:
            if key in data:
                updates.append(f'{field} = ?')
                args.append((data[key] or '').strip() or None)
        if 'latitude' in data:
            updates.append('latitude = ?')
            args.append(float(data['latitude']) if data['latitude'] not in (None, '') else None)
        if 'longitude' in data:
            updates.append('longitude = ?')
            args.append(float(data['longitude']) if data['longitude'] not in (None, '') else None)
        if 'radius_meters' in data:
            rm = data['radius_meters']
            if rm not in (None, ''):
                v = int(rm)
                if v < 50:
                    return jsonify({'error': 'Bán kính tối thiểu 50m'}), 400
                updates.append('radius_meters = ?')
                args.append(v)
            else:
                updates.append('radius_meters = ?')
                args.append(None)
        if 'is_active' in data:
            updates.append('is_active = ?')
            args.append(1 if data['is_active'] else 0)
        if not updates:
            return jsonify({'error': 'Không có trường cập nhật'}), 400
        args.append(oid)
        conn = None
        conn = get_db()
        cur = conn.execute('UPDATE offices SET ' + ', '.join(updates) + ' WHERE id = ?', args)
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy văn phòng'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offices/<int:oid>', methods=['DELETE'])
@require_auth
def delete_office(oid):
    """Xóa văn phòng."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('DELETE FROM offices WHERE id = ?', (oid,))
        conn.commit()
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Không tìm thấy văn phòng'}), 404
        # Reset AUTOINCREMENT if table is empty
        count = conn.execute('SELECT COUNT(*) FROM offices').fetchone()[0]
        if count == 0:
            conn.execute('DELETE FROM sqlite_sequence WHERE name = ?', ('offices',))
            conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════════
# EMPLOYEE TYPES (Loại nhân viên: Full-time, Part-time...)
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/employee-types', methods=['GET'])
@require_auth
def list_employee_types():
    """Lấy danh sách loại nhân viên."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT id, name, code, description, is_active, created_at FROM employee_types ORDER BY name')
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'name': r['name'],
            'code': r['code'],
            'description': r['description'],
            'is_active': bool(r['is_active']),
            'created_at': r['created_at'],
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-types', methods=['POST'])
@require_auth
def add_employee_type():
    """Tạo loại nhân viên mới."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        code = (data.get('code') or '').strip() or None
        description = (data.get('description') or '').strip() or None
        if not name:
            return jsonify({'error': 'Thiếu tên loại nhân viên'}), 400
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        try:
            conn.execute('INSERT INTO employee_types (name, code, description, created_at) VALUES (?, ?, ?, ?)',
                        (name, code, description, created_at))
            conn.commit()
            row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
            conn.close()
            return jsonify({
                'id': row_id, 'name': name, 'code': code, 'description': description,
                'is_active': True, 'created_at': created_at,
            }), 201
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': 'Tên loại nhân viên đã tồn tại'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-types/<int:etid>', methods=['PATCH'])
@require_auth
def update_employee_type(etid):
    """Cập nhật loại nhân viên."""
    conn = None
    try:
        data = _get_request_data()
        updates, args = [], []
        for field, key in [('name', 'name'), ('code', 'code'), ('description', 'description')]:
            if key in data:
                updates.append(f'{field} = ?')
                args.append((data[key] or '').strip() or None)
        if 'is_active' in data:
            updates.append('is_active = ?')
            args.append(1 if data['is_active'] else 0)
        if not updates:
            return jsonify({'error': 'Không có trường cập nhật'}), 400
        args.append(etid)
        conn = get_db()
        cur = conn.execute('UPDATE employee_types SET ' + ', '.join(updates) + ' WHERE id = ?', args)
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy loại nhân viên'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-types/<int:etid>', methods=['DELETE'])
@require_auth
def delete_employee_type(etid):
    """Xóa loại nhân viên."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('DELETE FROM employee_types WHERE id = ?', (etid,))
        conn.commit()
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Không tìm thấy loại nhân viên'}), 404
        count = conn.execute('SELECT COUNT(*) FROM employee_types').fetchone()[0]
        if count == 0:
            conn.execute('DELETE FROM sqlite_sequence WHERE name = ?', ('employee_types',))
            conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════════
# SALARY POLICIES (Chính sách lương)
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/salary-policies', methods=['GET'])
@require_auth
def list_salary_policies():
    """Lấy danh sách chính sách lương."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('''SELECT id, name, code, description, pay_frequency, standard_work_days,
                             standard_hours_per_day, overtime_multiplier, is_active, created_at
                             FROM salary_policies ORDER BY name''')
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'name': r['name'],
            'code': r['code'],
            'description': r['description'],
            'pay_frequency': r['pay_frequency'],
            'standard_work_days': float(r['standard_work_days'] or 26),
            'standard_hours_per_day': float(r['standard_hours_per_day'] or 8),
            'overtime_multiplier': float(r['overtime_multiplier'] or 1.5),
            'is_active': bool(r['is_active']),
            'created_at': r['created_at'],
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/salary-policies', methods=['POST'])
@require_auth
def add_salary_policy():
    """Tạo chính sách lương mới."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        code = (data.get('code') or '').strip() or None
        description = (data.get('description') or '').strip() or None
        pay_frequency = (data.get('pay_frequency') or 'monthly').strip().lower()
        standard_work_days = float(data.get('standard_work_days') or 26)
        standard_hours_per_day = float(data.get('standard_hours_per_day') or 8)
        overtime_multiplier = float(data.get('overtime_multiplier') or 1.5)
        if not name:
            return jsonify({'error': 'Thiếu tên chính sách lương'}), 400
        if pay_frequency not in ('monthly', 'weekly', 'biweekly', 'daily', 'hourly'):
            pay_frequency = 'monthly'
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        try:
            conn.execute('''INSERT INTO salary_policies (name, code, description, pay_frequency,
                             standard_work_days, standard_hours_per_day, overtime_multiplier, created_at)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                        (name, code, description, pay_frequency, standard_work_days,
                         standard_hours_per_day, overtime_multiplier, created_at))
            conn.commit()
            row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
            conn.close()
            return jsonify({
                'id': row_id, 'name': name, 'code': code, 'description': description,
                'pay_frequency': pay_frequency, 'standard_work_days': standard_work_days,
                'standard_hours_per_day': standard_hours_per_day,
                'overtime_multiplier': overtime_multiplier, 'is_active': True, 'created_at': created_at,
            }), 201
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': 'Tên chính sách lương đã tồn tại'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/salary-policies/<int:spid>', methods=['PATCH'])
@require_auth
def update_salary_policy(spid):
    """Cập nhật chính sách lương."""
    conn = None
    try:
        data = _get_request_data()
        updates, args = [], []
        for field, key in [('name', 'name'), ('code', 'code'), ('description', 'description')]:
            if key in data:
                updates.append(f'{field} = ?')
                args.append((data[key] or '').strip() or None)
        for field, key in [('pay_frequency', 'pay_frequency'), ('standard_work_days', 'standard_work_days'),
                           ('standard_hours_per_day', 'standard_hours_per_day'),
                           ('overtime_multiplier', 'overtime_multiplier')]:
            if key in data:
                updates.append(f'{field} = ?')
                args.append(float(data[key] or (26 if 'days' in key else 8 if 'hours' in key else 1.5)))
        if 'is_active' in data:
            updates.append('is_active = ?')
            args.append(1 if data['is_active'] else 0)
        if not updates:
            return jsonify({'error': 'Không có trường cập nhật'}), 400
        args.append(spid)
        conn = get_db()
        cur = conn.execute('UPDATE salary_policies SET ' + ', '.join(updates) + ' WHERE id = ?', args)
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy chính sách lương'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/salary-policies/<int:spid>', methods=['DELETE'])
@require_auth
def delete_salary_policy(spid):
    """Xóa chính sách lương."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('DELETE FROM salary_policies WHERE id = ?', (spid,))
        conn.commit()
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Không tìm thấy chính sách lương'}), 404
        count = conn.execute('SELECT COUNT(*) FROM salary_policies').fetchone()[0]
        if count == 0:
            conn.execute('DELETE FROM sqlite_sequence WHERE name = ?', ('salary_policies',))
            conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Import Offices ─────────────────────────────────────────────────────────────
def _parse_excel(f):
    """Parse file Excel, trả về list dicts có header ở dòng đầu."""
    ext = os.path.splitext(f.filename)[1].lower()
    content = f.read()
    import io as _io
    import openpyxl as _openpyxl
    wb = _openpyxl.load_workbook(_io.BytesIO(content))
    ws = wb.active
    rows = list(ws.values)
    if not rows:
        return None, 'File trống'
    header = None
    rows_data = []
    for i, row in enumerate(rows):
        parts = [str(c).strip() if c is not None else '' for c in row]
        if i == 0:
            header = [h.lower().strip() for h in parts]
            if 'name' not in header:
                return None, 'File thiếu cột "name"'
            continue
        if not parts or not any(parts):
            continue
        rows_data.append(dict(zip(header, parts)))
    if not rows_data:
        return None, 'Không có dữ liệu hợp lệ'
    return rows_data, None


@app.route('/api/offices/import', methods=['POST'])
@require_auth
def import_offices():
    """Import văn phòng từ file Excel.

    File Excel format:
      name,code,address,latitude,longitude,radius_meters,description
      Văn phòng chính,VP001,123 Nguyễn Huệ Q1,10.762912,106.679814,100,Trụ sở chính
      Chi nhánh HCM,CNHCM,456 Điện Biên Phủ,10.762912,106.679814,100,CN HCM
    """
    conn = None
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel'}), 400

        rows_data, err = _parse_excel(f)
        if err:
            return jsonify({'error': err}), 400

        conn = get_db()
        created = 0; updated = 0; errors = []
        created_at = datetime.utcnow().isoformat() + 'Z'

        for i, row in enumerate(rows_data, start=2):
            name = (row.get('name') or '').strip()
            if not name:
                errors.append(f'Dòng {i}: thiếu tên')
                continue
            code = (row.get('code') or '').strip() or None
            address = (row.get('address') or '').strip() or None
            description = (row.get('description') or '').strip() or None
            lat = None
            if row.get('latitude') not in (None, '') or row.get('vi_do') not in (None, ''):
                try:
                    lat = float(row.get('latitude') or row.get('vi_do'))
                except:
                    pass
            lng = None
            if row.get('longitude') not in (None, '') or row.get('kinh_do') not in (None, ''):
                try:
                    lng = float(row.get('longitude') or row.get('kinh_do'))
                except:
                    pass
            rm_raw = row.get('radius_meters') or row.get('ban_kinh_met')
            radius = int(float(rm_raw)) if rm_raw not in (None, '') else None
            if radius is not None and radius < 50:
                errors.append(f'Dòng {i}: bán kính tối thiểu 50m')
                radius = None

            existing = conn.execute('SELECT id FROM offices WHERE name = ?', (name,)).fetchone()
            if existing:
                conn.execute(
                    'UPDATE offices SET code=COALESCE(?,code), address=COALESCE(?,address), latitude=COALESCE(?,latitude), longitude=COALESCE(?,longitude), radius_meters=COALESCE(?,radius_meters), description=COALESCE(?,description) WHERE name=?',
                    (code, address, lat, lng, radius, description, name))
                updated += 1
            else:
                conn.execute('INSERT INTO offices (name, code, address, latitude, longitude, radius_meters, description, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                           (name, code, address, lat, lng, radius, description, created_at))
                created += 1

        conn.commit(); conn.close()
        return jsonify({'ok': True, 'created': created, 'updated': updated, 'errors': errors})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employee-types/import', methods=['POST'])
@require_auth
def import_employee_types():
    """Import loại nhân viên từ file Excel.

    File Excel format:
      name,code,description
      Full-time,FT,Nhân viên chính thức
      Part-time,PT,Nhân viên bán thời gian
      Thực tập,TT,Nhân viên thực tập
    """
    conn = None
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel'}), 400

        rows_data, err = _parse_excel(f)
        if err:
            return jsonify({'error': err}), 400

        conn = get_db()
        created = 0; updated = 0; errors = []
        created_at = datetime.utcnow().isoformat() + 'Z'

        for i, row in enumerate(rows_data, start=2):
            name = (row.get('name') or '').strip()
            if not name:
                errors.append(f'Dòng {i}: thiếu tên')
                continue
            code = (row.get('code') or '').strip() or None
            description = (row.get('description') or '').strip() or None

            existing = conn.execute('SELECT id FROM employee_types WHERE name = ?', (name,)).fetchone()
            if existing:
                if code or description:
                    conn.execute('UPDATE employee_types SET code=COALESCE(?,code), description=COALESCE(?,description) WHERE name=?',
                               (code, description, name))
                updated += 1
            else:
                conn.execute('INSERT INTO employee_types (name, code, description, created_at) VALUES (?, ?, ?, ?)',
                           (name, code, description, created_at))
                created += 1

        conn.commit(); conn.close()
        return jsonify({'ok': True, 'created': created, 'updated': updated, 'errors': errors})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/salary-policies/import', methods=['POST'])
@require_auth
def import_salary_policies():
    """Import chính sách lương từ file Excel.

    File Excel format:
      name,code,description,pay_frequency,standard_work_days,standard_hours_per_day,overtime_multiplier
      Lương tháng,L-thang,Lương theo tháng,monthly,26,8,1.5
      Lương tuần,L-tuan,Lương theo tuần,weekly,5,8,2.0
    """
    conn = None
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel'}), 400

        rows_data, err = _parse_excel(f)
        if err:
            return jsonify({'error': err}), 400

        conn = get_db()
        created = 0; updated = 0; errors = []
        created_at = datetime.utcnow().isoformat() + 'Z'
        valid_freqs = ('monthly', 'weekly', 'biweekly', 'daily', 'hourly')

        for i, row in enumerate(rows_data, start=2):
            name = (row.get('name') or '').strip()
            if not name:
                errors.append(f'Dòng {i}: thiếu tên')
                continue
            code = (row.get('code') or '').strip() or None
            description = (row.get('description') or '').strip() or None
            pay_frequency = (row.get('pay_frequency') or 'monthly').strip().lower()
            if pay_frequency not in valid_freqs:
                pay_frequency = 'monthly'
            try:
                standard_work_days = float(row.get('standard_work_days') or 26)
                standard_hours_per_day = float(row.get('standard_hours_per_day') or 8)
                overtime_multiplier = float(row.get('overtime_multiplier') or 1.5)
            except:
                errors.append(f'Dòng {i}: số không hợp lệ')
                continue

            existing = conn.execute('SELECT id FROM salary_policies WHERE name = ?', (name,)).fetchone()
            if existing:
                conn.execute('''UPDATE salary_policies SET code=COALESCE(?,code), description=COALESCE(?,description),
                               pay_frequency=COALESCE(?,pay_frequency), standard_work_days=COALESCE(?,standard_work_days),
                               standard_hours_per_day=COALESCE(?,standard_hours_per_day), overtime_multiplier=COALESCE(?,overtime_multiplier)
                               WHERE name=?''',
                           (code, description, pay_frequency, standard_work_days, standard_hours_per_day, overtime_multiplier, name))
                updated += 1
            else:
                conn.execute('''INSERT INTO salary_policies (name, code, description, pay_frequency, standard_work_days,
                               standard_hours_per_day, overtime_multiplier, created_at)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                           (name, code, description, pay_frequency, standard_work_days, standard_hours_per_day, overtime_multiplier, created_at))
                created += 1

        conn.commit(); conn.close()
        return jsonify({'ok': True, 'created': created, 'updated': updated, 'errors': errors})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/<int:emp_id>/account', methods=['POST'])
@require_auth
def create_employee_account(emp_id):
    """Cấp hoặc cập nhật tài khoản đăng nhập cho nhân viên.

    Body JSON: { username, password }.
    - Nếu chưa có tài khoản -> tạo mới.
    - Nếu đã có -> cho phép đổi mật khẩu, và chỉ chặn trùng username với NHÂN VIÊN KHÁC.
    """
    conn = None
    try:
        data = _get_request_data()
        username = (data.get('username') or '').strip()
        password = data.get('password') or ''
        if not username or not password:
            return jsonify({'error': 'Thiếu tên đăng nhập hoặc mật khẩu'}), 400

        ok, msg = validate_password(password)
        if not ok:
            return jsonify({'error': msg}), 400

        conn = get_db()

        # Kiểm tra nhân viên tồn tại + email/SĐT để đồng bộ quên mật khẩu
        cur = conn.execute('SELECT id, email, phone FROM employees WHERE id = ?', (emp_id,))
        emp_row = cur.fetchone()
        if not emp_row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        emp_email = (emp_row['email'] or '').strip() or None
        emp_phone = (emp_row['phone'] or '').strip() or None

        # Lấy thông tin tài khoản hiện tại (nếu có)
        cur = conn.execute('SELECT id, username FROM users WHERE employee_id = ?', (emp_id,))
        user_row = cur.fetchone()

        if user_row:
            # Đã có tài khoản: chỉ chặn khi username này đang được dùng bởi user khác
            cur = conn.execute('SELECT id FROM users WHERE username = ? AND id != ?', (username, user_row['id']))
            if cur.fetchone():
                conn.close()
                return jsonify({'error': 'Tên đăng nhập đã tồn tại cho nhân viên khác'}), 400
        else:
            # Chưa có tài khoản: chặn nếu username đã tồn tại ở bất kỳ user nào
            cur = conn.execute('SELECT id FROM users WHERE username = ?', (username,))
            if cur.fetchone():
                conn.close()
                return jsonify({'error': 'Tên đăng nhập đã tồn tại'}), 400

        created_at = datetime.utcnow().isoformat() + 'Z'
        pw_hash = generate_password_hash(password, method='pbkdf2:sha256')

        if user_row:
            conn.execute(
                'UPDATE users SET username = ?, password_hash = ?, email = ?, phone = ? WHERE id = ?',
                (username, pw_hash, emp_email, emp_phone, user_row['id'])
            )
        else:
            conn.execute(
                'INSERT INTO users (username, password_hash, employee_id, email, phone, created_at) VALUES (?, ?, ?, ?, ?, ?)',
                (username, pw_hash, emp_id, emp_email, emp_phone, created_at)
            )

        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'message': 'Đã cấp/ cập nhật tài khoản'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/leave-requests', methods=['GET'])
@require_auth
def list_leave_requests():
    """Danh sách đơn xin nghỉ. Query: employee_id (nếu không gửi và token có employee_id thì dùng), status, year."""
    conn = None
    try:
        employee_id = request.args.get('employee_id', type=int)
        status = request.args.get('status')
        year = request.args.get('year', type=int)
        if employee_id is None:
            payload = get_token_payload()
            if payload:
                conn0 = get_db()
                cur0 = conn0.execute('SELECT employee_id FROM users WHERE username = ?', (payload.get('sub'),))
                row0 = cur0.fetchone()
                conn0.close()
                if row0 and row0['employee_id']:
                    employee_id = row0['employee_id']
        conn = get_db()
        sql = '''SELECT l.id, l.employee_id, l.from_date, l.to_date, l.leave_type, l.reason, l.status, l.created_at, l.reviewed_at, l.reviewed_by, l.review_note,
                        e.name AS employee_name, e.code AS employee_code
                 FROM leave_requests l JOIN employees e ON l.employee_id = e.id WHERE 1=1'''
        args = []
        if employee_id is not None:
            sql += ' AND l.employee_id = ?'
            args.append(employee_id)
        if status:
            sql += ' AND l.status = ?'
            args.append(status)
        if year:
            sql += ' AND (l.from_date LIKE ? OR l.to_date LIKE ?)'
            args.extend([f'{year}%', f'{year}%'])
        sql += ' ORDER BY l.created_at DESC LIMIT 500'
        cur = conn.execute(sql, args)
        rows = cur.fetchall()
        conn.close()
        list_ = [{
            'id': r['id'], 'employee_id': r['employee_id'], 'from_date': r['from_date'], 'to_date': r['to_date'],
            'leave_type': r['leave_type'], 'reason': r['reason'], 'status': r['status'],
            'created_at': r['created_at'], 'reviewed_at': r['reviewed_at'], 'reviewed_by': r['reviewed_by'],
            'review_note': r['review_note'] or '',
            'employee_name': r['employee_name'], 'employee_code': r['employee_code'],
        } for r in rows]
        return jsonify(list_)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/leave-requests', methods=['POST'])
@require_auth
def create_leave_request():
    """Tạo đơn xin nghỉ. Body: { employee_id (optional nếu token có), from_date, to_date, leave_type: paid|unpaid, reason }."""
    conn = None
    try:
        data = _get_request_data()
        employee_id = data.get('employee_id')
        if employee_id is None:
            payload = get_token_payload()
            if payload:
                conn = get_db()
                cur = conn.execute('SELECT employee_id FROM users WHERE username = ?', (payload.get('sub'),))
                row = cur.fetchone()
                conn.close()
                if row and row['employee_id']:
                    employee_id = row['employee_id']
        if employee_id is None:
            return jsonify({'error': 'Thiếu employee_id hoặc đăng nhập bằng tài khoản nhân viên'}), 400
        employee_id = int(employee_id)
        from_date = (data.get('from_date') or '').strip()
        to_date = (data.get('to_date') or '').strip()
        leave_type = (data.get('leave_type') or 'unpaid').strip().lower()
        if leave_type not in ('paid', 'unpaid'):
            leave_type = 'unpaid'
        reason = (data.get('reason') or '').strip() or None
        if not from_date or not to_date:
            return jsonify({'error': 'Thiếu from_date hoặc to_date (YYYY-MM-DD)'}), 400
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        cur = conn.execute('SELECT id FROM employees WHERE id = ?', (employee_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        conn.execute(
            'INSERT INTO leave_requests (employee_id, from_date, to_date, leave_type, reason, status, created_at) VALUES (?, ?, ?, ?, ?, \'pending\', ?)',
            (employee_id, from_date, to_date, leave_type, reason, created_at)
        )
        conn.commit()
        row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
        conn.close()
        return jsonify({
            'id': row_id, 'employee_id': employee_id, 'from_date': from_date, 'to_date': to_date,
            'leave_type': leave_type, 'reason': reason, 'status': 'pending', 'created_at': created_at,
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/leave-requests/<int:req_id>', methods=['PATCH'])
@require_auth
def update_leave_request(req_id):
    """Duyệt/từ chối đơn. Body: { status: approved|rejected, review_note? }."""
    conn = None
    try:
        data = _get_request_data()
        status = (data.get('status') or '').strip().lower()
        if status not in ('approved', 'rejected'):
            return jsonify({'error': 'status phải là approved hoặc rejected'}), 400
        review_note = (data.get('review_note') or '').strip() or None
        reviewed_at = datetime.utcnow().isoformat() + 'Z'
        reviewed_by = g.current_user
        conn = get_db()
        cur = conn.execute(
            'UPDATE leave_requests SET status = ?, reviewed_at = ?, reviewed_by = ?, review_note = ? WHERE id = ?',
            (status, reviewed_at, reviewed_by, review_note, req_id)
        )
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy đơn'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/leave-balance', methods=['GET'])
@require_auth
def get_leave_balance():
    """Số buổi nghỉ có lương / không lương. Query: employee_id (optional nếu token có), year, month."""
    conn = None
    try:
        employee_id = request.args.get('employee_id', type=int)
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        if employee_id is None:
            payload = get_token_payload()
            if payload:
                conn = get_db()
                cur = conn.execute('SELECT employee_id FROM users WHERE username = ?', (payload.get('sub'),))
                row = cur.fetchone()
                conn.close()
                if row and row['employee_id']:
                    employee_id = row['employee_id']
        if employee_id is None:
            return jsonify({'error': 'Thiếu employee_id hoặc đăng nhập bằng tài khoản nhân viên'}), 400
        conn = get_db()
        cur = conn.execute('SELECT paid_leave_days_per_year FROM employees WHERE id = ?', (employee_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        total_paid_per_year = row['paid_leave_days_per_year'] or 12
        year = year or datetime.utcnow().year
        from datetime import date as _date
        year_start = _date(year, 1, 1)
        year_end = _date(year, 12, 31)
        cur = conn.execute(
            '''SELECT leave_type, from_date, to_date FROM leave_requests
               WHERE employee_id = ? AND status = 'approved' AND from_date <= ? AND to_date >= ?''',
            (employee_id, year_end.isoformat(), year_start.isoformat())
        )
        rows = cur.fetchall()
        def overlap_days(d1_str, d2_str, range_start, range_end):
            try:
                a = max(datetime.strptime(d1_str[:10], '%Y-%m-%d').date(), range_start)
                b = min(datetime.strptime(d2_str[:10], '%Y-%m-%d').date(), range_end)
                return max(0, (b - a).days + 1)
            except Exception:
                return 0
        used_paid = 0
        used_unpaid = 0
        for r in rows:
            d = overlap_days(r['from_date'], r['to_date'], year_start, year_end)
            if r['leave_type'] == 'paid':
                used_paid += d
            else:
                used_unpaid += d
        if month:
            import calendar
            month_start = _date(year, month, 1)
            month_end = _date(year, month, calendar.monthrange(year, month)[1])
            cur2 = conn.execute(
                '''SELECT leave_type, from_date, to_date FROM leave_requests
                   WHERE employee_id = ? AND status = 'approved' AND from_date <= ? AND to_date >= ?''',
                (employee_id, month_end.isoformat(), month_start.isoformat())
            )
            month_rows = cur2.fetchall()
            used_paid_month = sum(overlap_days(r['from_date'], r['to_date'], month_start, month_end) for r in month_rows if r['leave_type'] == 'paid')
            used_unpaid_month = sum(overlap_days(r['from_date'], r['to_date'], month_start, month_end) for r in month_rows if r['leave_type'] == 'unpaid')
            conn.close()
            return jsonify({
                'employee_id': employee_id,
                'year': year,
                'month': month,
                'paid_per_year': total_paid_per_year,
                'used_paid_year': used_paid,
                'remaining_paid_year': max(0, total_paid_per_year - used_paid),
                'used_unpaid_month': used_unpaid_month,
                'used_paid_month': used_paid_month,
            })
        conn.close()
        return jsonify({
            'employee_id': employee_id,
            'year': year,
            'paid_per_year': total_paid_per_year,
            'used_paid_year': used_paid,
            'remaining_paid_year': max(0, total_paid_per_year - used_paid),
            'used_unpaid_year': used_unpaid,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat/conversations', methods=['GET'])
@require_auth
def chat_conversations():
    """Danh sách hội thoại cho admin."""
    conn = None
    try:
        conn = get_db()
        current_employee_id = get_user_employee_id(conn, g.current_user)
        if current_employee_id is not None:
            conn.close()
            return jsonify({'error': 'Chỉ admin mới xem được toàn bộ hội thoại'}), 403

        rows = conn.execute(
            '''SELECT e.id AS employee_id, e.code, e.name, COALESCE(d.name, e.department) AS department,
                      (
                        SELECT CASE
                                 WHEN COALESCE(TRIM(cm.message), '') <> '' THEN cm.message
                                 WHEN cm.attachment_url IS NOT NULL THEN '[Tệp đính kèm]'
                                 ELSE ''
                               END
                        FROM chat_messages cm
                        WHERE cm.employee_id = e.id
                        ORDER BY cm.created_at DESC, cm.id DESC
                        LIMIT 1
                      ) AS last_message,
                      (
                        SELECT cm.created_at FROM chat_messages cm
                        WHERE cm.employee_id = e.id
                        ORDER BY cm.created_at DESC, cm.id DESC
                        LIMIT 1
                      ) AS last_created_at,
                      (
                        SELECT COUNT(*) FROM chat_messages cm
                        WHERE cm.employee_id = e.id AND cm.sender_type = 'employee' AND cm.is_read = 0
                      ) AS unread_count
               FROM employees e
               LEFT JOIN departments d ON e.department_id = d.id
               ORDER BY COALESCE(last_created_at, e.created_at) DESC, e.name ASC'''
        ).fetchall()
        conn.close()
        return jsonify([{
            'employee_id': r['employee_id'],
            'employee_code': r['code'],
            'employee_name': r['name'],
            'department': r['department'],
            'last_message': r['last_message'],
            'last_created_at': r['last_created_at'],
            'unread_count': int(r['unread_count'] or 0),
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat/messages', methods=['GET'])
@require_auth
def list_chat_messages():
    """Lấy lịch sử chat.
    - Admin: truyền employee_id để xem hội thoại 1 nhân viên.
    - Nhân viên: tự động lấy theo employee_id từ tài khoản.
    """
    conn = None
    try:
        conn = get_db()
        current_employee_id = get_user_employee_id(conn, g.current_user)
        requested_employee_id = request.args.get('employee_id', type=int)
        limit = request.args.get('limit', type=int) or 200
        limit = max(1, min(500, limit))

        if current_employee_id is None:
            if requested_employee_id is None:
                conn.close()
                return jsonify({'error': 'Thiếu employee_id'}), 400
            employee_id = requested_employee_id
        else:
            employee_id = current_employee_id
            if requested_employee_id is not None and requested_employee_id != current_employee_id:
                conn.close()
                return jsonify({'error': 'Bạn không có quyền xem hội thoại này'}), 403

        emp = conn.execute(
            'SELECT id, code, name FROM employees WHERE id = ?',
            (employee_id,)
        ).fetchone()
        if not emp:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404

        rows = conn.execute(
            '''SELECT id, employee_id, sender_type, sender_username, message,
                      attachment_name, attachment_url, attachment_type, attachment_size,
                      is_read, created_at
               FROM chat_messages
               WHERE employee_id = ?
               ORDER BY created_at ASC, id ASC
               LIMIT ?''',
            (employee_id, limit)
        ).fetchall()
        items = []
        for r in rows:
            items.append({
                'id': r['id'],
                'employee_id': r['employee_id'],
                'sender_type': r['sender_type'],
                'sender_username': r['sender_username'],
                'sender_display_name': _sender_display_name(conn, r['sender_username'], r['sender_type']),
                'message': r['message'],
                'attachment_name': r['attachment_name'],
                'attachment_url': r['attachment_url'],
                'attachment_type': r['attachment_type'],
                'attachment_size': int(r['attachment_size'] or 0),
                'is_read': bool(r['is_read']),
                'created_at': r['created_at'],
            })
        conn.close()
        return jsonify({
            'employee_id': emp['id'],
            'employee_code': emp['code'],
            'employee_name': emp['name'],
            'items': items,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat/messages', methods=['POST'])
@require_auth
def send_chat_message():
    """Gửi tin nhắn.
    - Admin: body cần employee_id và message.
    - Nhân viên: chỉ cần message, employee_id tự lấy từ tài khoản.
    """
    conn = None
    try:
        if request.content_type and request.content_type.startswith('multipart/form-data'):
            data = request.form or {}
            file_obj = request.files.get('file')
        else:
            data = _get_request_data()
            file_obj = None
        message = (data.get('message') or '').strip()
        if len(message) > 4000:
            return jsonify({'error': 'Tin nhắn quá dài (tối đa 4000 ký tự)'}), 400

        conn = get_db()
        current_employee_id = get_user_employee_id(conn, g.current_user)
        requested_employee_id = data.get('employee_id')
        if requested_employee_id is not None:
            try:
                requested_employee_id = int(requested_employee_id)
            except (TypeError, ValueError):
                conn.close()
                return jsonify({'error': 'employee_id không hợp lệ'}), 400

        if current_employee_id is None:
            if requested_employee_id is None:
                conn.close()
                return jsonify({'error': 'Thiếu employee_id'}), 400
            employee_id = requested_employee_id
            sender_type = 'admin'
        else:
            employee_id = current_employee_id
            sender_type = 'employee'
            if requested_employee_id is not None and requested_employee_id != current_employee_id:
                conn.close()
                return jsonify({'error': 'Bạn không có quyền gửi vào hội thoại này'}), 403

        emp = conn.execute('SELECT id FROM employees WHERE id = ?', (employee_id,)).fetchone()
        if not emp:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404

        attachment_name = None
        attachment_url = None
        attachment_type = None
        attachment_size = 0
        if file_obj and file_obj.filename:
            original_name = secure_filename(file_obj.filename)
            if not original_name:
                conn.close()
                return jsonify({'error': 'Tên tệp không hợp lệ'}), 400
            ext = os.path.splitext(original_name)[1].lower()
            # Block dangerous extensions + web shells
            blocked_exts = {
                '.exe', '.bat', '.cmd', '.com', '.js', '.msi', '.sh', '.ps1',
                '.php', '.phtml', '.php3', '.php4', '.php5', '.php7', '.phar',
                '.asp', '.aspx', '.cer', '.cgi', '.jsp', '.jspx', '.shtml',
                '.py', '.rb', '.pl', '.cgi', '.htaccess',
                '.html', '.htm', '.svg', '.xhtml',
            }
            if ext in blocked_exts:
                conn.close()
                return jsonify({'error': 'Loại tệp không được phép upload'}), 400
            unique_name = f"{int(datetime.utcnow().timestamp())}_{random.randint(1000, 9999)}_{original_name}"
            save_path = os.path.join(CHAT_UPLOAD_FOLDER, unique_name)
            file_obj.save(save_path)
            attachment_name = original_name
            attachment_url = request.url_root.rstrip('/') + '/uploads/chat/' + unique_name
            attachment_type = (file_obj.mimetype or '').strip() or None
            try:
                attachment_size = int(os.path.getsize(save_path))
            except Exception:
                attachment_size = 0

        if not message and not attachment_url:
            conn.close()
            return jsonify({'error': 'Nội dung tin nhắn không được để trống'}), 400

        created_at = datetime.utcnow().isoformat() + 'Z'
        conn.execute(
            '''INSERT INTO chat_messages (
                   employee_id, sender_type, sender_username, message,
                   attachment_name, attachment_url, attachment_type, attachment_size,
                   is_read, created_at
               ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?)''',
            (
                employee_id,
                sender_type,
                g.current_user,
                message,
                attachment_name,
                attachment_url,
                attachment_type,
                attachment_size,
                created_at,
            )
        )
        conn.commit()
        msg_id = conn.execute('SELECT last_insert_rowid() AS id').fetchone()['id']
        disp = _sender_display_name(conn, g.current_user, sender_type)
        conn.close()
        return jsonify({
            'id': msg_id,
            'employee_id': employee_id,
            'sender_type': sender_type,
            'sender_username': g.current_user,
            'sender_display_name': disp,
            'message': message,
            'attachment_name': attachment_name,
            'attachment_url': attachment_url,
            'attachment_type': attachment_type,
            'attachment_size': attachment_size,
            'is_read': False,
            'created_at': created_at,
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat/read', methods=['POST'])
@require_auth
def mark_chat_as_read():
    """Đánh dấu đã đọc tin nhắn đối phương."""
    conn = None
    try:
        data = _get_request_data()
        conn = get_db()
        current_employee_id = get_user_employee_id(conn, g.current_user)
        requested_employee_id = data.get('employee_id')
        if requested_employee_id is not None:
            try:
                requested_employee_id = int(requested_employee_id)
            except (TypeError, ValueError):
                conn.close()
                return jsonify({'error': 'employee_id không hợp lệ'}), 400

        if current_employee_id is None:
            if requested_employee_id is None:
                conn.close()
                return jsonify({'error': 'Thiếu employee_id'}), 400
            employee_id = requested_employee_id
            sender_to_mark = 'employee'
        else:
            employee_id = current_employee_id
            sender_to_mark = 'admin'
            if requested_employee_id is not None and requested_employee_id != current_employee_id:
                conn.close()
                return jsonify({'error': 'Bạn không có quyền thao tác hội thoại này'}), 403

        cur = conn.execute(
            'UPDATE chat_messages SET is_read = 1 WHERE employee_id = ? AND sender_type = ? AND is_read = 0',
            (employee_id, sender_to_mark)
        )
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'updated': cur.rowcount})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat/unread-count', methods=['GET'])
@require_auth
def get_chat_unread_count():
    """Số tin chưa đọc cho người dùng hiện tại (admin ↔ employee + p2p)."""
    conn = None
    try:
        conn = get_db()
        current_employee_id = get_user_employee_id(conn, g.current_user)
        if current_employee_id is None:
            row = conn.execute(
                "SELECT COUNT(*) AS c FROM chat_messages WHERE sender_type = 'employee' AND is_read = 0"
            ).fetchone()
        else:
            admin_unread = conn.execute(
                "SELECT COUNT(*) AS c FROM chat_messages WHERE employee_id = ? AND sender_type = 'admin' AND is_read = 0",
                (current_employee_id,)
            ).fetchone()
            p2p_unread = conn.execute(
                "SELECT COUNT(*) AS c FROM chat_messages WHERE room_id IS NOT NULL AND employee_id = ? AND sender_employee_id != ? AND is_read = 0",
                (current_employee_id, current_employee_id)
            ).fetchone()
            row = admin_unread
        conn.close()
        return jsonify({'unread_count': int(row['c'] or 0)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _p2p_room(a: int, b: int) -> str:
    """Tạo room ID bất đối xứng từ 2 employee ID."""
    return f"p2p_{min(a, b)}_{max(a, b)}"


@app.route('/api/chat/p2p/conversations', methods=['GET'])
@require_auth
def p2p_conversations():
    """Danh sách hội thoại peer-to-peer của nhân viên hiện tại."""
    conn = None
    try:
        conn = get_db()
        emp_id = get_user_employee_id(conn, g.current_user)
        if emp_id is None:
            conn.close()
            return jsonify({'error': 'Chỉ nhân viên mới dùng được tính năng này'}), 403

        rows = conn.execute(
            '''SELECT DISTINCT
                 CASE WHEN cm.sender_employee_id = ? THEN cm.employee_id
                      ELSE cm.sender_employee_id END AS peer_id,
                 e.code AS peer_code, e.name AS peer_name,
                 COALESCE(d.name, e.department) AS peer_department,
                 (
                   SELECT CASE
                     WHEN COALESCE(TRIM(cm2.message), '') <> '' THEN cm2.message
                     WHEN cm2.attachment_url IS NOT NULL THEN '[Tệp đính kèm]'
                     ELSE ''
                   END
                   FROM chat_messages cm2
                   WHERE cm2.room_id = cm.room_id
                   ORDER BY cm2.created_at DESC, cm2.id DESC LIMIT 1
                 ) AS last_message,
                 (
                   SELECT cm2.created_at FROM chat_messages cm2
                   WHERE cm2.room_id = cm.room_id
                   ORDER BY cm2.created_at DESC, cm2.id DESC LIMIT 1
                 ) AS last_created_at,
                 (
                   SELECT COUNT(*) FROM chat_messages cm2
                   WHERE cm2.room_id = cm.room_id
                     AND cm2.sender_employee_id != ?
                     AND cm2.is_read = 0
                 ) AS unread_count
               FROM chat_messages cm
               JOIN employees e ON e.id = CASE WHEN cm.sender_employee_id = ? THEN cm.employee_id ELSE cm.sender_employee_id END
               LEFT JOIN departments d ON e.department_id = d.id
               WHERE cm.room_id IS NOT NULL
                 AND (cm.sender_employee_id = ? OR cm.employee_id = ?)
               ORDER BY last_created_at DESC, e.name ASC''',
            (emp_id, emp_id, emp_id, emp_id, emp_id)
        ).fetchall()
        conn.close()
        return jsonify([{
            'peer_id': r['peer_id'],
            'peer_code': r['peer_code'],
            'peer_name': r['peer_name'],
            'peer_department': r['peer_department'],
            'last_message': r['last_message'] or '',
            'last_created_at': r['last_created_at'],
            'unread_count': int(r['unread_count'] or 0),
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat/p2p/messages/<int:peer_id>', methods=['GET'])
@require_auth
def p2p_messages(peer_id):
    """Lấy tin nhắn peer-to-peer giữa nhân viên hiện tại và peer_id."""
    conn = None
    try:
        conn = get_db()
        emp_id = get_user_employee_id(conn, g.current_user)
        if emp_id is None:
            conn.close()
            return jsonify({'error': 'Chỉ nhân viên mới dùng được tính năng này'}), 403

        if emp_id == peer_id:
            conn.close()
            return jsonify({'error': 'Không thể nhắn tin với chính mình'}), 400

        peer = conn.execute('SELECT id, code, name FROM employees WHERE id = ?', (peer_id,)).fetchone()
        if not peer:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404

        room = _p2p_room(emp_id, peer_id)
        rows = conn.execute(
            '''SELECT cm.id, cm.sender_employee_id, cm.employee_id, cm.sender_type,
                      cm.sender_username, cm.message, cm.attachment_name,
                      cm.attachment_url, cm.attachment_type, cm.attachment_size,
                      cm.is_read, cm.created_at,
                      COALESCE(e.name, u.display_name) AS sender_display_name
               FROM chat_messages cm
               LEFT JOIN employees e ON e.id = cm.sender_employee_id
               LEFT JOIN users u ON u.username = cm.sender_username AND u.employee_id IS NULL
               WHERE cm.room_id = ?
               ORDER BY cm.created_at ASC, cm.id ASC
               LIMIT 500''',
            (room,)
        ).fetchall()

        conn.execute(
            'UPDATE chat_messages SET is_read = 1 WHERE room_id = ? AND sender_employee_id != ? AND is_read = 0',
            (room, emp_id)
        )
        conn.commit()
        conn.close()

        return jsonify({
            'peer': {'id': peer['id'], 'code': peer['code'], 'name': peer['name']},
            'room': room,
            'items': [{
                'id': r['id'],
                'sender_employee_id': r['sender_employee_id'],
                'sender_type': r['sender_type'],
                'sender_username': r['sender_username'],
                'sender_display_name': r['sender_display_name'],
                'message': r['message'],
                'attachment_name': r['attachment_name'],
                'attachment_url': r['attachment_url'],
                'attachment_type': r['attachment_type'],
                'attachment_size': int(r['attachment_size'] or 0),
                'is_read': r['is_read'],
                'created_at': r['created_at'],
            } for r in rows]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/chat/p2p/send', methods=['POST'])
@require_auth
def p2p_send():
    """Gửi tin nhắn peer-to-peer từ nhân viên đến peer."""
    conn = None
    try:
        conn = get_db()
        emp_id = get_user_employee_id(conn, g.current_user)
        if emp_id is None:
            conn.close()
            return jsonify({'error': 'Chỉ nhân viên mới dùng được tính năng này'}), 403

        if request.content_type and request.content_type.startswith('multipart/form-data'):
            data = request.form or {}
            file_obj = request.files.get('file')
        else:
            data = _get_request_data()
            file_obj = None

        peer_id = data.get('peer_id')
        if peer_id is None:
            conn.close()
            return jsonify({'error': 'Thiếu peer_id'}), 400
        try:
            peer_id = int(peer_id)
        except (TypeError, ValueError):
            conn.close()
            return jsonify({'error': 'peer_id không hợp lệ'}), 400
        if peer_id == emp_id:
            conn.close()
            return jsonify({'error': 'Không thể nhắn tin với chính mình'}), 400

        peer = conn.execute('SELECT id, code, name FROM employees WHERE id = ?', (peer_id,)).fetchone()
        if not peer:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404

        message = (data.get('message') or '').strip()
        if len(message) > 4000:
            return jsonify({'error': 'Tin nhắn quá dài (tối đa 4000 ký tự)'}), 400

        room = _p2p_room(emp_id, peer_id)
        attachment_name = None
        attachment_url = None
        attachment_type = None
        attachment_size = 0
        if file_obj and file_obj.filename:
            original_name = secure_filename(file_obj.filename)
            if not original_name:
                conn.close()
                return jsonify({'error': 'Tên tệp không hợp lệ'}), 400
            ext = os.path.splitext(original_name)[1].lower()
            # Block dangerous extensions + web shells
            blocked_exts = {
                '.exe', '.bat', '.cmd', '.com', '.js', '.msi', '.sh', '.ps1',
                '.php', '.phtml', '.php3', '.php4', '.php5', '.php7', '.phar',
                '.asp', '.aspx', '.cer', '.cgi', '.jsp', '.jspx', '.shtml',
                '.py', '.rb', '.pl', '.cgi', '.htaccess',
                '.html', '.htm', '.svg', '.xhtml',
            }
            if ext in blocked_exts:
                conn.close()
                return jsonify({'error': 'Loại tệp không được phép upload'}), 400
            unique_name = f"{int(datetime.utcnow().timestamp())}_{random.randint(1000, 9999)}_{original_name}"
            save_path = os.path.join(CHAT_UPLOAD_FOLDER, unique_name)
            file_obj.save(save_path)
            attachment_name = original_name
            attachment_url = request.url_root.rstrip('/') + '/uploads/chat/' + unique_name
            attachment_type = (file_obj.mimetype or '').strip() or None
            try:
                attachment_size = int(os.path.getsize(save_path))
            except Exception:
                attachment_size = 0

        if not message and not attachment_url:
            conn.close()
            return jsonify({'error': 'Nội dung tin nhắn không được để trống'}), 400

        created_at = datetime.utcnow().isoformat() + 'Z'
        conn.execute(
            '''INSERT INTO chat_messages (
                   employee_id, sender_type, sender_username, sender_employee_id, room_id,
                   message, attachment_name, attachment_url, attachment_type, attachment_size,
                   is_read, created_at
               ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (peer_id, 'employee', g.current_user['username'], emp_id, room,
             message, attachment_name, attachment_url, attachment_type, attachment_size,
             0, created_at)
        )
        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'room': room,
            'message': message,
            'attachment_url': attachment_url,
            'created_at': created_at,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/bulk-import', methods=['POST'])
@require_auth
def bulk_import_employees():
    """Import hàng loạt nhân viên từ file Excel.

    File cần có các cột (header row bắt buộc):
      code, name, department, email?, phone?, position_id?, shift_id?,
      paid_leave_days_per_year?, daily_wage?, allowed_checkin?

    Hỗ trợ cột theo alias: ma/nhanvien/employee_code, ho-ten/name/ten,
    phongban/department/phong_ban, email, dienthoai/phone/so_dien_thoai.

    Query params (tùy chọn):
      skip_header=1    – bỏ qua dòng đầu tiên (mặc định: 1 nếu file có header)
      create_account=1 – tạo tài khoản cho mỗi nhân viên (mặc định: 0)
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Không có file được upload'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Chưa chọn file'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    skip_header = request.args.get('skip_header', '1') == '1'
    create_account = request.args.get('create_account', '0') == '1'

    if ext not in ('.xlsx', '.xls'):
        return jsonify({'error': 'Chỉ hỗ trợ file Excel (.xlsx, .xls)'}), 400

    try:
        # Đọc nội dung file Excel
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = [[str(cell.value).strip() if cell.value is not None else ''
                 for cell in row] for row in ws.iter_rows()]

        if len(rows) == 0:
            return jsonify({'error': 'File rỗng'}), 400

        # Xác định header
        header_row = 0 if skip_header else None
        if header_row is not None and len(rows) > header_row:
            raw_header = rows[header_row]
            # Chuẩn hóa: lower, bỏ dấu, thay khoảng trắng bằng _
            import unicodedata
            def norm(h):
                s = unicodedata.normalize('NFKD', h.lower())
                s = ''.join(c for c in s if not unicodedata.combining(c))
                s = s.replace(' ', '_').replace('-', '_')
                return s

            raw_header = [norm(h) for h in raw_header]
            # Map alias -> canonical
            ALIAS = {
                'code': ['code', 'ma', 'manv', 'employee_code', 'employeecode', 'msnv', 'emp_code'],
                'name': ['name', 'ten', 'ho_ten', 'hoten', 'fullname', 'full_name', 'ho_ten_nv'],
                'department': ['department', 'phongban', 'phong_ban', 'bo_phan', 'bophan', 'dept'],
                'email': ['email', 'mail', 'gmail'],
                'phone': ['phone', 'dienthoai', 'dien_thoai', 'so_dien_thoai', 'sdt', 'mobile'],
                'position_id': ['position_id', 'chucvu_id', 'chuc_vu_id', 'positionid'],
                'shift_id': ['shift_id', 'ca_id', 'ca_lam_id', 'shiftid'],
                'office_id': ['office_id', 'vanphong_id', 'van_phong_id', 'officeid', 'vp_id'],
                'employee_type_id': ['employee_type_id', 'loainv_id', 'loai_nv_id', 'employeetypeid', 'loai_nhan_vien_id'],
                'salary_policy_id': ['salary_policy_id', 'csluong_id', 'cs_luong_id', 'salarypolicyid', 'chinh_sach_luong_id'],
                'paid_leave': ['paid_leave_days_per_year', 'ngay_phep', 'paid_leave', 'paidleave'],
                'daily_wage': ['daily_wage', 'luong_ngay', 'luongngay', 'dailywage'],
                'allowed_checkin': ['allowed_checkin', 'allowed', 'cho_phep', 'checkin'],
            }
            col_map = {}
            for canon, aliases in ALIAS.items():
                for i, h in enumerate(raw_header):
                    if h in aliases:
                        col_map[canon] = i
                        break

            if 'code' not in col_map or 'name' not in col_map:
                return jsonify({
                    'error': f'File thiếu cột bắt buộc: code, name. Header hiện tại: {raw_header}'}), 400

            data_rows = rows[1:] if skip_header else rows
        else:
            col_map = {'code': 0, 'name': 1}
            data_rows = rows

        def get_val(row, canon):
            idx = col_map.get(canon)
            if idx is None or idx >= len(row):
                return ''
            v = row[idx]
            return v.strip() if isinstance(v, str) else str(v).strip()

        created = 0
        skipped = 0
        errors = []
        results = []

        conn = None
        conn = get_db()
        conn.execute('BEGIN IMMEDIATE')  # Khóa ngay để đảm bảo transaction
        now = datetime.utcnow().isoformat() + 'Z'

        for i, row in enumerate(data_rows):
            line_num = (header_row + 1 + i + 1) if header_row is not None else (i + 1)
            code = get_val(row, 'code')
            name = get_val(row, 'name')
            if not code or not name:
                errors.append({'row': line_num, 'error': 'Thiếu mã hoặc tên'})
                skipped += 1
                continue

            # Kiểm tra trùng mã
            existing = conn.execute('SELECT id FROM employees WHERE code = ?', (code,)).fetchone()
            if existing:
                errors.append({'row': line_num, 'error': f'Mã nhân viên "{code}" đã tồn tại'})
                skipped += 1
                continue

            email = get_val(row, 'email') or None
            phone = get_val(row, 'phone') or None
            department = get_val(row, 'department') or None
            position_id_raw = get_val(row, 'position_id')
            shift_id_raw = get_val(row, 'shift_id')
            paid_leave_raw = get_val(row, 'paid_leave')
            daily_wage_raw = get_val(row, 'daily_wage')
            allowed_raw = get_val(row, 'allowed_checkin')
            office_id_raw = get_val(row, 'office_id')
            employee_type_id_raw = get_val(row, 'employee_type_id')
            salary_policy_id_raw = get_val(row, 'salary_policy_id')

            office_id = int(office_id_raw) if office_id_raw and office_id_raw.isdigit() else None
            employee_type_id = int(employee_type_id_raw) if employee_type_id_raw and employee_type_id_raw.isdigit() else None
            salary_policy_id = int(salary_policy_id_raw) if salary_policy_id_raw and salary_policy_id_raw.isdigit() else None
            position_id = int(position_id_raw) if position_id_raw.isdigit() else None
            shift_id = int(shift_id_raw) if shift_id_raw.isdigit() else None
            paid_leave = int(paid_leave_raw) if paid_leave_raw.isdigit() else 12
            daily_wage = float(daily_wage_raw) if daily_wage_raw.replace('.', '', 1).isdigit() else 0.0
            allowed_checkin = 1 if allowed_raw.lower() in ('1', 'true', 'yes', 'co', 'cho_phep') else 0

            department_id = None
            if department:
                dept_row = conn.execute('SELECT id FROM departments WHERE name = ?', (department,)).fetchone()
                if dept_row:
                    department_id = dept_row['id']
                else:
                    conn.execute('INSERT INTO departments (name, created_at) VALUES (?, ?)',
                                  (department, now))
                    department_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

            if email:
                dup = conn.execute('SELECT id FROM employees WHERE email = ? AND email IS NOT NULL AND email != ""',
                                   (email,)).fetchone()
                if dup:
                    errors.append({'row': line_num, 'error': f'Email "{email}" đã được sử dụng'})
                    skipped += 1
                    continue

            if phone:
                dup = conn.execute('SELECT id FROM employees WHERE phone = ? AND phone IS NOT NULL AND phone != ""',
                                   (phone,)).fetchone()
                if dup:
                    errors.append({'row': line_num, 'error': f'SĐT "{phone}" đã được sử dụng'})
                    skipped += 1
                    continue

            try:
                conn.execute('''
                    INSERT INTO employees
                    (code, name, department, department_id, email, phone,
                     position_id, shift_id, office_id, employee_type_id, salary_policy_id,
                     paid_leave_days_per_year, daily_wage, allowed_checkin, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (code, name, department, department_id, email, phone,
                      position_id, shift_id, office_id, employee_type_id, salary_policy_id,
                      paid_leave, daily_wage, allowed_checkin, now))
                emp_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

                if create_account:
                    raw_pw = generate_random_password()
                    pw_hash = generate_password_hash(raw_pw, method='pbkdf2:sha256')
                    try:
                        conn.execute('''
                            INSERT INTO users (username, password_hash, employee_id, created_at)
                            VALUES (?, ?, ?, ?)
                        ''', (code, pw_hash, emp_id, now))
                    except sqlite3.IntegrityError as account_err:
                        # Username đã tồn tại - vẫn tạo employee nhưng không tạo account
                        errors.append({'row': line_num, 'warning': f'Tài khoản cho mã "{code}" đã tồn tại, bỏ qua tạo account. Lỗi: {account_err}'})

                conn.commit()
                created += 1
                item = {'row': line_num, 'code': code, 'name': name, 'id': emp_id}
                if create_account:
                    item['password'] = raw_pw
                results.append(item)
            except Exception as ex:
                conn.rollback()
                errors.append({'row': line_num, 'error': f'Lỗi khi tạo nhân viên: {ex}'})
                skipped += 1
                # Tiếp tục xử lý các dòng khác (rollback chỉ ảnh hưởng dòng lỗi)

        # Commit tất cả thay đổi thành công
        try:
            conn.commit()
        except Exception as commit_err:
            conn.rollback()
            return jsonify({
                'error': f'Lỗi khi lưu: {commit_err}. Không có dữ liệu nào được lưu.',
                'total_errors': len(errors),
                'errors': errors[:100],
            }), 500

        conn.close()
        return jsonify({
            'ok': True,
            'total_rows': len(data_rows),
            'created': created,
            'skipped': skipped,
            'total_errors': len(errors),
            'errors': errors[:100],
            'results': results[:100],
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/employees/<int:emp_id>', methods=['DELETE'])
@require_auth
def delete_employee(emp_id):
    """Xóa nhân viên và toàn bộ dữ liệu liên quan (attendance, zones, chat, leaves, users)."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT photo_filename FROM employees WHERE id = ?', (emp_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        photo_filename = row['photo_filename']

        # Xóa theo thứ tự để tránh FK constraint violation
        conn.execute('DELETE FROM employee_zones WHERE employee_id = ?', (emp_id,))
        conn.execute('DELETE FROM chat_messages WHERE employee_id = ?', (emp_id,))
        conn.execute('DELETE FROM leave_requests WHERE employee_id = ?', (emp_id,))
        conn.execute('DELETE FROM attendance WHERE employee_id = ?', (emp_id,))
        conn.execute('DELETE FROM users WHERE employee_id = ?', (emp_id,))
        conn.execute('DELETE FROM employees WHERE id = ?', (emp_id,))
        conn.commit()
        conn.close()

        # Xóa file ảnh khuôn mặt
        if photo_filename:
            path = os.path.join(UPLOAD_FOLDER, photo_filename)
            if os.path.exists(path):
                try:
                    os.unlink(path)
                except OSError:
                    pass
        return jsonify({'ok': True, 'message': 'Đã xóa nhân viên và dữ liệu liên quan'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/attendance', methods=['POST'])
@require_auth
def add_attendance():
    """Ghi nhận chấm công: body JSON { employee_id, check_type, check_at, latitude?, longitude?, reason?, shift_id? }.
    check_type: 'in' | 'out' | 'outside' (ra ngoài / công tác).
    'outside' luôn tính đủ công trong bảng lương.
    shift_id: ca làm việc (từ mobile app truyền lên, hoặc admin chọn tay)."""
    try:
        data = _get_request_data()
        employee_id = data.get('employee_id')
        check_type = (data.get('check_type') or '').strip().lower()
        check_at = data.get('check_at')
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        reason = data.get('reason') or ''
        shift_id = data.get('shift_id')
        if employee_id is None:
            return jsonify({'error': 'Thiếu employee_id'}), 400
        if check_type not in ('in', 'out', 'outside'):
            return jsonify({'error': 'check_type phải là "in", "out" hoặc "outside"'}), 400
        if check_type == 'outside' and not reason.strip():
            return jsonify({'error': 'Khi chấm công "Ra ngoài" cần nhập lý do (reason).'}), 400
        try:
            employee_id = int(employee_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'employee_id phải là số'}), 400
        if shift_id is not None:
            try:
                shift_id = int(shift_id)
            except (TypeError, ValueError):
                shift_id = None
        if check_at:
            try:
                dt = datetime.fromisoformat(check_at.replace('Z', '+00:00'))
                check_at = dt.strftime('%Y-%m-%dT%H:%M:%S')
            except Exception:
                check_at = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
        else:
            check_at = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
        lat = float(latitude) if latitude is not None else None
        lng = float(longitude) if longitude is not None else None

        conn = None
        conn = get_db()
        cur = conn.execute('SELECT id FROM employees WHERE id = ?', (employee_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'error': 'Không tìm thấy nhân viên'}), 404
        created_at = datetime.utcnow().isoformat() + 'Z'
        # 'outside' mặc định pending — cần admin phê duyệt mới tính lương
        att_status = 'pending' if check_type == 'outside' else 'approved'
        conn.execute(
            'INSERT INTO attendance (employee_id, check_type, check_at, latitude, longitude, reason, created_at, shift_id, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (employee_id, check_type, check_at, lat, lng, reason.strip(), created_at, shift_id, att_status)
        )
        conn.commit()
        row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
        conn.close()
        return jsonify({
            'id': row_id,
            'employee_id': employee_id,
            'check_type': check_type,
            'check_at': check_at,
            'latitude': lat,
            'longitude': lng,
            'reason': reason.strip(),
            'created_at': created_at,
            'shift_id': shift_id,
            'status': att_status,
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/attendance', methods=['GET'])
@require_auth
def list_attendance():
    """Danh sách chấm công. Query: employee_id (optional), year, month hoặc from, to (date)."""
    conn = None
    try:
        employee_id = request.args.get('employee_id', type=int)
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        from_date = request.args.get('from')
        to_date = request.args.get('to')

        conn = get_db()
        if year is not None and month is not None:
            start = f'{year}-{month:02d}-01T00:00:00'
            if month == 12:
                end = f'{year + 1}-01-01T00:00:00'
            else:
                end = f'{year}-{month + 1:02d}-01T00:00:00'
            if employee_id is not None:
                cur = conn.execute(
                    '''SELECT a.id, a.employee_id, a.check_type, a.check_at, a.image_path, a.latitude, a.longitude, a.reason, a.shift_id, a.status,
                       e.name AS employee_name, e.code AS employee_code
                       FROM attendance a JOIN employees e ON a.employee_id = e.id
                       WHERE a.employee_id = ? AND a.check_at >= ? AND a.check_at < ?
                       ORDER BY a.check_at ASC''',
                    (employee_id, start, end)
                )
            else:
                cur = conn.execute(
                    '''SELECT a.id, a.employee_id, a.check_type, a.check_at, a.image_path, a.latitude, a.longitude, a.reason, a.shift_id, a.status,
                       e.name AS employee_name, e.code AS employee_code
                       FROM attendance a JOIN employees e ON a.employee_id = e.id
                       WHERE a.check_at >= ? AND a.check_at < ?
                       ORDER BY a.check_at ASC''',
                    (start, end)
                )
        elif from_date and to_date:
            if employee_id is not None:
                cur = conn.execute(
                    '''SELECT a.id, a.employee_id, a.check_type, a.check_at, a.image_path, a.latitude, a.longitude, a.reason, a.shift_id, a.status,
                       e.name AS employee_name, e.code AS employee_code
                       FROM attendance a JOIN employees e ON a.employee_id = e.id
                       WHERE a.employee_id = ? AND a.check_at >= ? AND a.check_at <= ?
                       ORDER BY a.check_at ASC''',
                    (employee_id, from_date, to_date)
                )
            else:
                cur = conn.execute(
                    '''SELECT a.id, a.employee_id, a.check_type, a.check_at, a.image_path, a.latitude, a.longitude, a.reason, a.shift_id, a.status,
                       e.name AS employee_name, e.code AS employee_code
                       FROM attendance a JOIN employees e ON a.employee_id = e.id
                       WHERE a.check_at >= ? AND a.check_at <= ?
                       ORDER BY a.check_at ASC''',
                    (from_date, to_date)
                )
        else:
            if employee_id is not None:
                cur = conn.execute(
                    '''SELECT a.id, a.employee_id, a.check_type, a.check_at, a.image_path, a.latitude, a.longitude, a.reason, a.shift_id, a.status,
                       e.name AS employee_name, e.code AS employee_code
                       FROM attendance a JOIN employees e ON a.employee_id = e.id
                       WHERE a.employee_id = ?
                       ORDER BY a.check_at DESC LIMIT 200''',
                    (employee_id,)
                )
            else:
                cur = conn.execute(
                    '''SELECT a.id, a.employee_id, a.check_type, a.check_at, a.image_path, a.latitude, a.longitude, a.reason, a.shift_id, a.status,
                       e.name AS employee_name, e.code AS employee_code
                       FROM attendance a JOIN employees e ON a.employee_id = e.id
                       ORDER BY a.check_at DESC LIMIT 200'''
                )
        rows = cur.fetchall()
        conn.close()
        list_ = [{
            'id': r['id'],
            'employee_id': r['employee_id'],
            'check_type': r['check_type'],
            'check_at': r['check_at'],
            'image_path': r['image_path'],
            'latitude': r['latitude'],
            'longitude': r['longitude'],
            'reason': r['reason'] or '',
            'shift_id': r['shift_id'],
            'status': r['status'] or 'approved',
            'employee_name': r['employee_name'],
            'employee_code': r['employee_code'],
        } for r in rows]
        return jsonify(list_)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/attendance/<int:att_id>', methods=['PATCH'])
@require_auth
def update_attendance_status(att_id):
    """Duyệt hoặc từ chối bản ghi chấm công.
    Body JSON: { status: 'approved' | 'rejected' }
    Chỉ ảnh hưởng 'outside': chỉ tính lương khi status='approved'.
    Với 'outside': gửi tin nhắn vào kênh Nhắn tin cho nhân viên (approved/rejected).
    """
    conn = None
    try:
        data = _get_request_data()
        status = (data.get('status') or '').strip().lower()
        if status not in ('approved', 'rejected', 'pending'):
            return jsonify({'error': 'status phải là approved, rejected hoặc pending'}), 400
        conn = get_db()
        row = conn.execute(
            'SELECT employee_id, check_type, check_at FROM attendance WHERE id = ?',
            (att_id,),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy bản ghi'}), 404
        conn.execute(
            'UPDATE attendance SET status = ? WHERE id = ?',
            (status, att_id),
        )
        conn.commit()
        if row['check_type'] == 'outside' and status in ('approved', 'rejected'):
            emp_id = row['employee_id']
            day_str = (row['check_at'] or '')[:16].replace('T', ' ')
            if status == 'rejected':
                msg = (
                    '[Thông báo] Bản ghi Ra ngoài của bạn '
                    f'({day_str}) đã bị từ chối — không tính công/lương.'
                )
            else:
                msg = (
                    '[Thông báo] Bản ghi Ra ngoài của bạn '
                    f'({day_str}) đã được duyệt.'
                )
            admin_user = getattr(g, 'current_user', None) or 'Quản trị'
            created_at = datetime.utcnow().isoformat() + 'Z'
            conn.execute(
                '''INSERT INTO chat_messages (
                       employee_id, sender_type, sender_username, message,
                       attachment_name, attachment_url, attachment_type, attachment_size,
                       is_read, created_at
                   ) VALUES (?, ?, ?, ?, NULL, NULL, NULL, 0, 0, ?)''',
                (emp_id, 'admin', admin_user, msg, created_at),
            )
            conn.commit()
        conn.close()
        return jsonify({'ok': True, 'status': status})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/attendance/pending-count', methods=['GET'])
@require_auth
def attendance_pending_count():
    """Số bản ghi chấm công 'outside' đang chờ duyệt (pending)."""
    conn = None
    try:
        conn = get_db()
        row = conn.execute(
            "SELECT COUNT(*) AS c FROM attendance WHERE check_type = 'outside' AND status = 'pending'"
        ).fetchone()
        conn.close()
        return jsonify({'pending_count': int(row['c'] or 0)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/payroll', methods=['GET'])
@require_auth
def payroll_report():
    """Bảng lương theo tháng từ lịch sử chấm công.

    Query: year, month, employee_id (optional)
    Tính lương theo chức vụ:
    - Mỗi ngày đi làm: lương = (base_salary/standard_hours*actual_hours) + allowance/standard_hours*actual_hours
    - OT = số giờ tăng ca * hourly_rate * overtime_multiplier.
    - Nghỉ có lương: payable_days += paid_leave_days (tính theo daily_wage).
    - Nếu không có chức vụ: dùng daily_wage / 8 giờ tiêu chuẩn.
    """
    conn = None
    try:
        year = request.args.get('year', type=int) or datetime.utcnow().year
        month = request.args.get('month', type=int) or datetime.utcnow().month
        requested_employee_id = request.args.get('employee_id', type=int)
        if month < 1 or month > 12:
            return jsonify({'error': 'month phải trong khoảng 1..12'}), 400

        month_start = f'{year}-{month:02d}-01'
        if month == 12:
            next_month_start = f'{year + 1}-01-01'
        else:
            next_month_start = f'{year}-{month + 1:02d}-01'

        conn = get_db()
        cur_user = conn.execute(
            'SELECT employee_id FROM users WHERE username = ?',
            (g.current_user,)
        ).fetchone()
        token_employee_id = cur_user['employee_id'] if cur_user and cur_user['employee_id'] else None
        if token_employee_id is not None and requested_employee_id is not None and int(requested_employee_id) != int(token_employee_id):
            conn.close()
            return jsonify({'error': 'Bạn không có quyền xem bảng lương của nhân viên khác'}), 403
        employee_id = requested_employee_id if requested_employee_id is not None else token_employee_id

        # Tải toàn bộ chính sách lương (ưu tiên) và chức vụ (fallback)
        salary_policies = {}
        for sp in conn.execute('SELECT id, name, code, pay_frequency, standard_work_days, standard_hours_per_day, overtime_multiplier FROM salary_policies').fetchall():
            salary_policies[sp['id']] = {
                'name': sp['name'],
                'code': sp['code'],
                'pay_frequency': sp['pay_frequency'] or 'monthly',
                'standard_work_days': float(sp['standard_work_days'] or 26),
                'standard_hours_per_day': float(sp['standard_hours_per_day'] or 8),
                'overtime_multiplier': float(sp['overtime_multiplier'] or 1.5),
            }

        positions = {}
        for p in conn.execute('SELECT id, name, base_salary, allowance, standard_hours, hourly_rate, overtime_multiplier FROM positions').fetchall():
            positions[p['id']] = {
                'name': p['name'],
                'base_salary': float(p['base_salary'] or 0),
                'allowance': float(p['allowance'] or 0),
                'standard_hours': float(p['standard_hours'] or 8),
                'hourly_rate': float(p['hourly_rate'] or 0),
                'overtime_multiplier': float(p['overtime_multiplier'] or 1.5),
            }

        # Tải thông tin nhân viên: position_id, salary_policy_id, daily_wage
        emp_sql = (
            'SELECT e.id, e.code, e.name, COALESCE(d.name, e.department) AS department, '
            'e.position_id, e.salary_policy_id, COALESCE(e.daily_wage, 0) AS daily_wage FROM employees e '
            'LEFT JOIN departments d ON e.department_id = d.id'
        )
        emp_args = []
        if employee_id is not None:
            emp_sql += ' WHERE e.id = ?'
            emp_args.append(employee_id)
        emp_sql += ' ORDER BY e.name'
        employees = conn.execute(emp_sql, emp_args).fetchall()

        def _get_salary_config(emp):
            """Lấy cấu hình lương: ưu tiên salary_policy > position > daily_wage fallback."""
            sp_id = emp['salary_policy_id']
            pos_id = emp['position_id']
            dw = float(emp['daily_wage'] or 0)

            # Ưu tiên 1: salary_policy
            if sp_id and sp_id in salary_policies:
                sp = salary_policies[sp_id]
                std_h = sp['standard_hours_per_day']
                monthly = dw if dw > 0 else 0
                return {
                    'source': 'salary_policy',
                    'policy_name': sp['name'],
                    'base_salary': monthly,
                    'allowance': 0,
                    'standard_hours': std_h,
                    'standard_work_days': sp['standard_work_days'],
                    'hourly_rate': monthly / (sp['standard_work_days'] * std_h) if monthly > 0 else 0,
                    'overtime_multiplier': sp['overtime_multiplier'],
                    'pay_frequency': sp['pay_frequency'],
                }

            # Ưu tiên 2: position
            if pos_id and pos_id in positions:
                p = positions[pos_id]
                return {
                    'source': 'position',
                    'position_name': p['name'],
                    'base_salary': p['base_salary'],
                    'allowance': p['allowance'],
                    'standard_hours': p['standard_hours'],
                    'standard_work_days': 26,
                    'hourly_rate': p['hourly_rate'],
                    'overtime_multiplier': p['overtime_multiplier'],
                    'pay_frequency': 'monthly',
                }

            # Fallback: dùng daily_wage
            return {
                'source': 'daily_wage',
                'base_salary': dw,
                'allowance': 0,
                'standard_hours': 8.0,
                'standard_work_days': 26,
                'hourly_rate': dw / 8.0 if dw > 0 else 0,
                'overtime_multiplier': 1.5,
                'pay_frequency': 'monthly',
            }

        results = []

        # Batch load all attendance records for all employees (N+1 fix)
        emp_ids = [e['id'] for e in employees]
        att_map = {eid: [] for eid in emp_ids}
        if emp_ids:
            placeholders = ','.join('?' * len(emp_ids))
            all_att = conn.execute(
                f'''SELECT employee_id, check_type, check_at, shift_id, status
                     FROM attendance
                     WHERE employee_id IN ({placeholders})
                       AND check_at >= ?
                       AND check_at < ?''',
                emp_ids + [month_start + 'T00:00:00', next_month_start + 'T00:00:00']
            ).fetchall()
            for r in all_att:
                att_map[r['employee_id']].append(r)

        # Batch load all leave requests for all employees (N+1 fix)
        leave_map = {eid: [] for eid in emp_ids}
        if emp_ids:
            placeholders = ','.join('?' * len(emp_ids))
            all_leaves = conn.execute(
                f'''SELECT employee_id, leave_type, from_date, to_date
                     FROM leave_requests
                     WHERE employee_id IN ({placeholders})
                       AND status = 'approved'
                       AND from_date <= ?
                       AND to_date >= ?''',
                emp_ids + [next_month_start, month_start]
            ).fetchall()
            for lr in all_leaves:
                leave_map[lr['employee_id']].append(lr)

        for e in employees:
            emp_id = e['id']
            salary_config = _get_salary_config(e)

            # Sử dụng dữ liệu đã batch load
            attendance_rows = att_map.get(emp_id, [])

            days = {}
            for r in attendance_rows:
                d = (r['check_at'] or '')[:10]
                if not d:
                    continue
                status = r['status'] or 'approved'
                if d not in days:
                    days[d] = {'in': 0, 'out': 0, 'outside_approved': 0, 'outside_pending': 0, 'outside_rejected': 0, 'in_times': [], 'out_times': []}
                if r['check_type'] == 'in':
                    days[d]['in'] += 1
                    t = (r['check_at'] or '')[11:16]
                    if t:
                        days[d]['in_times'].append(t)
                elif r['check_type'] == 'out':
                    days[d]['out'] += 1
                    t = (r['check_at'] or '')[11:16]
                    if t:
                        days[d]['out_times'].append(t)
                elif r['check_type'] == 'outside':
                    if status == 'approved':
                        days[d]['outside_approved'] += 1
                    elif status == 'rejected':
                        days[d]['outside_rejected'] += 1
                    else:
                        days[d]['outside_pending'] += 1

            # attended_days: ngày có ít nhất 1 bản ghi (chỉ approved mới tính)
            attended_days = len(days)
            complete_days = sum(1 for _, v in days.items()
                                if (v['in'] > 0 and v['out'] > 0) or v['outside_approved'] > 0)
            # pending outside: không tính là đủ công, không tính lương
            pending_outside_days = sum(1 for _, v in days.items() if v['outside_pending'] > 0)
            rejected_outside_days = sum(1 for _, v in days.items() if v['outside_rejected'] > 0)
            incomplete_days = max(0, attended_days - complete_days)

            # Tính lương theo salary_config
            base_pay = 0.0
            allowance_pay = 0.0
            ot_pay = 0.0
            std_hours = salary_config['standard_hours']
            base_sal = salary_config['base_salary']
            allowance = salary_config['allowance']
            hourly_rate = salary_config['hourly_rate']
            ot_mult = salary_config['overtime_multiplier']
            std_days = salary_config.get('standard_work_days', 26)

            # Tính lương theo giờ tiêu chuẩn của salary_policy/position
            hourly_base = base_sal / (std_days * std_hours) if std_days * std_hours > 0 else 0
            hourly_allowance = allowance / (std_days * std_hours) if std_days * std_hours > 0 else 0

            for day, v in days.items():
                if v['outside_approved'] > 0:
                    # Ra ngoài đã duyệt: tính theo ngày, không phải toàn bộ tháng
                    daily_rate = (base_sal + allowance) / std_days if std_days > 0 else 0
                    base_pay += (base_sal / std_days) if std_days > 0 else 0
                    allowance_pay += (allowance / std_days) if std_days > 0 else 0
                    # Không tính OT cho ngày outside_approved
                elif v['outside_pending'] > 0:
                    # Ra ngoài chưa duyệt: không tính lương
                    pass
                else:
                    # Tính giờ làm thực tế từ giờ vào/ra đầu tiên và cuối cùng
                    actual_hours = 0.0
                    if v['in_times'] and v['out_times']:
                        first_in = v['in_times'][0]
                        last_out = v['out_times'][-1]
                        try:
                            fmt = '%H:%M'
                            t_in = datetime.strptime(first_in, fmt)
                            t_out = datetime.strptime(last_out, fmt)
                            diff_h = (t_out - t_in).seconds / 3600.0
                            if diff_h < 0:
                                diff_h += 24.0
                            actual_hours = diff_h
                        except Exception:
                            actual_hours = std_hours
                    elif v['in'] > 0 or v['out'] > 0:
                        actual_hours = std_hours  # không đọc được giờ -> coi = tiêu chuẩn

                    # Lương cơ bản + phụ cấp theo giờ làm thực tế
                    base_pay += hourly_base * actual_hours
                    allowance_pay += hourly_allowance * actual_hours

                    # Tính OT nếu vượt giờ tiêu chuẩn
                    if actual_hours > std_hours:
                        ot_hours = actual_hours - std_hours
                        ot_pay += ot_hours * hourly_rate * ot_mult

            # Nghỉ có lương - sử dụng dữ liệu đã batch load
            leave_rows = leave_map.get(emp_id, [])

            paid_leave_days = 0
            unpaid_leave_days = 0
            paid_leave_pay = 0.0
            for lr in leave_rows:
                try:
                    leave_start = datetime.strptime((lr['from_date'] or '')[:10], '%Y-%m-%d').date()
                    leave_end = datetime.strptime((lr['to_date'] or '')[:10], '%Y-%m-%d').date()
                    period_start = datetime.strptime(month_start, '%Y-%m-%d').date()
                    period_end = datetime.strptime(next_month_start, '%Y-%m-%d').date() - timedelta(days=1)
                    overlap_start = max(leave_start, period_start)
                    overlap_end = min(leave_end, period_end)
                    if overlap_start <= overlap_end:
                        overlap_days = (overlap_end - overlap_start).days + 1
                        if lr['leave_type'] == 'paid':
                            paid_leave_days += overlap_days
                            # Nghỉ có lương: tính theo daily_wage, không phải full salary
                            daily_rate = (base_sal + allowance) / std_days if std_days > 0 else 0
                            paid_leave_pay += overlap_days * daily_rate
                        else:
                            unpaid_leave_days += overlap_days
                except Exception:
                    continue

            payable_days = attended_days + paid_leave_days
            net_salary = round(base_pay + allowance_pay + ot_pay + paid_leave_pay, 2)

            # Tổng giờ OT (để hiển thị)
            total_ot_hours = 0.0
            for day, v in days.items():
                if v['outside_approved'] > 0 or v['outside_pending'] > 0 or not v['in_times'] or not v['out_times']:
                    continue
                try:
                    fmt = '%H:%M'
                    t_in = datetime.strptime(v['in_times'][0], fmt)
                    t_out = datetime.strptime(v['out_times'][-1], fmt)
                    diff_h = (t_out - t_in).seconds / 3600.0
                    if diff_h < 0:
                        diff_h += 24.0
                    if diff_h > std_hours:
                        total_ot_hours += (diff_h - std_hours)
                except Exception:
                    pass

            # Lấy thông tin nguồn lương
            salary_source = salary_config.get('source', 'unknown')
            salary_source_name = salary_config.get('policy_name') or salary_config.get('position_name') or None

            results.append({
                'employee_id': emp_id,
                'employee_code': e['code'],
                'employee_name': e['name'],
                'department': e['department'],
                'salary_policy_id': e['salary_policy_id'],
                'salary_policy_name': salary_source_name,
                'salary_source': salary_source,
                'position_id': e['position_id'],
                'position_name': salary_config.get('position_name'),
                'daily_wage': float(e['daily_wage'] or 0),
                'base_salary': base_sal,
                'allowance': allowance,
                'standard_hours': std_hours,
                'hourly_rate': hourly_rate,
                'attended_days': attended_days,
                'complete_days': complete_days,
                'incomplete_days': incomplete_days,
                'pending_outside_days': pending_outside_days,
                'rejected_outside_days': rejected_outside_days,
                'paid_leave_days': paid_leave_days,
                'unpaid_leave_days': unpaid_leave_days,
                'payable_days': payable_days,
                'base_pay': round(base_pay, 2),
                'allowance_pay': round(allowance_pay, 2),
                'ot_hours': round(total_ot_hours, 2),
                'ot_pay': round(ot_pay, 2),
                'paid_leave_pay': round(paid_leave_pay, 2),
                'net_salary': net_salary,
            })

        conn.close()
        total_net = round(sum(item['net_salary'] for item in results), 2)
        return jsonify({
            'year': year,
            'month': month,
            'employee_count': len(results),
            'total_net_salary': total_net,
            'items': results,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/change-password', methods=['POST'])
@require_auth
def change_password():
    """Đổi mật khẩu admin đang đăng nhập. Kiểm tra cả users và admin_users. Body: { current_password, new_password }."""
    conn = None
    try:
        data = _get_request_data()
        current_password = data.get('current_password') or ''
        new_password = data.get('new_password') or ''
        if not current_password or not new_password:
            return jsonify({'error': 'Thiếu mật khẩu hiện tại hoặc mật khẩu mới'}), 400
        ok, msg = validate_password(new_password)
        if not ok:
            return jsonify({'error': msg}), 400
        conn = get_db()
        username = g.current_user

        # Kiểm tra bảng admin_users trước (ưu tiên)
        cur = conn.execute('SELECT id, password_hash FROM admin_users WHERE username = ?', (username,))
        row = cur.fetchone()
        if row:
            if not check_password_hash(row['password_hash'], current_password):
                conn.close()
                return jsonify({'error': 'Mật khẩu hiện tại không đúng'}), 400
            new_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
            conn.execute('UPDATE admin_users SET password_hash = ? WHERE username = ?', (new_hash, username))
            conn.commit()
            conn.close()
            return jsonify({'ok': True, 'message': 'Đã đổi mật khẩu thành công.'})

        # Fallback: bảng users cũ
        cur = conn.execute('SELECT id, password_hash FROM users WHERE username = ?', (username,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Không tìm thấy tài khoản'}), 404
        if not check_password_hash(row['password_hash'], current_password):
            conn.close()
            return jsonify({'error': 'Mật khẩu hiện tại không đúng'}), 400
        new_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
        conn.execute('UPDATE users SET password_hash = ? WHERE username = ?', (new_hash, username))
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'message': 'Đã đổi mật khẩu thành công.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/attendance/daily-stats', methods=['GET'])
@require_auth
def attendance_daily_stats():
    """Thống kê chấm công hôm nay: tổng check-in, đi muộn (so với ca), OT, chưa check-in.
    Query: date (YYYY-MM-DD, mặc định hôm nay)."""
    conn = None
    try:
        date_str = request.args.get('date', '').strip()
        if not date_str:
            date_str = datetime.utcnow().strftime('%Y-%m-%d')
        from_ = date_str + 'T00:00:00'
        to_ = date_str + 'T23:59:59'

        conn = get_db()
        att_rows = conn.execute(
            '''SELECT a.employee_id, a.check_type, a.check_at, e.shift_start, e.shift_end
               FROM attendance a JOIN employees e ON a.employee_id = e.id
               WHERE a.check_at >= ? AND a.check_at < ? AND e.status = 'active' ''',
            (from_, to_)
        ).fetchall()

        att_by_emp = {}
        for r in att_rows:
            eid = r['employee_id']
            if eid not in att_by_emp:
                att_by_emp[eid] = {'ins': [], 'outs': []}
            check_time = r['check_at'] or ''
            ct_short = check_time[11:16] if len(check_time) >= 16 else ''
            if r['check_type'] == 'in':
                att_by_emp[eid]['ins'].append(ct_short)
            else:
                att_by_emp[eid]['outs'].append(ct_short)

        total_checked = len(att_by_emp)
        late_count = 0
        ot_count = 0

        # Batch query shifts for all employees to avoid N+1
        emp_ids = list(att_by_emp.keys())
        shifts_map = {}
        if emp_ids:
            placeholders = ','.join('?' * len(emp_ids))
            shift_rows = conn.execute(
                f'SELECT e.id AS emp_id, s.shift_start, s.shift_end FROM employees e LEFT JOIN shifts s ON e.shift_id = s.id WHERE e.id IN ({placeholders})',
                emp_ids
            ).fetchall()
            for sr in shift_rows:
                shifts_map[sr['emp_id']] = (sr['shift_start'], sr['shift_end'])

        for eid, data in att_by_emp.items():
            first_in = (data['ins'] or [''])[0]
            last_out = (data['outs'] or [''])[-1] if data['outs'] else ''
            shift_start = None
            shift_end = None
            if eid in shifts_map:
                shift_start, shift_end = shifts_map[eid]
                if shift_start:
                    shift_start = shift_start[:5]
                if shift_end:
                    shift_end = shift_end[:5]

            # Use LATE_THRESHOLD from config instead of hardcoded
            if shift_start and first_in and first_in > shift_start:
                try:
                    fmt = '%H:%M'
                    t_in = datetime.strptime(first_in, fmt)
                    t_shift = datetime.strptime(shift_start, fmt)
                    diff_minutes = (t_in - t_shift).total_seconds() / 60.0
                    if diff_minutes > LATE_THRESHOLD:
                        late_count += 1
                except Exception:
                    # Fallback: string comparison
                    if first_in > shift_start:
                        late_count += 1
            if shift_end and last_out and last_out > shift_end:
                try:
                    fmt = '%H:%M'
                    t_out = datetime.strptime(last_out, fmt)
                    t_shift_end = datetime.strptime(shift_end, fmt)
                    diff_minutes = (t_out - t_shift_end).total_seconds() / 60.0
                    if diff_minutes > LATE_THRESHOLD:
                        ot_count += 1
                except Exception:
                    if last_out > shift_end:
                        ot_count += 1

        active_emps = conn.execute(
            "SELECT COUNT(*) FROM employees WHERE status = 'active'"
        ).fetchone()[0] or 0
        not_checked = max(0, active_emps - total_checked)
        conn.close()

        return jsonify({
            'date': date_str,
            'total_checked': total_checked,
            'late_count': late_count,
            'ot_count': ot_count,
            'not_checked': not_checked,
            'active_employees': active_emps,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


@app.route('/api/backup', methods=['GET'])
@require_auth
def backup_database():
    """Tải file .db về (backup)."""
    try:
        if not os.path.exists(DB_PATH):
            return jsonify({'error': 'File database không tồn tại'}), 404
        return send_file(DB_PATH, as_attachment=True, download_name='attendance_backup.db')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/restore', methods=['POST'])
@require_auth
def restore_database():
    """Upload file .db để khôi phục. Tự động backup trước khi restore."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename or not f.filename.endswith('.db'):
            return jsonify({'error': 'File phải có đuôi .db'}), 400

        # Tự động backup database hiện tại trước khi restore
        import shutil as _shutil
        if os.path.exists(DB_PATH):
            backup_path = DB_PATH + f'.bak.{int(datetime.utcnow().timestamp())}'
            _shutil.copy2(DB_PATH, backup_path)

        # Verify uploaded file is valid SQLite before restoring
        try:
            import sqlite3 as _sqlite3
            f.seek(0)
            test_conn = _sqlite3.connect(':memory:')
            test_conn.close()
            f.seek(0)
        except Exception:
            return jsonify({'error': 'File không phải SQLite database hợp lệ'}), 400

        # Restore
        f.save(DB_PATH)

        # Verify restored file has required tables
        try:
            conn_verify = _sqlite3.connect(DB_PATH)
            cur = conn_verify.execute("SELECT name FROM sqlite_master WHERE type='table' LIMIT 5")
            tables = [r[0] for r in cur.fetchall()]
            conn_verify.close()
            required = {'employees', 'users', 'attendance', 'admin_users'}
            missing = required - set(tables)
            if missing:
                raise ValueError(f'Thiếu bảng: {missing}')
        except Exception as verify_err:
            # Restore failed - restore from backup
            if 'backup_path' in dir() and os.path.exists(backup_path):
                _shutil.copy2(backup_path, DB_PATH)
            return jsonify({'error': f'File restore không hợp lệ: {verify_err}'}), 400

        return jsonify({
            'ok': True,
            'message': 'Đã khôi phục database thành công. Làm mới trang để cập nhật.',
            'backup_created': 'backup_path' in dir()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Google Sheets sync (lazy import) ─────────────────────────────────────────
HAS_GOOGLE_API = False
try:
    from google.oauth2 import service_account
    import googleapiclient.discovery
    import googleapiclient.http
    HAS_GOOGLE_API = True
except ImportError:
    pass


def _get_google_services():
    """Trả về (sheets_svc, drive_svc) hoặc (None, None) nếu thiếu credentials hoặc thư viện Google."""
    if not HAS_GOOGLE_API:
        return None, None
    creds_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'google_credentials.json')
    if not os.path.isfile(creds_file):
        return None, None
    SCOPES = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive.readonly',
    ]
    creds = service_account.Credentials.from_service_account_file(creds_file, scopes=SCOPES)
    sheets = googleapiclient.discovery.build('sheets', 'v4', credentials=creds)
    drive  = googleapiclient.discovery.build('drive',  'v3', credentials=creds)
    return sheets, drive


def _extract_drive_file_id(url: str):
    import re
    for p in [r'/file/d/([a-zA-Z0-9_-]{10,})', r'id=([a-zA-Z0-9_-]{10,})',
              r'([a-zA-Z0-9_-]{10,})[?/&"\'\s]']:
        m = re.search(p, url)
        if m:
            return m.group(1)
    return None


def _compute_embedding_from_bytes(images_bytes, max_images=5):
    """Tính embedding trung bình từ danh sách bytes ảnh."""
    if not images_bytes:
        return None
    embeddings = []
    for b in images_bytes[:max_images]:
        emb = compute_embedding_from_images(
            [_save_temp_image(b)], max_images=1)
        if emb and len(emb) == 512:
            embeddings.append(np.array(emb, dtype=np.float32))
    if not embeddings:
        return None
    mean_emb = np.mean(embeddings, axis=0)
    norm = np.linalg.norm(mean_emb)
    if norm > 0:
        mean_emb = mean_emb / norm
    return mean_emb.tolist()


def _save_temp_image(image_bytes):
    import tempfile as _tf
    tmp = _tf.NamedTemporaryFile(suffix='.jpg', delete=False)
    tmp.write(image_bytes)
    tmp.flush()
    tmp.close()
    return tmp.name


SHEET_HEADERS = [
    'code', 'name', 'department', 'position', 'shift',
    'email', 'phone', 'birth_date', 'daily_wage', 'photo_url',
]


@app.route('/api/sync-from-sheet', methods=['POST'])
@require_auth
def api_sync_from_sheet():
    """
    Đồng bộ nhân viên từ Google Sheet.
    Body JSON (optional):
      sheet_id   – Google Sheet ID (bắt buộc nếu chưa có .env SHEET_ID)
      dry_run    – True: xem trước, không lưu gì
      force_face – True: cập nhật lại embedding dù đã có
    Trả:
      {ok, created, updated, skipped, errors, rows: [...]}
    """
    data = _get_request_data()
    dry_run    = bool(data.get('dry_run', False))
    force_face = bool(data.get('force_face', False))
    sheet_id   = data.get('sheet_id') or os.environ.get('SHEET_ID')

    if not sheet_id:
        return jsonify({'error': 'Thiếu sheet_id'}), 400

    sheets_svc, drive_svc = _get_google_services()

    # ── Đọc Sheet ───────────────────────────────────────────────────────────
    if sheets_svc:
        result = sheets_svc.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range='Employees!A:J',
        ).execute()
        raw_rows = result.get('values', [])
    else:
        # Public CSV fallback
        url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid=0'
        with urllib.request.urlopen(url, timeout=20) as resp:
            import csv, io as _io
            reader = csv.reader(_io.StringIO(resp.read().decode('utf-8-sig')))
            raw_rows = list(reader)

    if not raw_rows:
        return jsonify({'ok': True, 'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0, 'rows': []})

    # Parse
    records = []
    for row in raw_rows:
        if not row or not row[0].strip() or row[0].strip().lower() == 'code':
            continue
        rec = {h: (row[i].strip() if i < len(row) else '') for i, h in enumerate(SHEET_HEADERS)}
        records.append(rec)

    created = updated = skipped = errors = 0
    results = []

    # Dùng 1 connection cho toàn bộ sync (tránh mở/đóng liên tục)
    conn = None
    conn = get_db()

    for rec in records:
        code = rec.get('code', '').strip()
        name = rec.get('name', '').strip()
        if not code or not name:
            errors += 1
            results.append({'code': code, 'name': name, 'status': 'error', 'msg': 'Thiếu code hoặc name'})
            continue

        cur  = conn.cursor()

        # Resolve FKs
        def _get_or_create(tbl, name_col, val):
            if not val:
                return None
            val = val.strip()
            cur.execute(f"SELECT id FROM {tbl} WHERE {name_col} = ?", (val,))
            r = cur.fetchone()
            if r:
                return r['id']
            cur.execute(f"INSERT INTO {tbl} (name) VALUES (?)", (val,))
            return cur.lastrowid

        dept_id     = _get_or_create('departments', 'name', rec.get('department', ''))
        position_id = _get_or_create('positions',   'name', rec.get('position',   ''))
        shift_id    = _get_or_create('shifts',       'name', rec.get('shift',      ''))

        cur.execute("SELECT id, embedding FROM employees WHERE code = ?", (code,))
        existing = cur.fetchone()

        # ── Tải ảnh + embedding ─────────────────────────────────────────────
        embedding = None
        photo_url  = rec.get('photo_url', '').strip()
        if photo_url:
            img_bytes = None
            if drive_svc:
                fid = _extract_drive_file_id(photo_url)
                if fid:
                    try:
                        req  = drive_svc.files().get_media(fileId=fid)
                        fh   = io.BytesIO()
                        dler = googleapiclient.http.MediaIoBaseDownload(fh, req)
                        done = False
                        while not done:
                            _, done = dler.next_chunk()
                        img_bytes = fh.getvalue()
                    except Exception:
                        img_bytes = None
            else:
                fid = _extract_drive_file_id(photo_url)
                if fid:
                    pub_url = f"https://drive.google.com/uc?export=view&id={fid}"
                    try:
                        req = urllib.request.Request(pub_url, headers={'User-Agent': 'Mozilla/5.0'})
                        with urllib.request.urlopen(req, timeout=15) as resp2:
                            img_bytes = resp2.read()
                    except Exception:
                        img_bytes = None

            if img_bytes:
                embedding = _compute_embedding_from_bytes([img_bytes], max_images=1)

        if dry_run:
            conn.close()
            status = 'ok' if embedding else 'no_face'
            skipped += 1
            results.append({'code': code, 'name': name, 'status': status,
                            'msg': 'dry-run', 'has_face': bool(embedding)})
            continue

        now = datetime.now().isoformat()

        if existing:
            emp_id   = existing['id']
            old_emb  = existing['embedding']
            need_emb = (force_face or not old_emb or old_emb == 'null') and embedding
            fields   = {
                'name': name,
                'department': rec.get('department', ''),
                'department_id': dept_id,
                'position_id': position_id,
                'shift_id': shift_id,
                'email': rec.get('email', ''),
                'phone': rec.get('phone', ''),
                'birth_date': rec.get('birth_date', ''),
                'daily_wage': float(rec.get('daily_wage') or 0),
            }
            if need_emb:
                fields['embedding'] = json.dumps(embedding)
            set_cl  = ', '.join(f"{k} = ?" for k in fields)
            vals    = list(fields.values()) + [emp_id]
            cur.execute(f"UPDATE employees SET {set_cl} WHERE id = ?", vals)
            conn.commit()
            updated += 1
            results.append({'code': code, 'name': name, 'status': 'updated',
                            'msg': f'id={emp_id}', 'has_face': bool(embedding)})
        else:
            fields = {
                'code': code, 'name': name,
                'department': rec.get('department', ''),
                'department_id': dept_id,
                'position_id': position_id,
                'shift_id': shift_id,
                'allowed_checkin': 1,
                'email': rec.get('email', ''),
                'phone': rec.get('phone', ''),
                'birth_date': rec.get('birth_date', ''),
                'daily_wage': float(rec.get('daily_wage') or 0),
                'status': 'active', 'created_at': now,
            }
            if embedding:
                fields['embedding'] = json.dumps(embedding)
            cols = ', '.join(fields.keys())
            ph   = ', '.join('?' * len(fields))
            cur.execute(f"INSERT INTO employees ({cols}) VALUES ({ph})",
                        list(fields.values()))
            conn.commit()
            emp_id = cur.lastrowid
            created += 1
            results.append({'code': code, 'name': name, 'status': 'created',
                            'msg': f'id={emp_id}', 'has_face': bool(embedding)})

    conn.close()

    return jsonify({
        'ok': True,
        'dry_run': dry_run,
        'sheet_id': sheet_id,
        'total_rows': len(records),
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors,
        'rows': results,
    })


@app.route('/api/sync-from-sheet/config', methods=['GET', 'POST'])
@require_auth
def api_sync_config():
    """Lưu / đọc SHEET_ID vào .env."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')

    if request.method == 'GET':
        current = os.environ.get('SHEET_ID', '')
        return jsonify({'sheet_id': current, 'has_credentials': os.path.isfile(
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'google_credentials.json'))})

    data     = _get_request_data()
    sheet_id = (data.get('sheet_id') or '').strip()
    if not sheet_id:
        return jsonify({'error': 'Thiếu sheet_id'}), 400

    lines = []
    if os.path.isfile(env_path):
        with open(env_path, encoding='utf-8') as f:
            lines = f.readlines()

    new_lines = []
    found = False
    for line in lines:
        if line.strip().startswith('SHEET_ID='):
            new_lines.append(f'SHEET_ID={sheet_id}\n')
            found = True
        else:
            new_lines.append(line)
    if not found:
        new_lines.append(f'SHEET_ID={sheet_id}\n')

    with open(env_path, 'w', encoding='utf-8') as f:
        f.writelines(new_lines)
    os.environ['SHEET_ID'] = sheet_id
    return jsonify({'ok': True, 'sheet_id': sheet_id})


# ══════════════════════════════════════════════════════════════════════════════
#  Chat upload
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/uploads/chat/<path:filename>')
def serve_chat_upload(filename):
    return send_from_directory(CHAT_UPLOAD_FOLDER, filename)


# ════════════════════════════════════════════════════════════════════════════════
# WORK AREAS (Khu vực làm việc)
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/work-areas', methods=['GET'])
@require_auth
def list_work_areas():
    """Lấy danh sách khu vực làm việc."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT id, name, code, description, is_active, created_at FROM work_areas ORDER BY name')
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'name': r['name'],
            'code': r['code'],
            'description': r['description'],
            'is_active': bool(r['is_active']),
            'created_at': r['created_at'],
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/work-areas', methods=['POST'])
@require_auth
def add_work_area():
    """Tạo khu vực làm việc mới."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        code = (data.get('code') or '').strip() or None
        description = (data.get('description') or '').strip() or None
        if not name:
            return jsonify({'error': 'Thiếu tên khu vực làm việc'}), 400
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        try:
            conn.execute('INSERT INTO work_areas (name, code, description, created_at) VALUES (?, ?, ?, ?)',
                        (name, code, description, created_at))
            conn.commit()
            row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
            conn.close()
            return jsonify({
                'id': row_id, 'name': name, 'code': code, 'description': description,
                'is_active': True, 'created_at': created_at,
            }), 201
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': 'Tên khu vực làm việc đã tồn tại'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/work-areas/<int:waid>', methods=['PATCH'])
@require_auth
def update_work_area(waid):
    """Cập nhật khu vực làm việc."""
    conn = None
    try:
        data = _get_request_data()
        updates, args = [], []
        for field, key in [('name', 'name'), ('code', 'code'), ('description', 'description')]:
            if key in data:
                updates.append(f'{field} = ?')
                args.append((data[key] or '').strip() or None)
        if 'is_active' in data:
            updates.append('is_active = ?')
            args.append(1 if data['is_active'] else 0)
        if not updates:
            return jsonify({'error': 'Không có trường cập nhật'}), 400
        args.append(waid)
        conn = get_db()
        cur = conn.execute('UPDATE work_areas SET ' + ', '.join(updates) + ' WHERE id = ?', args)
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy khu vực làm việc'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/work-areas/<int:waid>', methods=['DELETE'])
@require_auth
def delete_work_area(waid):
    """Xóa khu vực làm việc."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('DELETE FROM work_areas WHERE id = ?', (waid,))
        conn.commit()
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Không tìm thấy khu vực làm việc'}), 404
        count = conn.execute('SELECT COUNT(*) FROM work_areas').fetchone()[0]
        if count == 0:
            conn.execute('DELETE FROM sqlite_sequence WHERE name = ?', ('work_areas',))
            conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/work-areas/import', methods=['POST'])
@require_auth
def import_work_areas():
    """Import khu vực làm việc từ file Excel.

    File Excel format:
      name,code,description
      Kỹ thuật,KT,Khối kỹ thuật
      Kinh doanh,KD,Khối kinh doanh
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel (.xlsx, .xls)'}), 400

        content = f.read()
        import io as _io
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(content))
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return jsonify({'error': 'File trống'}), 400

        header = None
        rows_data = []
        for i, row in enumerate(rows):
            parts = [str(c).strip() if c is not None else '' for c in row]

            if i == 0:
                header = [h.lower().strip() for h in parts]
                if 'name' not in header:
                    return jsonify({'error': 'File Excel thiếu cột "name"'}), 400
                continue
            if not parts or not any(parts):
                continue
            row_dict = dict(zip(header, parts))
            rows_data.append(row_dict)

        if not rows_data:
            return jsonify({'error': 'Không có dữ liệu hợp lệ trong file'}), 400

        conn = None
        conn = get_db()
        created = 0
        updated = 0
        errors = []
        created_at = datetime.utcnow().isoformat() + 'Z'

        for i, row in enumerate(rows_data, start=2):
            name = (row.get('name') or '').strip()
            if not name:
                errors.append(f'Dòng {i}: thiếu tên khu vực làm việc')
                continue
            code = (row.get('code') or '').strip() or None
            description = (row.get('description') or '').strip() or None

            existing = conn.execute('SELECT id FROM work_areas WHERE name = ?', (name,)).fetchone()
            if existing:
                if description is not None or 'description' in header:
                    conn.execute('UPDATE work_areas SET code = ?, description = ? WHERE name = ?', (code, description, name))
                updated += 1
            else:
                conn.execute('INSERT INTO work_areas (name, code, description, created_at) VALUES (?, ?, ?, ?)',
                           (name, code, description, created_at))
                created += 1

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'created': created,
            'updated': updated,
            'errors': errors,
            'total': len(rows_data),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════════
# POSITION TYPES (Loại chức vụ: Director, Manager, Staff...)
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/position-types', methods=['GET'])
@require_auth
def list_position_types():
    """Lấy danh sách loại chức vụ."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT id, name, code, description, is_active, created_at FROM position_types ORDER BY name')
        rows = cur.fetchall()
        conn.close()
        return jsonify([{
            'id': r['id'],
            'name': r['name'],
            'code': r['code'],
            'description': r['description'],
            'is_active': bool(r['is_active']),
            'created_at': r['created_at'],
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/position-types', methods=['POST'])
@require_auth
def add_position_type():
    """Tạo loại chức vụ mới."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        code = (data.get('code') or '').strip() or None
        description = (data.get('description') or '').strip() or None
        if not name:
            return jsonify({'error': 'Thiếu tên loại chức vụ'}), 400
        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = get_db()
        try:
            conn.execute('INSERT INTO position_types (name, code, description, created_at) VALUES (?, ?, ?, ?)',
                        (name, code, description, created_at))
            conn.commit()
            row_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
            conn.close()
            return jsonify({
                'id': row_id, 'name': name, 'code': code, 'description': description,
                'is_active': True, 'created_at': created_at,
            }), 201
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': 'Tên loại chức vụ đã tồn tại'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/position-types/<int:ptid>', methods=['PATCH'])
@require_auth
def update_position_type(ptid):
    """Cập nhật loại chức vụ."""
    conn = None
    try:
        data = _get_request_data()
        updates, args = [], []
        for field, key in [('name', 'name'), ('code', 'code'), ('description', 'description')]:
            if key in data:
                updates.append(f'{field} = ?')
                args.append((data[key] or '').strip() or None)
        if 'is_active' in data:
            updates.append('is_active = ?')
            args.append(1 if data['is_active'] else 0)
        if not updates:
            return jsonify({'error': 'Không có trường cập nhật'}), 400
        args.append(ptid)
        conn = get_db()
        cur = conn.execute('UPDATE position_types SET ' + ', '.join(updates) + ' WHERE id = ?', args)
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy loại chức vụ'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/position-types/<int:ptid>', methods=['DELETE'])
@require_auth
def delete_position_type(ptid):
    """Xóa loại chức vụ."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('DELETE FROM position_types WHERE id = ?', (ptid,))
        conn.commit()
        if cur.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Không tìm thấy loại chức vụ'}), 404
        count = conn.execute('SELECT COUNT(*) FROM position_types').fetchone()[0]
        if count == 0:
            conn.execute('DELETE FROM sqlite_sequence WHERE name = ?', ('position_types',))
            conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/position-types/import', methods=['POST'])
@require_auth
def import_position_types():
    """Import loại chức vụ từ file Excel.

    File Excel format:
      name,code,description
      Giám đốc,GD,Quản lý cao cấp
      Trưởng phòng,TP,Quản lý cấp trung
      Nhân viên,NV,Nhân viên thường
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel (.xlsx, .xls)'}), 400

        content = f.read()
        import io as _io
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(content))
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return jsonify({'error': 'File trống'}), 400

        header = None
        rows_data = []
        for i, row in enumerate(rows):
            parts = [str(c).strip() if c is not None else '' for c in row]

            if i == 0:
                header = [h.lower().strip() for h in parts]
                if 'name' not in header:
                    return jsonify({'error': 'File Excel thiếu cột "name"'}), 400
                continue
            if not parts or not any(parts):
                continue
            row_dict = dict(zip(header, parts))
            rows_data.append(row_dict)

        if not rows_data:
            return jsonify({'error': 'Không có dữ liệu hợp lệ trong file'}), 400

        conn = None
        conn = get_db()
        created = 0
        updated = 0
        errors = []
        created_at = datetime.utcnow().isoformat() + 'Z'

        for i, row in enumerate(rows_data, start=2):
            name = (row.get('name') or '').strip()
            if not name:
                errors.append(f'Dòng {i}: thiếu tên loại chức vụ')
                continue
            code = (row.get('code') or '').strip() or None
            description = (row.get('description') or '').strip() or None

            existing = conn.execute('SELECT id FROM position_types WHERE name = ?', (name,)).fetchone()
            if existing:
                if description is not None or 'description' in header:
                    conn.execute('UPDATE position_types SET code = ?, description = ? WHERE name = ?', (code, description, name))
                updated += 1
            else:
                conn.execute('INSERT INTO position_types (name, code, description, created_at) VALUES (?, ?, ?, ?)',
                           (name, code, description, created_at))
                created += 1

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'created': created,
            'updated': updated,
            'errors': errors,
            'total': len(rows_data),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════════════
# TIMESHEETS (Bảng chấm công)
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/timesheets', methods=['GET'])
@require_auth
def list_timesheets():
    """Lấy danh sách bảng chấm công với chi tiết từng ngày."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT id, name, code, description, work_type, is_active, created_at FROM timesheets ORDER BY name')
        rows = cur.fetchall()
        timesheets = []
        for r in rows:
            # Lấy chi tiết các ca trong ngày
            details_cur = conn.execute(
                '''SELECT id, timesheet_id, shift_name, day_of_week, is_working_day,
                          check_in_start, check_in_end, check_out_start, check_out_end,
                          shift_order, work_hours, is_day_off
                   FROM timesheet_details WHERE timesheet_id = ? ORDER BY shift_order, day_of_week''',
                (r['id'],)
            )
            details = [{
                'id': d['id'],
                'timesheet_id': d['timesheet_id'],
                'shift_name': d['shift_name'],
                'day_of_week': d['day_of_week'],
                'is_working_day': bool(d['is_working_day']),
                'check_in_start': d['check_in_start'],
                'check_in_end': d['check_in_end'],
                'check_out_start': d['check_out_start'],
                'check_out_end': d['check_out_end'],
                'shift_order': d['shift_order'],
                'work_hours': float(d['work_hours'] or 8),
                'is_day_off': bool(d['is_day_off']),
            } for d in details_cur.fetchall()]
            timesheets.append({
                'id': r['id'],
                'name': r['name'],
                'code': r['code'],
                'description': r['description'],
                'work_type': r['work_type'],
                'is_active': bool(r['is_active']),
                'created_at': r['created_at'],
                'details': details,
            })
        conn.close()
        return jsonify(timesheets)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/timesheets/<int:tsid>', methods=['GET'])
@require_auth
def get_timesheet(tsid):
    """Lấy chi tiết một bảng chấm công."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('SELECT * FROM timesheets WHERE id = ?', (tsid,))
        r = cur.fetchone()
        if not r:
            conn.close()
            return jsonify({'error': 'Không tìm thấy bảng chấm công'}), 404
        details_cur = conn.execute(
            '''SELECT id, timesheet_id, shift_name, day_of_week, is_working_day,
                      check_in_start, check_in_end, check_out_start, check_out_end,
                      shift_order, work_hours, is_day_off
               FROM timesheet_details WHERE timesheet_id = ? ORDER BY shift_order, day_of_week''',
            (tsid,)
        )
        details = [{
            'id': d['id'],
            'timesheet_id': d['timesheet_id'],
            'shift_name': d['shift_name'],
            'day_of_week': d['day_of_week'],
            'is_working_day': bool(d['is_working_day']),
            'check_in_start': d['check_in_start'],
            'check_in_end': d['check_in_end'],
            'check_out_start': d['check_out_start'],
            'check_out_end': d['check_out_end'],
            'shift_order': d['shift_order'],
            'work_hours': float(d['work_hours'] or 8),
            'is_day_off': bool(d['is_day_off']),
        } for d in details_cur.fetchall()]
        conn.close()
        return jsonify({
            'id': r['id'],
            'name': r['name'],
            'code': r['code'],
            'description': r['description'],
            'work_type': r['work_type'],
            'is_active': bool(r['is_active']),
            'created_at': r['created_at'],
            'details': details,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/timesheets', methods=['POST'])
@require_auth
def add_timesheet():
    """Tạo bảng chấm công mới với chi tiết các ca.

    Body JSON:
      name: tên bảng chấm công (bắt buộc)
      code: mã (tùy chọn)
      description: mô tả
      work_type: 'single' | 'double' (1 ca hoặc 2 ca sáng+chiều)
      details: array của {
        shift_name: tên ca ("Ca sáng", "Ca chiều"),
        day_of_week: "mon,tue,wed,thu,fri,sat,sun" hoặc "mon fri",
        is_working_day: 1|0,
        check_in_start: "08:00",
        check_in_end: "08:30",
        check_out_start: "12:00",
        check_out_end: "17:00",
        shift_order: 1|2,
        work_hours: 8,
        is_day_off: 0|1
      }
    """
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        code = (data.get('code') or '').strip() or None
        description = (data.get('description') or '').strip() or None
        work_type = (data.get('work_type') or 'single').strip().lower()
        if work_type not in ('single', 'double'):
            work_type = 'single'
        if not name:
            return jsonify({'error': 'Thiếu tên bảng chấm công'}), 400

        details = data.get('details', [])
        if not details:
            return jsonify({'error': 'Thiếu chi tiết bảng chấm công (details)'}), 400

        created_at = datetime.utcnow().isoformat() + 'Z'
        conn = None
        conn = get_db()
        try:
            conn.execute(
                'INSERT INTO timesheets (name, code, description, work_type, created_at) VALUES (?, ?, ?, ?, ?)',
                (name, code, description, work_type, created_at)
            )
            ts_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]

            # Thêm chi tiết các ca
            for d in details:
                shift_name = (d.get('shift_name') or '').strip()
                day_of_week = (d.get('day_of_week') or 'mon,tue,wed,thu,fri').strip().lower()
                is_working = 1 if d.get('is_working_day', 1) else 0
                check_in_start = (d.get('check_in_start') or '08:00')[:5]
                check_in_end = (d.get('check_in_end') or '09:00')[:5]
                check_out_start = (d.get('check_out_start') or '12:00')[:5]
                check_out_end = (d.get('check_out_end') or '18:00')[:5]
                shift_order = int(d.get('shift_order') or 1)
                work_hours = float(d.get('work_hours') or 8)
                is_day_off = 1 if d.get('is_day_off', 0) else 0

                conn.execute(
                    '''INSERT INTO timesheet_details
                       (timesheet_id, shift_name, day_of_week, is_working_day,
                        check_in_start, check_in_end, check_out_start, check_out_end,
                        shift_order, work_hours, is_day_off, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (ts_id, shift_name, day_of_week, is_working,
                     check_in_start, check_in_end, check_out_start, check_out_end,
                     shift_order, work_hours, is_day_off, created_at)
                )

            conn.commit()
            conn.close()
            return jsonify({
                'id': ts_id,
                'name': name,
                'code': code,
                'description': description,
                'work_type': work_type,
                'is_active': True,
                'created_at': created_at,
            }), 201
        except Exception as ex:
            conn.rollback()
            conn.close()
            return jsonify({'error': str(ex)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/timesheets/<int:tsid>', methods=['PUT'])
@require_auth
def update_timesheet(tsid):
    """Cập nhật bảng chấm công và chi tiết."""
    conn = None
    try:
        data = _get_request_data()
        name = (data.get('name') or '').strip()
        code = (data.get('code') or '').strip() or None
        description = (data.get('description') or '').strip() or None
        work_type = (data.get('work_type') or 'single').strip().lower()
        if work_type not in ('single', 'double'):
            work_type = 'single'
        if not name:
            return jsonify({'error': 'Thiếu tên bảng chấm công'}), 400

        conn = get_db()
        cur = conn.execute('SELECT id FROM timesheets WHERE id = ?', (tsid,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'error': 'Không tìm thấy bảng chấm công'}), 404

        conn.execute(
            'UPDATE timesheets SET name = ?, code = ?, description = ?, work_type = ? WHERE id = ?',
            (name, code, description, work_type, tsid)
        )

        # Xóa chi tiết cũ và thêm mới
        conn.execute('DELETE FROM timesheet_details WHERE timesheet_id = ?', (tsid,))

        details = data.get('details', [])
        for d in details:
            shift_name = (d.get('shift_name') or '').strip()
            day_of_week = (d.get('day_of_week') or 'mon,tue,wed,thu,fri').strip().lower()
            is_working = 1 if d.get('is_working_day', 1) else 0
            check_in_start = (d.get('check_in_start') or '08:00')[:5]
            check_in_end = (d.get('check_in_end') or '09:00')[:5]
            check_out_start = (d.get('check_out_start') or '12:00')[:5]
            check_out_end = (d.get('check_out_end') or '18:00')[:5]
            shift_order = int(d.get('shift_order') or 1)
            work_hours = float(d.get('work_hours') or 8)
            is_day_off = 1 if d.get('is_day_off', 0) else 0
            created_at = datetime.utcnow().isoformat() + 'Z'

            conn.execute(
                '''INSERT INTO timesheet_details
                   (timesheet_id, shift_name, day_of_week, is_working_day,
                    check_in_start, check_in_end, check_out_start, check_out_end,
                    shift_order, work_hours, is_day_off, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (tsid, shift_name, day_of_week, is_working,
                 check_in_start, check_in_end, check_out_start, check_out_end,
                 shift_order, work_hours, is_day_off, created_at)
            )

        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/timesheets/<int:tsid>', methods=['DELETE'])
@require_auth
def delete_timesheet(tsid):
    """Xóa bảng chấm công."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('DELETE FROM timesheets WHERE id = ?', (tsid,))
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy bảng chấm công'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/timesheets/<int:tsid>/details/<int:did>', methods=['DELETE'])
@require_auth
def delete_timesheet_detail(tsid, did):
    """Xóa chi tiết một ca trong bảng chấm công."""
    conn = None
    try:
        conn = get_db()
        cur = conn.execute('DELETE FROM timesheet_details WHERE id = ? AND timesheet_id = ?', (did, tsid))
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            return jsonify({'error': 'Không tìm thấy chi tiết'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/timesheets/import', methods=['POST'])
@require_auth
def import_timesheets():
    """Import bảng chấm công từ file Excel.

    File Excel format (mỗi dòng là một ca):
      timesheet_name,shift_name,day_of_week,is_working_day,check_in_start,check_in_end,check_out_start,check_out_end,shift_order,work_hours,is_day_off
      Giờ hành chính,Ca sáng,mon tue wed thu fri,1,08:00,09:00,12:00,17:30,1,8,0
      Giờ hành chính,Ca chiều,mon tue wed thu fri,1,13:00,13:30,17:30,18:00,2,4,0
      Nghỉ cuối tuần,Nghỉ,Sat Sun,0,,,,,,,1
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Chưa chọn file'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Chưa chọn file'}), 400
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            return jsonify({'error': 'Chỉ hỗ trợ file Excel (.xlsx, .xls)'}), 400

        content = f.read()
        import io as _io
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(content))
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return jsonify({'error': 'File trống'}), 400

        header = None
        rows_data = []
        for i, row in enumerate(rows):
            parts = [str(c).strip() if c is not None else '' for c in row]

            if i == 0:
                header = [h.lower().strip() for h in parts]
                if 'timesheet_name' not in header and 'name' not in header:
                    return jsonify({'error': 'File Excel thiếu cột "timesheet_name" hoặc "name"'}), 400
                continue
            if not parts or not any(parts):
                continue
            row_dict = dict(zip(header, parts))
            rows_data.append(row_dict)

        if not rows_data:
            return jsonify({'error': 'Không có dữ liệu hợp lệ trong file'}), 400

        conn = None
        conn = get_db()
        created = 0
        updated = 0
        errors = []
        created_at = datetime.utcnow().isoformat() + 'Z'
        ts_ids = {}  # Map timesheet_name -> id

        for i, row in enumerate(rows_data, start=2):
            ts_name = (row.get('timesheet_name') or row.get('name') or '').strip()
            if not ts_name:
                errors.append(f'Dòng {i}: thiếu tên bảng chấm công')
                continue

            # Tạo hoặc lấy timesheet
            if ts_name not in ts_ids:
                existing = conn.execute('SELECT id FROM timesheets WHERE name = ?', (ts_name,)).fetchone()
                if existing:
                    ts_id = existing['id']
                else:
                    conn.execute('INSERT INTO timesheets (name, created_at) VALUES (?, ?)', (ts_name, created_at))
                    ts_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()[0]
                    created += 1
                ts_ids[ts_name] = ts_id

            ts_id = ts_ids[ts_name]
            shift_name = (row.get('shift_name') or 'Ca chính').strip()
            day_of_week = (row.get('day_of_week') or 'mon,tue,wed,thu,fri').strip().lower()
            is_working = 1 if str(row.get('is_working_day', '1')).strip() in ('1', 'yes', 'true') else 0
            check_in_start = (row.get('check_in_start') or '')[:5]
            check_in_end = (row.get('check_in_end') or '')[:5]
            check_out_start = (row.get('check_out_start') or '')[:5]
            check_out_end = (row.get('check_out_end') or '')[:5]
            shift_order = int(float(row.get('shift_order') or 1))
            work_hours = float(row.get('work_hours') or 8)
            is_day_off = 1 if str(row.get('is_day_off', '0')).strip() in ('1', 'yes', 'true') else 0

            conn.execute(
                '''INSERT INTO timesheet_details
                   (timesheet_id, shift_name, day_of_week, is_working_day,
                    check_in_start, check_in_end, check_out_start, check_out_end,
                    shift_order, work_hours, is_day_off, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (ts_id, shift_name, day_of_week, is_working,
                 check_in_start, check_in_end, check_out_start, check_out_end,
                 shift_order, work_hours, is_day_off, created_at)
            )

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'created': created,
            'updated': updated,
            'errors': errors,
            'total': len(rows_data),
            'timesheets_created': len(ts_ids),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    init_db()
    port = DEFAULT_PORT
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() in ('1', 'true', 'yes')
    app.run(host=os.environ.get('FLASK_HOST', '127.0.0.1'), port=port, debug=debug)
